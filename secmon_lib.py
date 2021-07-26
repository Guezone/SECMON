# -*- coding: utf-8 -*-
""" 
SECMON - Library for SECMON python web backend.
"""
__author__ = "Aubin Custodio"
__copyright__ = "Copyright 2021, SECMON"
__credits__ = ["Aubin Custodio","Guezone"]
__license__ = "CC BY-NC-SA 4.0"
__version__ = "2.1"
__maintainer__ = "Aubin Custodio"
__email__ = "custodio.aubin@outlook.com"
from flask import Flask, url_for, render_template, send_from_directory, request, flash, redirect, safe_join, session, url_for, session, abort
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from time import sleep
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Color, Alignment, Border, Side, colors, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension
from deep_translator import GoogleTranslator, MyMemoryTranslator
import jinja2.exceptions, sqlite3,requests, feedparser, re, smtplib, os, tweepy, base64, time, socket, random
from datetime import datetime, timedelta
from werkzeug.security import generate_password_hash, check_password_hash
from flask_sqlalchemy import sqlalchemy, SQLAlchemy
from flask_simplelogin import is_logged_in
from bs4 import BeautifulSoup
import cve_searchsploit as CS
import secmon_monitor,traceback
from multiprocessing import Process, Queue
from datetime import *
rss_feeds = [ 
    'https://cyware.com/allnews/feed',
    'https://www.cshub.com/rss/categories/attacks',
    'https://www.schneier.com/feed/atom/',
    'https://www.cshub.com/rss/categories/threat-defense',
    'https://www.cshub.com/rss/categories/malware',
    'https://www.fortiguard.com/rss/ir.xml',
    'https://nakedsecurity.sophos.com/feed/',
    'https://www.techrepublic.com/rssfeeds/topic/security/?feedType=rssfeeds',
    'https://www.cert.ssi.gouv.fr/feed/',
    'https://us-cert.cisa.gov/ncas/all.xml',
    'https://www.zataz.com/feed/',
]
messages = {
    'login_success': 'Login success.',
    'login_failure': 'Bad Username, bad password or bad request.',
    'logout': 'You have been successfully logged out.',
    'login_required': 'Please sign-in.',
    'access_denied': 'Access denied.',
    'auth_error': 'Error： {0}'
}
def sortWordsByOccurrences(word_stats):
    word_stats.sort(key = lambda x: x[1])
    return word_stats[::-1]

def getNewsTopSubject():
    top_words = []
    fr_common_words = ["de","la","le","et","les","des","en","un","du","une","que","est","pour","qui","dans","a","par","plus","pas","au","sur","ne","se","Le","ce","il","sont","La","Les","ou","avec","son","Il","aux","d'un","En","cette","d'une","ont","ses","mais","comme","on","tout","nous","sa","Mais","fait","été","aussi","leur","bien","peut","ces","y","deux","A","ans","l","encore","n'est","marché","d","Pour","donc","cours","qu'il","moins","sans","C'est","Et","si","entre","Un","Ce","faire","elle","c'est","peu","vous","Une","prix","On","dont","lui","également","Dans","effet","pays","cas","De","millions","Belgique","BEF","mois","leurs","taux","années","temps","groupe","ainsi","toujours","société","depuis","tous","soit","faut","Bruxelles","fois","quelques","sera","entreprises","F","contre","francs","je","n'a","Nous","Cette","dernier","était","Si","s'est","chez","L","monde","alors","sous","actions","autres","Au","ils","reste","trois","non","notre","doit","nouveau","milliards","avant","exemple","compte","belge","premier","s","nouvelle","Elle","l'on","terme","avait","produits","cela","d'autres","fin","niveau","bénéfice","toute","travail","partie","trop","hausse","secteur","part","beaucoup","Je","valeur","croissance","rapport","USD","aujourd'hui","année","base","Bourse","lors","vers","souvent","vie","l'entreprise","autre","peuvent","bon","surtout","toutes","nombre","fonds","point","grande","jour","va","avoir","nos","quelque","place","grand","personnes","plusieurs","certains","d'affaires","permet","politique","cet","chaque","chiffre","pourrait","devrait","produit","l'année","Par","rien","mieux","celui","qualité","France","Ils","Ces","s'agit","vente","jamais","production","action","baisse","Avec","résultats","Des","votre","risque","début","banque","an","voir","avons","qu'un","qu'","elles","moment","qu'on","question","pouvoir","titre","doute","long","petit","d'ailleurs","notamment","FB","droit","qu'elle","heures","cependant","service","Etats-Unis","qu'ils","l'action","jours","celle","demande","belges","ceux","services","bonne","seront","économique","raison","car","situation","Depuis","entreprise","me","nouvelles","n'y","possible","toutefois","tant","nouveaux","selon","parce","dit","seul","qu'une","sociétés","vient","jusqu'","quatre","marchés","mise","seulement","Van","semble","clients","Tout","Cela","serait","fort","frais","lieu","gestion","font","quand","capital","gouvernement","projet","grands","réseau","l'autre","données","prendre","plan","points","outre","pourtant","Ainsi","ni","type","Europe","pendant","Comme","mesure","actuellement","public","dire","important","mis","partir","parfois","nom","n'ont","veut","présent","passé","forme","autant","développement","mettre","grandes","vue","investisseurs","D","trouve","maison","mal","l'an","moyen","choix","doivent","NLG","direction","Sur","simple","période","enfants","dollars","personnel","assez","programme","général","banques","eux","semaine","président","personne","européenne","moyenne","tard","loi","petite","certaines","savoir","loin","explique","plupart","jeunes","cinq","contrat","Banque","valeurs","seule","rendement","nombreux","fonction","offre","client","activités","eu","environ","ministre","cadre","sens","étaient","sécurité","recherche","Paris","sorte","décembre","Son","suite","davantage","ensuite","janvier","donne","vrai","cause","d'abord","conditions","suis","juin","peine","certain","septembre","sommes","famille","l'indice","pris","laquelle","directeur","qu'en","propose","gens","derniers","étant","fut","chose","portefeuille","obligations","afin","différents","technique","Aujourd'hui","ailleurs","P","l'ensemble","américain","ventes","Selon","rue","livre","octobre","vraiment","sein","Or","dollar","Enfin","haut","Plus","petits","porte","tel","durée","domaine","aurait","jeune","présente","passe","PC","lorsque","choses","puis","Vous","aucun","l'un","n'en","tandis","coup","existe","propre","carte","crise","importante","atteint","revenus","montant","forte","ici","s'il","Quant","vu","rapidement","j'ai","ville","etc","mars","s'en","mon","premiers","bas","marque","véritable","ligne","longtemps","propres","devant","passer","départ","pu","total","série","quoi","particulier","concurrence","élevé","position","connu","principe","tendance","court","n","pages","évidemment","résultat","aura","parmi","Sans","américaine","face","trouver","durant","femmes","construction","désormais","distribution","telle","difficile","autour","européen","pratique","centre","vendre","juillet","mai","région","sociale","filiale","film","h","besoin","mode","Pas","représente","réalité","femme","vaut","Tél","aucune","hommes","donner","titres","l'Europe","nombreuses","différentes","moyens","formation","chiffres","Générale","dix","prochain","l'Etat","genre","bureau","communication","participation","gros","pourquoi","estime","devient","réalisé","création","novembre","l'évolution","pourra","semaines","consommation","faible","terrain","site","droits","moitié","puisque","Du","reprise","compris","projets","avril","vont","call","donné","simplement","six","firme","perte","Bien","Philippe","sait","prend","vite","via","stratégie","vos","jeu","J'","petites","marketing","presque","Michel","manque","réaliser","financiers","Car","Comment","voiture","chef","constitue","Internet","J'ai","enfin","net","charge","nature","second","payer","actuel","Elles","investissements","dispose","financier","d'achat","membres","date","avaient","gamme","revanche","comment","décision","l'avenir","tour","actionnaires","s'y","solution","créer","l'économie","concerne","l'époque","belle","lequel","tél","seconde","version","Pays-Bas","cher","chacun","lire","techniques","décidé","mouvement","conseil","nécessaire","meilleur","double","sujet","généralement","restent","celles","politiques","malgré","confiance","homme","d'actions","Certains","ayant","papier","commerce","Région","Wallonie","Windows","termes","met","contraire","informations","l'industrie","trimestre","E","différence","certaine","formule","jusqu'au","voit","programmes","actuelle","permis","dossier","Quand","l'heure","guerre","acheter","rendre","février","ma","l'emploi","main","voire","bons","technologie","européens","Sa","éléments","unique","l'eau","venir","générale","courant","suffit","l'ordre","conserver","maximum","force","fax","Que","largement","milliard","soient","Pierre","devenir","l'Union","franc","minimum","mort","responsable","possibilité","presse","affaires","longue","travers","M","BBL","relativement","moi","Deux","présence","européennes","devraient","groupes","ensemble","santé","New","pense","bénéfices","but","compagnie","publique","coeur","revenu","mesures","table","nettement","questions","d'avoir","permettre","l'homme","Chez","retour","qu'elles","C","majorité","potentiel","moindre","récemment","secteurs","réduction","large","traitement","perdu","étrangers","parents","l'une","fond","capacité","vitesse","activité","l'exercice","l'objet","quel","tient","taille","éviter","risques","Jean","Pourtant","Allemagne","parler","propos","quant","signifie","voie","jouer","prévoit","blanc","noir","parti","logiciel","continue","Notre","bois","meilleure","l'argent","perspectives","développer","celui-ci","oeuvre","structure","suivre","tiers","prise","professionnels","raisons","néanmoins","preuve","social","bénéficiaire","couleurs","mondial","Cet","maintenant","essentiellement","prévu","Japon","prévisions","centrale","Alors","international","yeux","PME","l'a","ait","bonnes","opérations","pied","l'art","pourraient","Londres","juge","devra","uniquement","corps","divers","Parmi","numéro","réduire","Tous","texte","tenu","budget","l'étranger","pression","mes","n'était","style","économiques","Jacques","montre","population","analystes","S","processus","placement","classique","dividende","rester","publics","fortement","plein","wallonne","DEM","Express","faudra","travailler","Crédit","directement","prime","Flandre","crédit","monnaie","précise","appel","Autre","travaux","l'occasion","juste","Chaque","put","tableau","terre","permettent","devenu","rouge","mémoire","partenaires","rapide","travailleurs","joue","objectif","salle","parle","musique","milieu","d'entreprise","autorités","chute","régime","d'autant","liste","opération","bout","performances","électronique","haute","responsables","lancé","voitures","patron","Malgré","affiche","situe","B","l'image","études","Microsoft","condition","retrouve","Aux","revient","Belgacom","route","Ensuite","Luxembourg","campagne","comptes","hors","culture","Commission","d'entre","possibilités","semestre","actifs","finalement","internationale","l'achat","monétaire","passage","of","justice","page","tels","poids","celle-ci","commercial","entendu","l'investisseur","mondiale","accord","diverses","totalement","fil","clair","vin","biens","euro","York","parfaitement","viennent","division","réseaux","principal","lancer","supérieur","atteindre","référence","téléphone","management","vins","proche","collection","fiscale","Ceci","informatique","investissement","volume","matériel","publicité","train","coupon","progression","tenir","protection","l'aide","couleur","nouvel","Lorsque","change","changement","garantie","somme","Belge","plaisir","fils","laisse","importants","privé","ses","besoins","oeuvres","américains","relations","peau","moteur","augmentation","suivi","volonté","beau","bancaire","laisser","bureaux","principalement","intéressant","logiciels","sommet","l'activité","d'en","vivre","élevés","Robert","contrats","oublier","performance","réponse","d'exploitation","concept","obtenir","poste","attendre","lignes","consiste","augmenté","vert","Ou","figure","mot","développé","l'histoire","magasins","collaboration","répondre","TVA","holding","G","livres","convient","fonctions","fera","pouvait","million","Paul","britannique","d'entreprises","voix","Grande-Bretagne","disque","affaire","minutes","quelle","contexte","limite","mains","commun","réduit","Pourquoi","particuliers","verre","wallon","d'Etat","allemand","effets","Chine","meilleurs","rend","applications","d'ici","procédure","l'opération","devait","profit","méthode","pose","commence","idée","l'Internet","d'eau","créé","nuit","Nord","capitaux","options","consommateur","cartes","soi","métier","probablement","aller","d'investissement","facile","International","importantes","Marc","capitale","devise","prochaine","transport","Street","demander","utilisateurs","l'affaire","image","l'idée","propriétaire","facilement","publiques","croire","disponible","Louis","d'or","veulent","Charleroi","Ne","consommateurs","devises","difficultés","sort","national","machines","annoncé","choisi","découvrir","soutien","avez","perdre","cuisine","telles","D'autres","travaille","R","ouvert","phase","certainement","télévision","pratiquement","annuel","bord","paiement","Bank","institutions","seuls","arrive","constate","marques","nationale","regard","représentent","Belges","état","Qui","libre","rachat","Toutefois","portes","sortir","commandes","permettant","manager","fiscal","cinéma","histoire","zone","sauf","avantages","l'information","voici","dur","effectivement","puisse","réel","The","puissance","fixe","Belgium","contact","époque","rythme","principaux","vendu","utilisé","étude","Leur","sensible","Bref","rencontre","L'entreprise","spécialistes","brut","mauvais","néerlandais","supplémentaire","mots","reprises","nécessaires","Non","soir","Prix","machine","penser","CD","parts","comprend","fusion","acquis","totale","voyage","logique","l'échéance","concurrents","idées","trouvé","dette","Sud","réellement","financement","disponibles","vieux","lance","marge","dirigeants","avis","changer","conséquence","sociales","supérieure","Certes","faisant","ordinateur","partenaire","warrant","fabrication","redressement","suffisamment","délégué","pourront","poursuit","chemin","emplois","l'environnement","réalise","FRF","évolution","Cour","automobile","Premier","ancien","note","parties","pension","professionnel","assure","garder","Rien","Actuellement","DE","S'il","l'administration","Guy","est-il","IBM","climat","d'acheter","SICAV","département","sept","partout","immobilier","lancement","rating","réussi","patrimoine","feu","expérience","Anvers","anciens","graphique","Fortis","faveur","retrouver","droite","responsabilité","commande","Kredietbank","d'argent","direct","l'inflation","n'avait","utiliser","tonnes","l'origine","connaissance","acheté","Ici","américaines","clairement","semblent","biais","futur","neuf","chance","faillite","km","équipe","musée","compagnies","documents","pertes","sortie","m'a","seraient","d'autre","choisir","l'instant","tellement","industriel","précompte","d'Europe","immédiatement","avantage","qu'au","constituent","déchets","sport","van","demeure","garde","maisons","Solvay","conséquences","KB","l'offre","active","dépenses","donnent","employés","sites","élections","détient","n'importe","obligation","fruits","véhicule","l'égard","Conseil","investi","mission","profiter","visite","comprendre","professionnelle","affirme","l'intérieur","Wall","charges","privée","rares","succession","liberté","rentabilité","suivant","efficace","assurer","images","agences","impossible","m","John","enfant","fournisseurs","photo","salaires","Avant","compter","l'Est","disposition","formes","bénéficiaires","lesquels","maintenir","précisément","couple","enregistré","recul","offrir","peur","hauteur","centres","voulu","industrielle","positif","Luc","administrateur","intéressante","commerciale","interne","pleine","passant","vision","GSM","faits","retard","certes","l'air","lundi","Outre","porter","écrit","cesse","locaux","délai","trouvent","classiques","commencé","réalisée","Alain","vigueur","gagner","Celui-ci","Philips","ceux-ci","favorable","pouvoirs","participations","annonce","génération","élément","devenue","touche","conseils","devoir","mer","souligne","respectivement","rapports","vacances","lieux","naturellement","d'y","lorsqu'il","statut","USA","ceci","destiné","défaut","objectifs","récente","saison","d'art","industriels","Suisse","catégorie","complexe","huit","l'obligation","fisc","obtenu","repris","occupe","sérieux","émis","Quelques","comportement","limité","vingt","conjoncture","gauche","marche","d'origine","l'utilisateur","ordre","mobilier","parcours","perspective","normes","recours","l'esprit","Communauté","annuelle","T","lecteur","objets","fabricant","niveaux","Entre","réalisation","amateurs","conséquent","présenter","Celle-ci","vise","types","détail","mauvaise","professeur","progressé","signe","passée","approche","p","Reste","return","jardin","l'espace","flamand","Namur","bilan","Vif","sensiblement","Trois","utilise","commune","dimanche","option","partis","analyse","films","surface","warrants","GBP","prises","secret","historique","journée","l'ancien","Pendant","allemande","d'assurance","André","fille","l'importance","proposer","avions","x","augmenter","parc","Delhaize","the","Lors","limitée","appareils","villes","au-dessus","diminution","prochaines","servir","Bernard","commission","faiblesse","plus-value","souhaite","internationales","producteur","producteurs","code","belles","cabinet","fonctionnement","FF","gérer","cm","mouvements","pratiques","régions","dossiers","meilleures","Parce","entrée","vendredi","actif","sociaux","supplémentaires","café","message","physique","Société","communes","dizaine","faute","sélection","source","facteurs","milliers","soleil","tirer","concernant","Bourses","fallait","sentiment","bénéficier","débat","l'Allemagne","élevée","ouvrage","police","pouvez","attention","a-t-il","bel","constructeurs","contribuable","moderne","passion","primes","in","suit","auquel","dépasse","spécialisée","bruxellois","déclaration","multiples","quartier","vidéo","dépend","l'école","liquidités","correction","CA","comité","Web","cherche","filiales","Sous","signé","leader","calcul","gaz","D'abord","Rens","artistes","déficit","cadres","fédéral","probable","remboursement","and","efforts","restaurant","Toutes","couverture","domicile","soins","devront","luxe","complet","danger","indispensable","syndicats","comporte","faite","juridique","langue","rendez-vous","d'informations","demandé","respect","continuer","l'organisation","lesquelles","local","l'impression","n'existe","rare","restructuration","automatiquement","plat","boursier","sol","c'était","cotées","décide","L'action","Cependant","Certaines","matériaux","ordinateurs","tradition","V","progressivement","capable","classe","familiale","réserve","fonctionne","solutions","LA","fabricants","paie","Finances","l'été","réelle","changé","masse","unités","considéré","fer","auront","noms","riche","Patrick","proposé","salon","territoire","fixé","magasin","candidats","marges","asiatique","inférieur","réaction","fleurs","l'effet","record","tribunal","recettes","poursuivre","dessous","portant","Aussi","Sabena","acteurs","dehors","constructeur","l'auteur","relation","offrent","spectaculaire","LUF","produire","confort","familles","investir","reprend","sert","montrer","mérite","places","Soit","judiciaire","textes","quasi","SNCB","jeux","permettra","étudiants","membre","photos","positions","sud","Cockerill","lendemain","cent","gagné","japonais","l'absence","mark","pointe","solide","Voici","anglais","n'ai","présentent","décisions","législation","médias","victimes","écran","nécessairement","découverte","l'assuré","club","environnement","noter","crée","exportations","négociations","Jan","répond","BEL","entier","business","peinture","s'était","voisins","faibles","location","nord","promotion","technologies","auraient","caisse","entend","simples","maladie","menu","chances","commerciaux","printemps","Benelux","poser","Asie","l'utilisation","usage","PIB","actionnaire","prennent","résistance","Dow","II","surprise","Etats","mariage","nécessité","Puis","cote","Plusieurs","beauté","exclusivement","lettre","payé","rendu","s'ils","software","utile","gestionnaires","bénéficie","procédé","vaste","crois","normal","Centre","construire","démarche","emprunts","naissance","D'autant","Co","d'information","distance","tourner","Club","attendant","quantité","roi","l'assureur","tourne","ajoute","bancaires","ajouter","géant","automatique","faux","attend","litres","présenté","argent","confirme","indépendants","l'ordinateur","énorme","destinés","l'avantage","véhicules","ressources","standard","auparavant","construit","Quelle","principales","quelqu'un","disposer","global","écoles","Quel","réputation","fameux","rappelle","conseille","heure","veille","difficulté","l'état","limites","commerciales","samedi","palais","vend","vit","Tractebel","connaissent","reprendre","village","emploi","amis","budgétaire","croit","mises","souci","contient","habitants","Weekend","bras","beaux","bruxelloise","faisait","introduit","intérieur","outils","précis","chercheurs","taxe","salaire","transactions","Christian","chambre","portée","réflexion","AG","C'était","d'emploi","hasard","matin","assureurs","réforme","Beaucoup","fournir","recherches","liés","tenue","proposent","aide","ferme","l'enfant","l'or","secondes","CGER","contenu","quotidien","flamande","centaines","course","billet","critique","l'arrivée","naturel","principale","support","week-end","Dehaene","Gand","chargé","économies","Nos","augmente","guide","proposition","laissé","spécialiste","francophones","importance","vent","conception","préférence","spectacle","avenir","d'entrée","grave","commencer","d'années","diminuer","chercher","bonheur","dizaines","LE","d'environ","exactement","outil","scénario","Jones","coups","émissions","éventuellement","Royale","l'agence","soumis","d'exercice","lecture","monter","Grand","central","exigences","assuré","contacts","consacré","l'attention","d'administration","due","faut-il","réussite","échéance","recevoir","tableaux","arriver","évident","PS","art","Italie","amélioration","auteurs","estimé","quinze","Russie","demain","précédent","vendeur","événements","autrement","experts","fortes","furent","possibles","circonstances","placer","publication","l'écran","réserves","sauce","venu","Charles","collaborateurs","implique","l'assurance","obligataire","établi","CD-Rom","H","forcément","l'essentiel","l'enseignement","remarquable","vol","Claude","tourisme","internationaux","directe","compétences","conseiller","facteur","l'est","plastique","rarement","Royal","affiché","lutte","relative","actuels","envie","l'équipe","ministres","secrétaire","capitalisation","langage","positive","circulation","convaincre","notion","visage","vouloir","ajoutée","caractéristiques","Eric","Union","paix","puisqu'il","courrier","disposent","développe","présentation","barre","comparaison","déterminer","firmes","fournisseur","informatiques","luxembourgeois","achats","solde","Serge","globale","propriété","stratégique","Renault","partage","porté","sources","Kong","cour","destinée","N","absolument","branche","l'objectif","ouvre","plans","productivité","Résultat","améliorer","d'obtenir","joué","Parlement","dépit","fichiers","personnalité","constitué","gestionnaire","né","profession","qualités","conscience","médecin","celles-ci","design","décor","faudrait","participer","appelle","forces","suisse","appareil","conduite","D'une","longueur","tarifs","vérité","lien","locales","francophone","clubs","correspond","coupons","d'émission","estiment","défi","protéger","réalisés","d'emplois","d'éviter","l'ouverture","méthodes","revenir","superbe","volontiers","document","nommé","tente","financer","scientifique","Georges","travaillent","l'investissement","lié","zones","aime","lettres","ouverte","Hong","L'année","murs","philosophie","rappeler","utilisés","suivante","d'année","représentant","traduit","remettre","situé","différente","longs","économie","discours","distributeur","domaines","l'introduction","régional","faites","italien","restera","usine","Group","l'informatique","personnage","portent","attendu","l'option","Jean-Pierre","articles","changements","fallu","léger","mener","propriétaires","spécifique","récupérer","voyages","procéder","locale","médecins","privés","transmission","concurrent","courte","quart","baisser","pieds","publié","Ford","menace","réunion","transfert","composé","dimension","personnages","ralentissement","conclusion","l'usage","agents","parfum","rémunération","difficiles","l'entrée","mettent","pierre","proches","réglementation","salles","grimpé","prochains","prévue","électrique","dynamique","exposition","installé","plancher","distributeurs","déclare","connue","n'avons","préparation","réalisées","beurre","opérateurs","achat","province","spécifiques","Albert","l'usine","l'existence","renforcer","téléphonique","comptable","effectuer","trafic","degré","l'ont","définitivement","humain","optique","remarque","talent","appelé","modifier","définition","peintre","respecter","stade","statistiques","certificats","s'attend","limiter","livraison","placements","raconte","volumes","immobiliers","Fax","anciennes","chevaux","médicaments","Peter","feuilles","football","identique","pouvons","remise","structures","tenter","accords","cotisations","indice","neutre","Mon","constituer","d'accord","montrent","placé","loyer","proximité","voient","épouse","Canada","entrer","postes","précision","cité","concours","patrons","populaire","pétrole","négatif","allemands","d'activité","roman","victime","italienne","ménages","repas","PetroFina","langues","tendances","D'autre","pire","prudence","savent","Néanmoins","conduit","mille","rénovation","égard","Américains","exercice","l'étude","s'impose","avance","effectué","fortune","fournit","lecteurs","Morgan","découvert","l'inverse","différent","emploie","bleu","royal","technologique","télécommunications","Amsterdam","fiscales","indique","information","lourd","signal","Ed","Mieux","aider","ancienne","apporte","nette","prestations","publicitaires","sensibles","communauté","l'émission","lit","volatilité","étape","assurance","jusqu'en","lancée","résoudre","garanti","modification","revue","spéciale","www","chacune","l'analyse","différences","messages","priorité","recommandation","récent","charme","dividendes","Olivier","passent","finale","immeubles","logement","pourcentage","rire","stabilité","difficilement","défense","l'ancienne","magazine","D'un","Y","eaux","jeunesse","l'intention","continuent","révolution","étonnant","organisation","constater","dos","emprunt","oui","éditions","Daniel","sel","utilisée","compartiment","publicitaire","EN","article","bande","capacités","centrales","considérée","milieux","occasion","quasiment","pouvant","Vermeulen-Raemdonck","visiteurs","chambres","considérablement","demi","découvre","essentiel","broker","dettes","mardi","reconnaissance","salariés","formules","grosse","heureux","perd","radio","allait","multimédia","partiellement","seules","Gérard","Oui","Securities","toucher","jugement","l'oeuvre","considérer","remplacer","couvrir","précieux","segment","dessins","espace","indices","refuse","chefs","exemples","rejoint","spécialisé","l'amour","l'exportation","objet","précédente","rose","versions","d'études","destination","Encore","deviennent","ET","l'Italie","personnelle","plats","vingtaine","l'expérience","virus","Faut-il","chasse","longues","Toute","bases","cotée","final","monnaies","travaillé","apporter","aspects","disparu","David","Management","port","racheter","relever","Celui","ING","catalogue","centaine","chaleur","profil","représentants","SA","conclu","réside","scientifiques","Chambre","secondaire","Fin","serveur","XIXe","exige","grimper","immeuble","l'Université","montants","paysage","vendus","ton","assurances","catégories","dure","décote","soutenir","édition","dangereux","agréable","voulait","combien","d'application","disparition","optimiste","plus-values","tomber","erreur","l'augmentation","situations","spécialisés","subi","suivent","Jusqu'au","classement","l'exemple","norme","rentable","sang","socialiste","tombe","Justice","attitude","mines","qu'aux","liée","plantes","vague","General","l'immobilier","légumes","Ceux-ci","conflit","excellent","licence","travailleur","appris","est-elle","gagne","mari","préparer","purement","située","vérifier","Jean-Luc","gain","métal","surfaces","L'objectif","d'épargne","douze","expliquer","lorsqu'on","meubles","yen","chaussures","créée","institution","l'accent","solidarité","Maastricht","basée","journal","soin","sourire","Guerre","bouteilles","flexibilité","maintient","appartient","moments","rouges","L'an","basé","devons","installations","Bacob","association","d'obligations","format","City","Page","disques","modem","mélange","ordinaire","vide","chimique","disent","pharmaceutique","d'assurances","numérique","porteur","répartition","blanche","composants","future","parvient","évoque","Durant","calme","cru","Electrabel","culturel","grosses","baissé","lois","moteurs","principes","trente","éventuelle","Peu","prévoir","tours","Pentium","acheteur","dimensions","fonctionnaires","organisé","rencontré","russe","savoir-faire","établissements","Fédération","Toujours","créativité","top","application","dépasser","importe","jaune","l'application","marqué","mécanique","socialistes","tranche","Quelles","envisage","traiter","Surtout","acheteurs","chinois","claire","l'Institut","vécu","Objectif","bail","demandes","diversification","montré","renseignements","souscription","Tokyo","entendre","tests","Siemens","filles","unité","Bekaert","Dr","UCB","composition","resté","sinon","agence","fini","modifications","Cash","industrielles","obtient","permanence","restaurants","réels","échange","florins","l'accord","terrains","émergents","atouts","offrant","LES","bouche","champ","chaud","l'annonce","monte","preneur","présents","quitte","tarif","facture","fiscaux","modeste","processeur","Fund","avenue","compétition","relevé","tenté","Est-ce","Musée","W","bijoux","différentiel","déclaré","institutionnels","l'employeur","traité","Intel","traditionnels","victoire","connus","correctement","pub","Dominique","Tant","accessible","rencontrer","stocks","Art","espérer","jouent","menée","nécessite","provenant","utilisent","affichent","délais","inférieure","sent","spécial","Amérique","acquérir","album","idéal","l'écart","véritables","associé","candidat","connaissances","l'énergie","signes","cheveux","conserve","stress","d'Anvers","d'action","directeurs","donnée","endroit","l'emprunt","l'impact","der","traditionnelle","Martin","ciel","convention","obligataires","prouver","Espagne","O","Petit","Source","dessin","humaine","l'huile","lait","Seule","Thierry","boursiers","continent","destinées","flamands","néerlandaise","pensions","commencent","considérable","nationales","nul","s'adresse","conjoint","crédits","militaire","morceaux","privatisation","repose","sommeil","traditionnel","PSC","Seul","capables","combat","finances","puissant","s'agissait","Bill","Renseignements","physiques","Richard","allant","créations","toile","évidence","convaincu","excellente","or","retraite","théorie","transformer","Tour","transaction","visant","Deutsche","Mons","attentes","cycle","détails","Votre","héros","l'artiste","l'université","sérieusement","uns","Ceux","considération","impose","propositions","Autrement","cap","forts","l'Afrique","usines","Afin","Quels","aisément","ressemble","risquent","totalité","imaginer","originale","intégré","intéressantes","l'extérieur","loyers","auxquels","circuit","indépendant","intérieure","jus","maintien","cotisation","l'Asie","moyennes","quitter","stable","CVP","Compaq","galerie","liens","souffle","GIB","apprendre","concert","l'exception","l'échelle","liquide","nez","noire","température","transparence","école","champion","diminué","désir","ressort","voulons","équipé","alimentaire","den","organisations","présidence","raisonnable","ratio","recommande","utilisant","accepter","accepté","cache","chocolat","chuté","comparer","courts","figurent","passagers","prison","viande","associés","esprit","froid","jeudi","liées","revu","satisfaction","satisfaire","test","tiennent","vraie","contrairement","dépassé","extérieur","qu'avec","ami","American","Etat","complémentaire","déclarations","réactions","Fonds","artiste","conclure","déduction","remis","L'indice","déterminée","fiscalité","grand-chose","humaines","réponses","équipes","ITL","Michael","Systems","aspect","commercialisation","manger","RTBF","engagé","obligé","proportion","signature","étranger","imposé","s'applique","silence","vote","Afrique","Mobistar","cible","contemporain","fondateur","Jean-Claude","communiquer","d'investir","existent","majeure","ouvrir","électroniques","JPY","TGV","compétitivité","erreurs","notation","rang","Apple","GB","accident","certificat","exceptionnel","http","proprement","riches","Barco","Quoi","violence","adapté","bénéficient","récession","sentir","armes","arrivé","crainte","garanties","l'automne","ménage","officiellement","ouvriers","Autant","discussion","rejoindre","époux","citoyens","concernés","d'inflation","définir","L'idée","Paribas","Telecom","d'aller","fabrique","feront","née","oblige","patients","pensent","responsabilités","SP","doublé","fraude","l'article","organise","Henri","conclut","désire","l'appareil","l'association","l'installation","législateur","écrans","choc","gratuit","mobile","naturelle","dialogue","révision","familial","lourde","poche","décider","négociation","tort","Maison","Trésor","constante","cotation","déterminé","l'instar","managers","opté","transformation","Life","anniversaire","compétence","géographique","I","mandat","réservé","établir","Business","fins","richesse","CAD","commente","intermédiaire","l'univers","retrouvé","sciences","Sun","banquier","former","monté","parfait","veux","René","investit","l'oeil","n'aurait","parvenir","vieille","Di","collections","dirige","fonctionner","mauvaises","tapis","venus","Contrairement","Suez","piste","pistes","tensions","campagnes","investis","proposés","sac","tabac","bataille","britanniques","fine","liégeois","partenariat","privées","remplir","supérieurs","Beaux-Arts","Christie's","laser","restauration","Dutroux","chimie","rendent","textile","Brabant","Colruyt","James","National","Quatre","préalable","souvenir","venue","Communal","avocat","comparable","consolidé","critiques","interdit","l'initiative","mine","quotidienne","rigueur","réduite","tissu","Invest","pain","participants","procédures","profondeur","retrouvent","rues","taxation","Mexique","asiatiques","conducteur","demandent","environs","fermeture","gris","rumeurs","accueille","amoureux","d'augmenter","défendre","l'immeuble","pure","souffre","créneau","d'énergie","journaux","s'explique","seuil","Jeux","Office","auteur","cash-flow","fichier","foi","instruments","quelles","séance","véritablement","Yves","attirer","civil","civile","d'aujourd'hui","eau","l'épargne","station","courbe","hectares","influence","ingénieurs","tables","vivent","Exemple","L'un","blancs","couche","cuir","devenus","extraordinaire","patient","peux","aient","animaux","associations","d'utiliser","foie","initiative","l'Amérique","poursuite","survie","Face","K","apparemment","consultant","expansion","l'exposition","séjour","champagne","commentaires","complexes","cylindres","décennie","rendements","retenu","sais","sujets","cuivre","offert","réagir","sec","varie","Fondation","artistique","communications","monétaires","métaux","permanente","positifs","électriques","Ph","basse","concentration","investisseur","provoqué","doux","stations","coin","modifié","avocats","estimations","original","souplesse","Attention","Frank","Hainaut","Suite","annuels","cellule","clause","exemplaires","malheureusement","minute","normale","Frédéric","NT","Sud-Est","atout","latine","logements","pilotes","susceptibles","Roger","XVIIIe","ordres","remarquer","actuelles","bouteille","constat","opportunités","prépare","vendeurs","accrue","fruit","g","jugé","l'amélioration","loisirs","pur","trentaine","bus","gendarmerie","air","alimentaires","coté","modernes","préciser","réussir","laissent","parfaite","spécialement","évoluer","Dewaay","Désormais","Groupe","maladies","négligeable","tension","Lion","chansons","dite","festival","négative","préféré","restant","Cera","adopté","coopération","distingue","douceur","retirer","technologiques","Editions","Parfois","bruit","comptant","démocratie","exception","mercredi","offres","sucre","vedette","évolue","British","Leurs","compromis","hauts","élevées","émission","Faire","attendue","d'appel","jusqu'ici","lourds","quels","soirée","événement","alternative","chimiques","conférence","quitté","serveurs","Brésil","CD-ROM","correspondant","l'avis","locataire","matériau","périodes","utilisées","ai","d'emblée","l'aspect","morale","équilibre","Sony","fixer","gratuitement","trait","Trop","adultes","consacrer","d'importance","normalement","parole","prochainement","suscite","verra","clé","mesurer","notes","potentiels","relatives","Flamands","Francfort","L'homme","Palais","Plan","République","l'armée","transports","Portugal","couvert","joueurs","Malheureusement","coupe","dispositions","effort","endroits","aides","contribution","insiste","s'inscrit","souhaitent","communal","impact","progresser","Sambre","US","administrateurs","d'ordre","deviendra","dégager","formations","l'ouvrage","souscrire","cellules","facilité","gras","militaires","passés","quinzaine","souvient","Se","automobiles","bref","confortable","essentielle","officiel","vive","vols","Marcel","Top","combinaison","distinction","définitive","japonaise","liaison","tissus","cadeau","canadien","distribué","existants","ordinaires","servi","surveillance","l'architecture","l'aéroport","médecine","n'aura","n'étaient","revoir","récentes","voies","L'obligation","Rappelons","comptabilité","fabriquer","fasse","intéressants","peintures","quartiers","valable","étapes","bénéficié","couvre","diminue","envers","introduire","missions","s'attendre","Petrofina","apparition","coffre","digne","fibres","initiatives","littérature","rembourser","retrait","Bundesbank","D'ailleurs","Ma","Pascal","Pologne","consacre","employeur","favorables","l'approche","manquent","assurée","battre","chantier","conclusions","consulter","craindre","d'utilisation","vivant","Chacun","internes","apprend","liégeoise","observe","provenance","sortes","Marie","cessé","céder","estimée","marchandises","Poste","balance","copie","cuisson","négocier","spéciaux","traite","Bruges","hollandais","peut-on","porteurs","régler","soutenue","suivie","Stanley","accueillir","médical","notoriété","provoquer","sensibilité","su","vocation","L'investisseur","for","impression","l'ampleur","séduit","conflits","imposable","journalistes","manifeste","provoque","wallons","éditeurs","EUR","canal","fondamentale","futurs","graves","mené","mur","pommes","racheté","remonte","solides","suffisante","chargée","chers","discussions","garantit","indicateurs","provient","soutenu","sportif","systématiquement","zéro","comptent","recette","récit","subir","évolué","Johan","accorde","faciliter","hausses","Macintosh","Services","d'imposition","débuts","garantir","portefeuilles","susceptible","universités","Glaverbel","Sotheby's","actes","brasserie","caractéristique","cherchent","cp","favoriser","justement","prudent","stock","échelle","énormément","Standard","compose","couronne","exceptionnelle","flux","j'étais","justifier","réfugiés","t","téléphoniques","Monsieur","Ville","accepte","inspiré","l'ombre","pollution","situent","allemandes","boissons","douce","gouvernements","intervention","motifs","primaire","World","entrepreneurs","l'efficacité","représentation","Thomas","apparaissent","complémentaires","cycliques","franchement","instrument","rayon","Food","Roi","conversion","partager","retenue","simplicité","Comité","confirmé","devaient","expériences","front","jeter","logistique","reconnu","Affaires","Heureusement","comédie","historiques","imposer","l'actionnaire","obligatoire","recourir","références","traces","témoigne","GBL","Java","Vu","acte","appliquer","catastrophe","conduire","contribué","fais","intervenir","mettant","pilote","plafond","remplacement","tire","Berlin","Vincent","portable","profonde","refusé","repos","béton","fermé","juges","parlementaires","prévention","Donc","d'électricité","dispositif","forment","neige","suffisant","Louvain","TV","diffusion","fédération","lentement","prenant","souris","tu","contenter","douleur","intervient","j'avais","look","manoeuvre","parquet","poussé","arguments","billets","consacrée","dirigeant","décoration","holdings","justifie","levier","majeur","midi","recyclage","robe","Entre-temps","appels","directive","initial","intéressés","pousser","pouvaient","secrets","surpris","univers","d'avis","poisson","spécialisées","séduire","verser","d'investissements","générations","nettoyage","ouverts","réductions","vélo","Anne","Compagnie","Souvent","d'Amsterdam","explique-t-il","l'abri","l'intégration","officielle","résolution","Service","courses","l'exploitation","pari","pousse","revendre","trace","abonnés","craint","croissant","juger","régionale","symbole","touristes","Rome","actives","communautaire","contraintes","journaliste","traditionnelles","variable","amour","atelier","budgets","budgétaires","clef","d'ores","détriment","nationaux","paquet","relatif","Francis","Rupo","d'enfants","diesel","gare","l'acquisition","parlent","rapporte","regarder","éventuel","Clabecq","carrés","psychologique","rupture","téléphonie","Air","Danemark","Sauf","citoyen","four","permettrait","puissent","rapides","Marketing","Tendances","dit-il","développements","enregistre","envoyé","intermédiaires","l'issue","liquidité","réagi","Allemands","L'autre","Louise","connues","consolidation","créateur","idéale","l'espoir","profité","prévus","résulte","similaire","Boeing","Didier","Dieu","MB","Willy","agir","coins","constaté","d'eux","danse","occidentale","optimistes","pensée","professionnelles","Computer","San","Tournai","appliquée","chanson","déroule","franchir","liquidation","morts","nouveauté","prestigieux","suppression","Laurent","Mercedes","existantes","pleinement","simultanément","établissement","cercle","corruption","discipline","familiales","l'avant","laboratoire","livrer","montée","participe","Personne","adresse","finance","génie","leasing","versement","bits","concernées","dents","inclus","maximale","précédemment","routes","variations","équipements","Declerck","chemins","constituée","d'effectuer","globalement","libres","proposant","souligner","Bon","ambitions","croissante","décennies","fou","l'influence","littéralement","motivation","rubrique","souvenirs","surprises","vendue","Celles-ci","bébé","plainte","stockage","écrire","énergie","Spector","annonceurs","d'olive","débats","ferait","grain","sont-ils","séparation","tournant","vendues","Compte","Cools","Volvo","accessoires","constitution","consultants","dommages","occupé","s'appelle","échanges","Seconde","adresses","efficacité","fixée","frappe","l'apparition","monopole","panneaux","restée","sentiments","terminé","utiles","Bruno","Seuls","appliqué","donnant","fondamentaux","fréquemment","l'aventure","métiers","planche","royale","suppose","Inc","Moins","fourni","japonaises","mm","payés","profond","programmation","résolument","L'Europe","d'amour","d'ouvrir","golf","poudre","","","proposées","étoiles","PRL","attaché","concevoir","dommage","l'opinion","main-d'oeuvre","récents","stratégiques","vitesses","Peugeot","Philip","apprécié","connexion","hommage","jardins","remonter","supplément","Canal","Tessenderlo","cheval","entretien","inutile","l'Espagne","laissant","mécanisme","nouveautés","placés","repli","régionales","régionaux","souple","symbolique","troubles","évaluer","Aucun","Mac","Régions","cession","confie","moyennant","numéros","portrait","établie","cinquantaine","d'assurer","peuple","promis","retenir","réception","sexe","utilisation","visiblement","X","acteur","créateurs","dites","déposer","expositions","handicap","lourdes","plastiques","procure","proviennent","sous-jacente","Ni","Quick","Virgin","auxquelles","banquiers","baptisé","finit","venait","volant","Fiat","Joseph","Lyonnais","enseignants","geste","l'UCL","sérieuse","Mignon","Royaume-Uni","Vers","classes","doigts","encadré","froide","niche","prévision","servent","Baudouin","Nicolas","Smeets","arrivée","domestique","envisager","espaces","filet","inflation","posé","promouvoir","roues","Assurances","Capital","immense","incontestablement","lot","pharmacie","restructurations","sportive","L'ensemble","ci-dessus","d'activités","engagements","humains","introduction","organisée","Delvaux","assiste","couverts","franchise","L'histoire","annuellement","arrivent","causes","pierres","valent","volet","Hanart","Karel","Lotus","intention","l'acheteur","manifestement","prendra","profondément","relance","suivantes","suspension","commissions","divisions","développée","employé","fourchette","qu'est","s'occupe","vendent","Clinton","Jean-Marie","Maurice","Nationale","compenser","d'octobre","essayer","fondé","formidable","graphiques","professeurs","tester","George","Histoire","boutique","caméra","d'avance","fondée","heureusement","label","montagne","pensons","plate-forme","temporaire","tombé","tribunaux","évite","BMW","Monde","PB","condamné","culturelle","d'air","entre-temps","entrées","installer","perception","sauver","thé","Fermé","Peut-on","Unilever","accompagné","externe","franchi","jadis","manifestation","miracle","moral","refus","réunit","révéler","s'installe","Etienne","Evidemment","bateau","conseillé","d'écart","décrit","fréquence","l'occurrence","s'adresser","taxes","Company","concentrer","consultation","dorénavant","dynamisme","installée","profite","réunions","amateur","avoirs","calculé","d'atteindre","estimation","exerce","bloc","circuits","couper","courante","d'améliorer","d'instruction","effectués","fameuse","intéressé","montage","prévues","subsides","séduction","traités","trouvera","équipés","Aucune","ingénieur","réclame","rémunérations","tentent","tournent","égale","émetteurs","Prenons","agent","attentif","d'aide","d'oeil","existant","fluctuations","gré","l'administrateur","médicament","partiel","permanent","s'installer","situés","sportifs","vertu","Intranet","L'évolution","Quelque","allons","appartements","duquel","kilos","sicav","toit","versées","chaussée","d'huile","futures","individuelle","manifestations","raisonnement","sports","Christophe","DES","absolue","appelée","contente","d'idées","d'investisseurs","intense","money","répondent","tranches","Waterloo","assurent","calculer","choisit","citer","doté","fixes","inférieurs","mensuel","promoteurs","relais","sorti","télé","voisin","Corée","Lynch","dit-on","hiver","l'Association","l'ULB","naturelles","preuves","présentés","souffert","Qu'est-ce","attendent","camions","camp","contenant","curieux","détente","effectue","géants","l'endroit","l'intermédiaire","légale","n'étant","prestation","publiés","rente","réalisent","ski","soigneusement","vif","Cie","conviction","doubler","morceau","racines","tenant","universitaires","visiter","Center","Global","démarrage","entamé","fondamental","l'intervention","magique","procurer","records","universitaire","vrais","L'une","ateliers","avion","confronté","contribuables","doigt","drame","féminin","habitudes","l'immédiat","lutter","pétrolier","supérieures","vois","AEX","Bell","afficher","confirmer","conservé","d'offrir","détour","fusions","l'avons","l'équilibre","lever","malades","ouvrages","paradis","prouve","prévoient","remplacé","spéculation","Rwanda","concernent","départements","dérivés","identiques","marquée","n'avaient","prince","produisent","résidence","voulez","L'opération","Turquie","allocations","démontrer","enregistrée","individuelles","oublié","parking","proposée","Commerce","Guide","Tom","comprenant","débuté","engagement","fit","légal","participé","passées","présentant","présentes","quantités","échapper","Maystadt","Software","acquisitions","affirment","alentours","assureur","autonomie","canaux","inverse","l'adresse","l'automobile","modes","signaler","signée","Goldman","Notons","cancer","carnet","convergence","foule","indispensables","intégrée","nucléaire","opérateur","paiements","palette","pence","priori","promesses","tentative","Belgian","Corporation","Dutch","Tel","aérienne","boutiques","craignent","débiteur","entités","ouverture","procureur","puisqu'elle","sommets","supporter","traitements","voyageurs","Bureau","anglaise","argument","d'établir","imaginé","l'appui","mécanismes","personnelles","privilégié","satisfait","science","terrasse","tiré","trésorerie","télécoms","D'ici","chaude","coupé","esthétique","inscrit","poissons","refuser","s'effectue","tennis","Moi","Unix","appartement","clavier","démontre","organismes","pressions","regroupe","secours","sous-traitance","théorique","accessibles","courants","d'été","judiciaires","l'innovation","l'opérateur","précédentes","réaliste","aventure","d'Internet","effectifs","gains","l'opposition","l'unité","musées","rock","Coupe","Netscape","bain","déposé","espoirs","majoritaire","semblait","Digital","accorder","attire","d'échange","feuille","initiale","installation","krach","malade","opérationnel","pauvres","pont","préserver","publier","rechercher","recrutement","représenter","révélé","sanctions","traditionnellement","vapeur","Cobepa","Salon","confier","considérés","cultures","hypothécaire","illustre","introduite","l'échec","menus","multinationales","paient","pareil","problématique","quarantaine","rentrée","soutient","terminée","voudrait","carré","exemplaire","lorsqu'ils","nulle","posent","pratiquer","sida","versements","visites","étions","étrange","CBR","berline","cash","distinguer","durs","défend","efficaces","essence","exclu","jolie","photographe","propriétés","veau","DU","Journal","Nobel","Vieux","atteinte","chapitre","concertation","dégage","extérieurs","médicale","pareille","patience","recueillis","substance","transforme","voile","échec","Léopold","enthousiasme","fédérale","gloire","préparations","transmettre","visiteur","éd","Ajouter","Brederode","Européens","Jean-Louis","Tony","apporté","d'importantes","l'acier","libéralisation","observateurs","panique","présentée","réserver","signer","tendre","to","touristique","Récemment","brillant","conventions","décret","généreux","industries","joie","stars","égal","Sachs","continué","dessert","espagnol","est-ce","légende","passera","rapprochement","salariale","scolaire","Monétaire","assurément","contraint","coton","curiosité","entité","entré","l'architecte","libéraux","logo","parlementaire","parviennent","portables","provisoirement","routier","réservée","tourné","veiller","FN","Hoogovens","XVIIe","arbres","communs","employeurs","exercices","faisons","l'alimentation","magazines","maintenu","roses","répondu","spécialité","Citibank","Moscou","Times","accidents","adapter","amené","avoue","collectif","d'évaluation","dessus","indépendante","l'institution","l'établissement","peintres","rappel","réalisations","s'avérer","architectes","comprise","essentielles","examen","fidélité","héritiers","l'actualité","préférable","relancer","s'adapter","s'engage","sable","semestriels","significative","suisses","Grande","Nouveau","cadeaux","comportements","constamment","contribuer","d'images","offerts","périphérie","varient","Michelin","caisses","conscient","cédé","effectuées","faisaient","personnalités","s'engager","syndicat","Arbed","OPA","abandonné","cents","destin","drogue","fines","identité","invités","l'événement","modalités","négatifs","paru","répertoire","s'intéresse","Disney","Isabelle","Japonais","Roland","William","annoncée","champignons","défis","générer","russes","situer","supprimer","élu","Jean-Paul","Spa","accordé","acquise","courtier","d'attente","foulée","noirs","résister","section","signaux","sombre","susciter","compartiments","correspondance","créances","discret","dépassent","florin","formé","frappé","papiers","représentait","saurait","versé","absence","d'Or","d'acquérir","d'avenir","degrés","envoyer","joli","occupent","on-line","percée","priorités","processeurs","restés","résume","soie","travaillant","économistes","Etant","affirmer","ambitieux","cerveau","consensus","coordination","d'options","l'appel","magistrats","qualifié","rangs","tournée","Alcatel","Toyota","anonyme","c","cassation","cf (usually cf.)","confusion","discrétion","fondamentalement","initialement","installés","l'assemblée","l'entretien","l'émetteur","maman","nuances","paraissent","parfums","saine","vedettes","DM","Nikkei","dirigée","duo","enseigne","indiqué","kg","lourdement","module","prononcer","réalisateur","réformes","star","équivalent","Danone","Site","adopter","commis","couches","explication","joint-venture","malaise","pantalon","pomme","reine","sacs","saumon","soeur","toiles","échéant","Agusta","bond","courir","expert","glace","l'enseigne","multiplier","pluie","salons","teint","European","Finalement","Maintenant","adaptée","diriger","gérant","répartis","saveurs","souscrit","substances","vieilles","vraisemblablement","élaboré","émettre","certitude","champions","cotés","cyclique","détenteurs","explications","fonctionnent","générales","invite","l'expression","pauvre","successeur","zinc","Big","Claes","Six","brochure","cave","codes","configuration","d'enregistrement","fragile","féminine","issus","magnifique","maintenance","manuel","qu'a","recommandé","spectaculaires","subit","traduction","évidente","Conséquence","Fabrimétal","KBC","adaptés","chronique","d'IBM","enregistrés","fibre","jazz","jusque","louer","médiatique","peser","rentables","réussit","s'élevait","saisir","semble-t-il","visible","Financial","Singapour","absolu","blanches","boulevard","commissaire","comprennent","créent","faculté","histoires","individus","issue","multiplient","prétexte","quotidiens","réfléchir","satellites","souffrent","standards","Washington","commercialise","directs","diversité","gratuite","l'Office","logiquement","ouvertes","renoncer","calculs","compléter","couples","d'entrer","d'esprit","d'importants","l'acte","organiser","payant","paysages","récupération","slogan","Electric","PVC","administratives","arts","avancé","carrément","changes","crédibilité","déplacement","l'avance","parvenu","relatifs","revues","veste","Celle","FGTB","Moody's","assurés","créés","d'éléments","immédiat","jambes","litre","mousse","prestige","sentent","souhait","touché","élus","Belle","Telinfo","abrite","considérables","d'urgence","disait","faillites","oeil","religieux","rédaction","séries","terres","vice-président","MHz","System","XXe","cure","dirigé","don","enregistrer","juridiques","pouce","précises","prétend","réunis","salade","trouvait","évaluation","Cinq","Fort","confié","cuire","indicateur","l'avait","origines","parlé","remet","spéciales","terrible","témoignent","étonnante","Buffett","Catherine","Research","SAP","Véronique","achetée","généraux","imposée","l'organisme","l'édition","mention","merveille","opposition","réorganisation","satellite","scanner","Milan","Notamment","a-t-elle","acier","ch","conteste","créanciers","d'acier","intégrés","l'habitude","multiplication","panier","pharmaceutiques","quelconque","rayons","spectateurs","transformé","troupes","Madame","Tandis","effectuée","fromage","géré","interlocuteur","législatives","motif","métalliques","placée","réclamation","schéma","surplus","transition","trio","Coca-Cola","Motors","Proximus","Wallons","atteignent","bleus","chair","conforme","costume","d'accueil","intentions","l'horizon","l'électricité","manqué","sortent","subsiste","supermarchés","D'Ieteren","Européenne","Lorsqu'on","amélioré","avantageux","d'applications","engagée","espoir","exceptions","fausse","l'expansion","l'équivalent","plage","plaide","poivre","CHF","Livres","cadastral","chips","comptait","craintes","d'ordinateurs","durable","démocratique","exceptionnels","factures","fonctionnaire","fondation","indépendance","inventé","issu","maturité","mobilité","musiciens","organisme","recommandations","spéculatif","suscité","titulaire","traverse","évolutions","BD","Fed","calendrier","collective","disposant","dévaluation","l'honneur","pauvreté","poursuivi","qualifier","savait","suédois","termine","traduire","valait","CSC","Forges","Hugo","Max","VVPR","appartiennent","confrontés","demeurent","divorce","dramatique","déductibles","efficacement","existence","fermeté","imagine","intégrer","larges","locataires","orienté","pensé","variété","administrations","aériennes","complexité","entrent","exercer","photographie","sauvage","terminer","venant","Corp","amortissements","champs","déplacer","désigné","déterminant","opportunité","piano","remontée","s'agisse","étroite","AT","Difficile","Dix","Recticel","bar","concerné","constructions","l'identité","merveilleux","min","moindres","réunir","survivre","ultime","étudié","Lambert","RC","caractérise","choisie","distribuer","décidément","limités","livré","luxembourgeoise","modules","progresse","promet","redresser","tombée","bains","d'hommes","dessine","enfance","finition","jury","mythe","optimale","pair","plateau","poussée","resteront","Zaventem","assurance-vie","composée","d'entretien","décident","hélas","instant","jet","laine","mobiles","parcs","préoccupations","ramener","représenté","soudain","éditeur","José","L'auteur","Morris","Nasdaq","administrative","autorise","banking","humour","jouit","l'actuel","market","n'ait","organisateurs","peint","s'annonce","s'assurer","sculptures","superbes","équipée","ASBL","CMB","Gates","bronze","catholique","citron","contributions","couture","disquette","démarrer","excellence","fatigue","imprimantes","industrie","l'aménagement","l'effort","l'encontre","laboratoires","menées","meuble","mondiaux","réduits","sont-elles","sous-traitants","talents","Christine","Henry","administratif","administration","ailes","aérien","carrosserie","d'économie","découvertes","exclure","hautes","hiérarchie","impressionnant","massivement","métro","possession","remporté","strictement","suédoise","utilisateur","vais","émises","étage","d'arbitrage","devez","expliquent","file","hebdomadaire","intéresse","l'hiver","l'élaboration","marbre","performant","personnels","prévenir","suivants","verte","viendra","Angleterre","Association","Hongrie","L'affaire","Louvain-la-Neuve","OS","apportent","automne","bourgmestre","branches","carton","contraste","courage","d'analyse","datant","dépendra","feux","importations","plantations","sidérurgie","signale","FMI","Jean-Michel","Léon","Super","UN","Venise","adaptation","allure","attachés","exploite","folie","instance","naturels","olympique","populaires","reprenant","valorisation","villa","villages","Est-il","Renaissance","Shell","Vienne","architecture","authentique","autonome","complicité","d'au","d'ouverture","dépendance","dépense","fiable","invention","lancés","partagent","rencontres","renouvellement","évoluent","Akzo","Combien","Marché","Xavier","ampleur","analyses","bandes","canard","collectionneurs","compliqué","culturelles","d'avril","donnera","déplacements","fermer","jugée","l'aise","médaille","notaire","peut-il","privilégier","prototype","regain","regarde","wallonnes","Emile","Volkswagen","accru","caoutchouc","cinquante","communautaires","conjoncturel","créant","durer","délicat","exigent","précédents","renforce","s'ouvre","évalué","Lille","débute","définitif","engagés","exploiter","fur","positives","réparation","soupe","transferts","Ostende","Propos","Victor","limitées","nourriture","offertes","ramené","reculé","remédier","similaires","triste","écarts","Data","Industries","abaissé","boire","break","chien","consacrés","cours-bénéfice","fuite","gigantesque","imprimante","l'Ouest","l'emballage","l'église","remplace","salariaux","spectacles","vache","velours","étudie","ABN","Auparavant","Cité","Continent","Guido","Meuse","Mo","Question","d'exemple","dotée","défini","définit","délicate","démission","extérieure","interventions","jouant","l'engagement","n'ayant","noires","obligés","Bruxellois","Mark","Motorola","accéder","affichait","chemise","espagnole","fleur","gardé","habitation","huile","l'accueil","légales","multiplié","revers","architecte","assister","axes","concerts","contemporains","discuter","dose","détiennent","folle","l'éditeur","magie","pompe","provisions","rapidité","témoignages","Cap","Festival","Finlande","NDLR","contribue","demandeurs","démonstration","exact","numériques","participent","poignée","puissants","spécialités","G-Banque","III","Livre","Peeters","SICAFI","Technology","applique","copies","flacon","lunettes","mixte","nullement","plante","provisoire","publie","puissante","regrette","s'ajoute","stratégies","typique","vocale","Anhyp","Brothers","brokers","concentre","diagnostic","faciles","gestes","guise","hardware","opérer","orientée","passionné","refusent","scénarios","suffisent","vagues","écart","Chrysler","Sénat","Via","ambiance","appartenant","assisté","attrayant","bagages","blocs","d'essai","d'histoire","d'étude","déduire","forfait","manquer","restait","surprenant","sérénité","vertus","écouter","DKK","Dirk","Gevaert","HP","Santé","Wim","accueilli","affichés","affronter","appeler","coloris","composent","contiennent","contrepartie","fondamentales","impressionnante","largeur","peaux","proportions","reconversion","revente","significatif","écrite","énormes","J'aime","Network","aiment","cherché","chinoise","décharge","député","essais","indiquent","infrastructures","jouets","musicale","mutation","obstacle","partant","perdent","étudiant","J'avais","Sinon","accordée","adjoint","débarrasser","débit","dégustation","déjeuner","glisse","individu","l'éducation","l'électronique","organisées","produite","prétendre","quotidiennement","s'étend","secondaires","soucieux","sous-évaluation","verts","écologique","émet","Hollywood","Legrand","Lorsqu'il","Pro","améliorée","bat","e-mail","excessive","favorise","joueur","l'OCDE","marks","office","phrase","promenade","prometteur","stimuler","séances","tiendra","valoir","Martine","Québec","acquisition","augmentent","baisses","distribue","dus","massif","médiocre","obtenus","rentrer","sales","semblable","transmis","Julie","Place","ZAR","bouquet","ceinture","coalition","comptables","corporate","d'actifs","d'attendre","différemment","dits","italiens","journées","l'assurance-vie","linguistique","marchands","n'avoir","opinion","originales","registre","requis","synergies","tunnel","vogue","Malaisie","charbon","emballages","esprits","examiner","fléchi","l'outil","librement","mentalité","miroir","occidentaux","parité","progressive","sensation","sonore","supports","synonyme","vinaigre","Début","Euro","Hollandais","alliance","barres","chargés","d'habitants","dois","fier","gouverneur","l'atelier","l'humour","n'avez","origine","payée","pétroliers","signalé","variation","Né","Point","XVIe","aliments","caméras","comportant","consultance","contemporaine","déclin","effectif","invité","j'en","l'actif","licenciement","match","millénaire","salarié","studio","tenus","triple","équipement","étoile","Bob","Californie","Devant","Smet","abonnement","baptisée","commerces","creux","facilite","flamandes","jurisprudence","l'ai","l'attitude","noyau","portraits","prononcé","publications","puce","qu'aujourd'hui","sinistre","terminal","Dexia","Mes","augmentations","batterie","cinéaste","compare","guides","inconvénients","instances","l'avion","retourner","sympathique","évaluée","L'Etat","achetant","bailleur","bonus","colonne","compensation","conseillers","continu","courbes","déclarer","enregistrées","généré","innovations","ira","jusqu'aux","lente","occuper","pesé","pot","quarts","épreuve","Bois","Congo","Courtrai","Powerfin","admet","attribuer","championnat","cités","comble","conquérir","d'encre","d'oeuvres","d'office","devenues","excessif","incertitudes","intitulé","l'évaluation","périphériques","réclamer","réelles","s'étaient","Ecolo","Nivelles","Qu'il","Travail","allures","camps","dues","exclus","grandeur","homard","illustré","inévitable","inévitablement","l'équipement","mariés","modération","ont-ils","positivement","profits","quarante","sculpture","spots","stage","universelle","vainqueur","édité","étendue","Arts","Communications","Media","Novell","Poor's","Stéphane","Word","changent","communiqué","conversation","d'artistes","effective","interlocuteurs","l'Administration","l'ambiance","n'aime","patronales","permettront","pneus","qualifiés","religion","souffrir","évoqué","Chirac","Chris","Forest","Herman","Hubert","Opel","Parti","SEK","Terre","Vie","alternatives","anversoise","bateaux","battu","brillante","d'introduire","désert","entrepreneur","essayé","interface","intégralement","j'aime","lu","modifie","personnellement","systématique","Arthur","Park","admis","blocage","calls","développent","individuel","l'ONU","l'appréciation","modestes","multinationale","out","parlant","porcelaine","pénétrer","respecte","soupapes","spéculateurs","étudier","Nestlé","abus","combler","conservation","donation","fiabilité","l'exclusion","m'ont","parcourir","parisien","remarquables","retournement","returns","EASDAQ","Kodak","PDG","collecte","d'alcool","déception","détérioration","l'avoir","l'échange","lorsqu'elle","palme","phases","privatisations","répéter","s'imposer","valu","","","","","voulais","Almanij","Infos","Procter","Smith","Tubize","actuariel","australien","croient","d'intervention","d'objets","encourager","fiscalement","hautement","l'assiette","marchand","néerlandaises","plaintes","reproche","retient","sillage","soldats","témoins","urbain","FEB","L'économie","adopte","boutons","chuter","conjoints","convaincus","coopérative","correspondent","director","n'hésite","niches","savez","stables","tend","vain","CV","Gamble","L'art","Quinze","Servais","Seules","apport","chauffage","commercialiser","d'attirer","d'existence","d'organisation","dangers","foyer","ingrédients","négocie","révolutionnaire","score","sidérurgique","techniciens","voyageur","Brown","Corluy","Herstal","Horta","L'avenir","attiré","com","conférences","constatation","d'Amérique","douzaine","duration","détenir","indemnités","lion","nuits","plomb","soumise","sportives","verres","attribué","corriger","d'hiver","domestiques","faille","foot","home","indemnité","romantique","simulation","Brussels","L'avantage","Swissair","autrefois","choisis","communales","d'Angleterre","dessinée","disponibilité","détenu","engager","exceptionnelles","figurer","habitant","hollandaise","immédiate","intégration","média","électeurs","Amro","DOS","Moniteur","Parc","acceptable","apprécier","centre-ville","d'elle","envisagé","fantaisie","habituellement","posséder","pourrez","tentatives","touches","visibilité","Creyf's","Heineken","Régie","Sterk","Tchéquie","analyser","autorisé","complets","contrainte","costumes","d'agir","doucement","démarré","eut","posée","raffinement","rond","sidérurgiste","ABB","Ensemble","L'offre","Me","accroissement","ajouté","assiette","autoroutes","batteries","d'Asie","dame","disciplines","décrocher","essai","essaie","fréquent","génétique","inconnu","l'avocat","majeurs","multimédias","plume","probabilité","préavis","publiée","scandale","spot","sérieuses","tomates","égales","Coup","Die","L'investissement","animé","bleue","d'utilisateurs","danoise","essentiels","fondateurs","fonde","l'Atlantique","l'épreuve","maquillage","mexicaine","oeufs","opérationnelle","prestigieuse","renforcé","rumeur","soigner","témoin","Blok","Golf","Nouvelle","Prudential","Tonneau","Wavre","affichée","attendus","ballon","bouton","chanteur","chiens","d'écrire","entrepris","exprimé","nomination","perfection","photographies","renommée","sous-évaluée","universel","vives","BT","Eaux","Jacobs","Raymond","axée","cacher","défauts","l'aube","l'octroi","méritent","occidentales","planning","rage","testé","Barbara","Britanniques","Interbrew","Technologies","Visa","acquiert","adulte","affichant","agricole","annonces","anversois","atteignait","cabinets","centenaire","confection","culturels","d'aucuns","destinations","doutes","développant","ha","l'ASBL","l'autorisation","l'émotion","masque","méfiance","officiels","outre-Atlantique","panne","perdue","pouces","protégé","rires","sacré","silhouette","soeurs","virtuelle","vues","écrits","épreuves","Impossible","Madrid","appelés","candidature","chargement","documentation","dominante","députés","indications","l'indépendance","leaders","listes","mince","opte","énergétique","étendu","Ci-dessus","L'exercice","University","aisé","concepts","d'achats","d'agences","d'alarme","disquettes","domine","développés","envoie","exercé","existante","fauteuil","habitations","italiennes","modifiée","nets","puces","réclament","tomate","tons","CERA","Campo","Val","aluminium","assumer","confortables","d'ajouter","dates","démarre","exotiques","expose","hall","l'aluminium","l'enregistrement","licences","navigation","opter","rapprocher","émotions","Bertrand","CPAS","CompuServe","Jamais","Jo","Klein","Swiss","acquéreur","courtiers","doublement","défaire","l'Ecole","libération","lin","percevoir","perdus","plonger","poésie","process","s'accompagne","saisi","signification","songe","séparément","tentation","trou","variables","Collignon","Los","accusé","affecté","apparu","complément","créatif","exposé","financé","incite","investissent","limitation","montagnes","onze","originaux","ouvrier","partagé","professions","préalablement","tonne","trajet","visent","L'industrie","Magritte","Power","St","balle","chercheur","communale","compression","crises","d'articles","démontré","détention","détermination","endettement","gel","inchangé","incontournable","l'enfance","l'once","l'écriture","lent","modernisation","organes","promotions","présentées","relief","remboursé","rencontrent","sage","Monnaie","accuse","axe","distances","dépréciation","détenteur","fournissent","gendarmes","horaires","imposables","intercommunales","lancent","multitude","particularité","partisans","provinces","quelques-uns","reviennent","s'améliorer","vernis","volontaire","ambition","baux","d'apprendre","directives","dis","délégation","détecter","détermine","exprimer","individuels","l'unanimité","localisation","miel","optiques","passif","performants","piles","politiciens","poussent","privilégie","restreint","souligné","supporte","vallée","ventre","éthique","Anglais","Assubel","Frans","L'utilisateur","Saint","Supposons","VLD","activement","analogue","bassin","boulot","chapeau","claires","connaissait","consortium","culte","d'administrateur","dégradation","hypothécaires","l'exécution","l'incertitude","lumineux","maux","perles","porte-parole","privilégiée","privilégiés","qualifiée","réagit","résumé","Australie","Axa","Etre","Fernand","Jules","L'activité","Lui","apparence","ci-contre","comédiens","connecter","continuera","convertibles","correcte","dessiné","durées","enjeux","incapable","libellées","millier","populations","épargne","évoquent","CLT","Mobile","capot","charger","communiste","comparables","congé","correct","d'inventaire","gagnant","intérimaire","l'UEM","l'autorité","l'envie","libéral","liquides","orientale","passait","policiers","redoutable","solidité","traitées","Bond","Mme","SME","affiches","attente","basés","comparé","consommer","d'armes","inconvénient","jette","maigre","masculin","négocié","profitent","vaches","vitrine","Avis","Finance","Hotel","Index","administratifs","allaient","avancée","compétitions","confiée","contenir","d'Art","d'opérations","description","dormir","exclusif","gouvernementale","joliment","l'Intérieur","mailing","modérée","pénal","raisins","rempli","royaume","sanction","spéculative","terminaux","Chanel","Grand-Duché","Leo","accélération","analyste","bourse","coupable","d'expérience","d'honneur","fabriqués","mettra","morte","paye","prudents","récolte","réduites","s'intéresser","saisons","sexuelle","sondage","strict","tenues","tranquille","agit","briques","concernée","d'envoyer","d'occupation","débiteurs","illustrations","paroles","passager","payées","piscine","plages","rapporter","Gillette","S'ils","banlieue","commander","commentaire","composer","d'appareils","distribués","détenus","fabuleux","génocide","hésité","indication","indiquer","industrialisés","l'imagination","l'individu","l'évidence","minimale","musical","nerveux","plaisirs","pop","positionner","rassurer","réalités","réservés","sons","séminaires","trésors","uniforme","vis","Blue","Boston","Ch","Gold","Lisbonne","Nul","SNCI","VW","affiliés","appréciation","avancées","cafés","casser","chat","compatible","d'Afrique","d'actualité","franchisés","l'apport","lots","moules","mécaniques","papa","préparé","qu'auparavant","souhaité","traits","éventuels","Cannes","Chicago","Chinois","Lee","cou","d'espace","d'expansion","d'offres","d'évaluer","discrimination","dépassant","expliqué","externes","fédérations","metteur","mobiliers","paire","peseta","polices","productions","préoccupation","rappellent","rentrées","réparties","spécialisation","statuts","traditions","truffe","éliminer","étroit","Hélas","Napoléon","Onkelinx","audio","av","baroque","brique","carburant","conducteurs","docteur","délocalisation","désirent","e","estimons","flotte","harmonie","l'écrivain","livrent","marquer","obstacles","rapporté","s'ajoutent","sobre","tandem","timing","travaillons","électorale","établis","Avermaete","Baudoux","Caisse","Elisabeth","Picasso","Proton","abandonner","assiettes","axé","civilisation","compréhension","confirmation","crime","d'énormes","diversifier","dresse","extension","fantastique","ignore","incendie","levé","minoritaires","négliger","pacte","panneau","pratiquent","raisonnables","retenus","retombées","revenue","sponsoring","étages","Bayer","Contre","Inutile","McDonald's","Nike","Pr","aborde","adoptée","bouger","cassettes","certification","exigeant","frein","fréquente","imposées","l'infrastructure","manipulation","optimisme","pourvu","profondes","rassemble","retiré","shopping","stabilisation","vertes","volontairement","établit","Dior","accompagnée","admettre","aurez","automatiques","concret","constant","d'organiser","drogues","fourneaux","influencer","interprétation","intime","magistrat","occupée","prouvé","recommander","sélectionner","séparer","tube","voyager","écrivain","équivaut","étroitement","évoquer","AUD","Age","Design","Oracle","Petite","accueil","approuvé","boom","construite","continuité","cotait","dresser","décors","gravité","l'Eglise","légers","menaces","mondialisation","passionnés","s'exprimer","stricte","styles","superficie","tas","terroir","versée","Dernier","Total","applicable","avancer","boisson","camion","commercialisé","composantes","concrétiser","conservateur","dames","diable","diffuser","fixés","formulaire","freiner","l'enthousiasme","l'élément","mourir","oublie","placées","péril","raffinage","rapproche","remarqué","rendue","résiste","réunies","sponsors","transférer","verse","Budget","Denis","Hoechst","Hotels","Pfizer","RNIS","Walter","averti","conseillons","disposé","défunt","failli","grammes","libérer","nickel","popularité","pratiqués","rassembler","regards","requiert","spécifiquement","successives","théoriquement","timide","volontaires","Barcelone","Dell","Ligue","Mettre","Simon","TEXTE","Zurich","associée","ci-dessous","connaissons","d'apporter","douter","fabriqué","figures","finesse","innovation","jeté","justesse","l'endettement","mémoires","neuve","pile","plaque","promesse","pénétration","rouler","répondant","s'agira","sachant","situait","Chili","Florence","Information","Qu'en","acceptent","accélérer","adaptées","adeptes","adolescents","attendait","d'émettre","foire","habitués","incontestable","interview","l'attente","l'indique","ndlr","noix","oiseaux","permettait","qu'aucun","qualification","rachats","rendant","réputé","syndicale","tombent","touristiques","valoriser","voler","écu","Alexandre","Dupont","Environ","atteints","clarté","compact","concurrentiel","décroché","enseignement","herbes","intelligent","licenciements","négatives","rédigé","répercussions","salariales","tempérament","tutelle","varier","verdure","Cologne","PNB","SOMMAIRE","aisée","bijou","bénéficiant","combattre","d'humour","donnait","exacte","exposés","expression","fonctionnant","gérés","impérativement","indépendantes","l'opportunité","levée","lorsqu'un","non-ferreux","nostalgie","nourrir","pratiqué","richesses","robot","saisie","Andersen","King","aborder","assortie","brutalement","brute","disparaissent","engendre","heureuse","injection","jaunes","l'intelligence","montent","nuance","officielles","physiquement","prudente","refaire","segments","supposer","synthétique","vieillissement","échéances","Autres","Bonne","Dupuis","Elio","J'étais","TENDANCES","cassette","colonnes","d'impression","dépression","fouet","judicieux","l'actuelle","l'entrepreneur","lac","portait","riz","sauvages","secrétariat","situées","solaire","soucis","souscripteur","séparés","transparent","truc","typiquement","Blanc","Café","Généralement","KLM","United","améliore","annoncer","assemblée","cahier","chanter","compétitif","concession","connais","crédible","d'accéder","d'équipement","dangereuse","destruction","envisagée","glisser","grimpe","infrastructure","insuffisant","isolé","minorité","passionnant","possédait","recommandée","régimes","s'interroger","s'offrir","syndical","trouble","Angeles","Barings","Défense","Equity","L'agence","Merrill","Mouscron","Spitaels","Xerox","abouti","adéquate","archives","attribue","boeuf","directes","directions","filets","fleuve","indéniable","interrogées","l'alcool","observer","prescription","revendications","rude","successifs","tri","trimestriels","télécommunication","Communication","Eddy","Hasselt","L'expérience","aboutir","anticipé","arrivés","cohérence","collaborateur","compositeur","cravate","d'excellents","d'outils","dons","indirectement","interviennent","l'Histoire","l'estimation","multiplie","qualifie","retenues","réalisme","suffisait","variétés","échappe","Antoine","Communautés","Hauspie","Larry","Petercam","Seulement","Sofina","accessoire","aimé","bébés","canadienne","cuisinier","d'imposer","détenue","existait","explosion","flexible","licencié","livraisons","malheur","militants","minime","notions","opposé","organisés","particules","ratios","reprennent","semblables","tournage","trous","Holding","IP","L'accord","Las","Pékin","artistiques","avancés","cosmétiques","fournis","héritage","implantation","l'adoption","originaire","partielle","passions","proposera","recherché","renforcement","sexualité","suites","téléphones","témoignage","vignes","vitamines","Code","Général","Inde","Johnson","Mobutu","Multimédia","Méditerranée","Notes","Vendredi","amoureuse","brevet","consomme","cristal","créées","d'image","descendre","décline","entame","faim","freins","inspire","l'explosion","maritime","menacé","pervers","poursuivent","savons","sols","sorties","soumettre","strictes","venues","Chantal","Exchange","Galerie","Hans","Maroc","Nathalie","Nombre","PJ","arme","bombe","bouge","chantiers","chic","extraits","gants","informer","insuffisante","l'unique","liaisons","limitent","nourrit","précautions","rencontrés","renforcée","résolu","s'ajouter","s'occuper","surveiller","trimestres","échappé","éclairage","élargir","Bosnie","Lyon","Nicole","Primo","Tintin","accompli","comportent","considérées","croyait","di","définie","enseignes","estimer","exiger","fiches","forcé","forfaitaire","gourmand","identifier","joint","l'extension","massive","millésime","minoritaire","plaques","restants","s'imposent","stabiliser","virage","vraies","éventail","Belcofi","Vlaams","audacieux","aventures","banal","communautés","douche","détection","exprime","honoraires","jugés","l'avenue","littéraire","marginal","montres","peloton","procédés","protocole","rendus","renoncé","subissent","suggestions","sépare","turbo","voulaient","Faute","Stock","achetées","agricoles","circulent","confirment","coucher","d'aluminium","détenues","extérieures","finement","fondre","fous","horaire","imposés","inciter","l'adaptation","l'employé","mondiales","médicaux","occidental","occupait","offerte","orientés","prévaut","pureté","reposer","tendresse","élégance","ISO","NC","Sport","XEU","alliances","amont","annoncent","appliquées","appréciable","appréciée","artisans","assis","brasseur","clairs","cubes","d'afficher","d'auteur","d'avion","divisé","débourser","décevants","emprunte","fermement","générosité","horizon","l'Emploi","l'accroissement","l'envoi","l'examen","originalité","parent","plongée","scission","sondages","soumises","surprenante","triomphe","vides","virtuel","émise","Botanique","Busquin","Consulting","Jean-Marc","L'exemple","RAM","Vietnam","Young","achetés","africains","apprécie","attaque","bruits","cabine","clichés","criminalité","densité","exclusive","extrait","facilités","fixation","gastronomie","inconnue","investie","l'agent","l'humanité","maintenue","mythique","piliers","psychologie","raconter","récompense","résidentiel","saint","signés","situant","songer","thérapie","touchent","tromper","échappent","équivalente","America","Avenue","Dont","Explorer","Jeune","Lux","Malines","Vos","artificielle","assistée","branché","brutale","clos","complice","compétent","emprunté","enthousiaste","entourage","fierté","formats","genres","loue","marié","mentalités","ministre-président","peines","poupe","prolonger","pénale","queue","ressemblent","rotation","savoureux","sensations","tuer","Faites","Federal","Hewlett-Packard","Investment","Jack","Johnny","Lernout","Retour","Réserve","boule","chaise","disposons","délit","démarches","entouré","informé","l'élégance","loué","majoré","n'hésitent","pessimisme","planification","pompes","puts","syndicales","triangle","vitres","épargnants","First","Mozart","Real","Rotterdam","albums","amener","appellent","baron","brevets","carbone","carnets","combine","consolider","cycles","d'accepter","d'épuration","diamant","douloureux","dépendent","développées","frigo","l'Hexagone","masculine","modernité","nucléaires","ouvrant","pavillon","recevra","restrictions","revirement","rétrospective","solitaire","spontanément","stages","tubes","Ackermans","Ben","Entreprises","Institute","L'argent","Warner","calculée","chasseurs","circonstance","cite","commandé","concurrencer","conditionnement","d'escompte","d'ordinateur","disposait","exceptionnellement","favorablement","fiables","frapper","fédéraux","grossesse","généralisée","inquiétant","l'UE","l'anglais","m'en","n'avais","précieuse","précieuses","prévoyons","rejet","résiduelle","réuni","seconds","solitude","trésor","épais","BASF","Européen","Roumanie","Sept","Telenet","ad","approfondie","catalogues","center","circulaire","colle","composés","concentré","concepteurs","contribuent","d'architecture","décidée","délicieux","excellentes","inspirée","intervenants","joindre","médicales","plaindre","plongé","prévisible","qu'aucune","romans","répétition","stagnation","stand","suivies","ticket","vanille","économiste","élégant","Airlines","Culture","Exposition","Indonésie","L'occasion","Lunch","Net","Néerlandais","South","Sécurité","Trends-Tendances","accumuler","ajoutant","armée","assorti","bénéfique","cigarettes","conte","d'économies","désireux","fiction","finir","fixées","handicapés","investies","l'écoute","manoeuvres","nationalité","onéreux","paquets","ponts","prenne","rattraper","recouvre","reflet","reporter","rwandais","serait-il","texture","traverser","visibles","élargi","établies","Front","Irlande","Moyen","Style","Verviers","bords","composant","conjoncturelle","couteau","credo","d'installer","d'édition","descente","désirs","gagnent","galeries","imprimés","intellectuel","job","l'appellation","l'intéressé","ludique","parisienne","prédilection","publiées","rebond","reconstruction","éventuelles","Bonn","Cofinimmo","Force","Line","Midi","améliorations","angle","brochures","chaleureux","conviennent","d'André","détruire","examine","exonération","familiaux","l'Américain","laver","linge","légendaire","matériels","mensuelle","n'est-il","noblesse","prévenu","refroidissement","retrouvera","réguliers","saut","saveur","surprendre","tir","versés","DSM","Erik","L'administration","Loin","Lyonnaise","Pire","Vlaamse","chante","chemises","confirmée","connexions","d'identité","déroulera","fausses","grille","habite","implantée","imposant","influencé","intéresser","l'autonomie","mondes","moniteur","métallique","n'arrive","neufs","pondération","prédécesseur","repreneur","s'efforce","sera-t-il","sous-évalué","trompe","va-t-il","vendant","vidéos","voudrais","étiquettes","évalue","Franki","Geert","Liégeois","Mitterrand","PLC","adéquat","autorité","concessions","convivialité","cordes","courtes","d'avantages","d'étudiants","désigner","essor","exigence","exploitation","exportateurs","habituel","lancées","livrés","mets","mineurs","parlait","parlement","pause","primordial","publicités","recherchent","émotion","Casterman","Deceuninck","Découvertes","ESP","GM","School","Shanghai","West","accompagner","assemblées","championnats","classé","col","couvrent","d'antan","d'envergure","dictionnaire","exigé","formés","gérée","jeans","juifs","justifié","l'abandon","l'attribution","lenteur","manifesté","oreilles","prononcée","reconnaissent","reconnue","reproduction","seize","serait-ce","soixantaine","sortira","spectateur","vice-Premier","vingt-quatre","voyez","élégante","AU","Dyck","Hollande","Laurette","Manager","Partners","Rover","Tu","Warren","affirmé","allé","capteurs","chant","clinique","corde","d'annoncer","douces","dés","excellents","exercée","faiblesses","incapables","moeurs","nation","occupés","pesant","prolonge","proposons","prospérité","présentait","reprocher","réglé","s'adressent","sauvegarde","sauvetage","vitamine","Douglas","EOE","L'étude","Nouvelles","Porsche","Produit","Transports","Winterthur","bruxelloises","calcule","causé","chauffeur","concrets","conscients","conservée","conversations","crevettes","d'Albert","découpe","déterminées","envisagent","fermes","fermés","imposition","infrarouge","interactif","l'ambition","l'interdiction","l'obtention","langoustines","oreille","pencher","pessimiste","pessimistes","pilotage","policier","porc","prétendent","quotité","remous","représentaient","référendum","répartir","répression","semblerait","souffrance","substantielle","voisine","Italiens","Logique","Mhz","Traité","Walibi","animal","appellation","couler","coulisses","d'usage","disant","diversifié","déboucher","débouchés","dégagé","déguster","dérive","fumée","garage","inculpé","intellectuelle","l'actionnariat","l'essence","lisse","morales","peut-elle","pro","prospectus","raisonnablement","rapportent","retombé","régulier","s'attaque","sauter","scolaires","servira","soirées","suffira","ultérieurement","usagers","Autriche","Delta","Economique","Home","ITT","Iris","Reweghs","UNE","assurant","cardiaque","catastrophique","chaussure","cocktail","compatibles","d'adresses","d'immeubles","durement","détecteurs","implanté","intervenu","intervenue","jargon","l'Inde","l'instruction","mixtes","peso","promoteur","précisé","ralentir","rayonnement","remplit","remporte","successivement","séminaire","travaillait","urgence","équipées","Francisco","Gabriel","London","Mélissa","Nederland","Pacific","Portrait","Reine","Sipef","Sophie","académique","africain","allie","bougent","chaises","condamnation","crus","d'intégrer","digitale","doubles","décédé","exclue","fromages","habitude","imaginaire","jugent","l'utiliser","l'étiquette","mortes","notations","","","","obtenue","partent","retrouvée","roue","récits","sortant","tolérance","trouveront","vice","visuel","visé","vivement","vivons","émetteur","Apparemment","BP","Dassault","Elf","Genk","Ltd","Museum","accusations","bactéries","bouillon","brillants","clauses","collaborer","consentis","cri","d'amélioration","d'arriver","desserts","dynamiques","débouché","déposée","désigne","emballage","fumé","gibier","inscrits","intelligence","l'Angleterre","l'assistance","l'instauration","l'optimisme","l'égalité","longuement","légion","ménagers","noble","normales","polémique","repérer","secs","sombres","teneur","traversé","trucs","évitant","BNB","Banques","Belfox","Don","Hervé","KUL","Moyen-Orient","N'oubliez","Pays","Photo","aidant","anonymes","attrait","avez-vous","capitalisme","censé","clefs","considérations","d'Atlanta","d'échelle","demeurant","high-tech","linguistiques","manteau","monstre","n'auront","posées","proposait","préférer","remarquablement","s'appuie","sagesse","simplifier","surprend","tablent","te","trains","écoulée","écoute","étiquette","Constitution","Rouge","Valérie","aboutit","conformes","conjointement","conquis","d'aucune","d'espoir","d'intégration","l'époque","devienne","donnés","faveurs","gratuits","impliqué","inclut","intelligente","l'Irlande","légalement","nécessitent","occasions","perdant","portera","prenait","propice","rebelles","religieuse","réduisant","sale","sobriété","sophistiqué","sophistiqués","spécificité","transporteurs","urgent","usages","Haaren","Jusqu'ici","L'Union","Man","Moteur","Notre-Dame","Products","Robeco","Verts","Villa","abondamment","admirablement","assistance","casserole","chaos","citées","confiant","convenu","d'échéance","danois","durables","dépenser","extraordinaires","figurant","installe","l'agriculture","manifestent","marcher","nappes","paradoxe","partagés","protégés","pédagogique","raté","renvoie","résultant","s'exprime","soulignent","temporairement","trancher","unes","émanant","épaules","Brande","Honda","L'appareil","Nissan","Star","Tourisme","capitaine","communistes","compositions","convertible","d'imagination","doter","détenait","fassent","fermentation","fréquents","incité","indexé","intéressées","majoration","marine","maxi","meurt","musicien","musiques","navire","oppose","parkings","performante","produites","remontent","rigide","récolter","s'intéressent","salarial","salée","spectre","spread","studios","tirent","trading","yens","électricité","étendre","Black","DAX","Electrafina","Ernst","Gestion","Gilbert","Jim","L'utilisation","Lorsqu'un","Pol","Royaume","Wide","accompagne","annoncées","anticipée","clés","condamnés","contentent","d'apprentissage","d'il","d'éducation","del","diamants","doré","douleurs","défavorable","dénomination","entretient","farine","favori","fiche","fluide","huiles","incroyable","l'erreur","l'implantation","mentionner","modéré","multiple","musulmans","n'exclut","participant","prioritairement","préjudice","préventive","réalisant","suivis","tué","épices","équilibrée","Banksys","Cornet","Fontaine","Golfe","L'analyse","L'effet","L'image","LVMH","Laisser","Nm","Secundo","Soir","Texas","alimentation","annoncés","assistant","attrayante","caché","chinoises","communaux","comédien","continents","crayon","d'adapter","d'attention","d'égalité","dessiner","dignes","dira","dotés","enveloppe","exporte","facettes","gosses","indépendamment","interrogations","intégrale","issues","jupe","l'apprentissage","livret","n'est-ce","opportun","planches","présidents","pénurie","ratings","regroupement","respectifs","restons","retourne","soixante","séjours","sénateur","tailleur","volaille","voter","week-ends","Autour","BSR","J'en","Jean-Jacques","L'émission","Mondiale","SERGE","VAN","Vilvorde","adversaires","agressive","complété","conservent","coutume","croix","d'attribution","d'expression","d'exécution","d'habitation","d'imaginer","demandant","disposés","estimant","gardent","hoc","hormis","juristes","mystérieux","plonge","réflexe","s'agir","s'interroge","s'offre","servant","souples","souscrite","sus","teintes","Bus","Cartier","Dixon","Geschäftsidee","Lundi","POUR","Telle","affectés","amie","apportées","assume","autorisation","biologique","cachent","compatriotes","conservant","copains","couvrant","engouement","grandir","intégrées","l'affichage","l'américain","l'attrait","l'importation","maigres","merci","mélanger","plumes","politiquement","portugais","pratiquée","progressif","proie","qu'avant","qu'ont","reculer","relier","représentations","respecté","s'améliore","soif","statistique","sympathie","tenait","trouvant","vocabulaire","Armani","Caterpillar","Champagne","GIMV","Koramic","Ouvert","Raison","SRIW","Steve","Ter","VTM","Velde","Wathelet","arbre","d'appliquer","desquels","déboires","entamer","escompté","exister","illustration","indirecte","justifiée","l'angle","l'autoroute","l'emporte","leadership","légitime","nettoyer","nier","observé","opéré","orientation","placent","projette","rachetée","raffiné","rails","remises","retiendra","roulant","réussie","satisfaits","signatures","tienne","vastes","voyait","états","Airways","Ardennes","Danvers","Degroof","Greenspan","L'usine","News","Olivetti","Rabobank","Radio","Systemat","Vanden","Xeikon","aéroports","barreau","casse","consenti","convertir","créance","d'hier","d'inspiration","descend","dominé","dérapage","filtre","finissent","fériés","inutiles","l'épouse","librairie","malheureux","manifester","n'auraient","nippon","parier","patronat","pis","profile","renseignement","ressent","robes","réglementations","réputés","satin","sculpteur","sien","transporter","Albin","Focus","Gaston","Greenpeace","L'essentiel","Prague","Prince","Rue","Steel","Temps","amendes","audit","avéré","confrontée","contradiction","d'enfant","d'intérim","dédié","explosé","exécutif","foires","imprimer","l'accident","l'union","mandats","ongles","pensez-vous","pensionnés","persuadé","poil","provision","provoquent","radicalement","remboursements","renouveler","représentés","régie","s'établir","sexy","supprimé","suscitent","téléphoner","vignoble","échoué","écologiques","éducation","Airbus","Bangkok","Carlo","Edouard","Grand-Place","Janssens","Kinshasa","Nouvel","Time","Trust","anticiper","asbl","attendons","basées","biographie","cliquet","combiné","conformément","conservateurs","d'Espagne","d'actionnaires","d'adaptation","d'elles","d'occasion","designer","douteuses","enlever","ferroviaire","fossé","giron","généreuse","haine","hier","implantations","inférieures","l'ONSS","lorsqu'une","obligatoirement","paisible","plaire","pressé","reliés","révolte","s'établit","sexuel","souhaitait","stocker","survivant","tactique","terriblement","union","viandes","visés","écrin","évoquant","Baan","Desimpel","Engineering","L'inflation","argumente","attractif","clan","colorés","comporter","d'anciens","demi-heure","devenant","définis","départs","envoyés","express","forcer","frappant","hormones","islamistes","l'Autriche","poursuites","pénible","qu'est-ce","rangement","regroupant","représentée","requise","riverains","sourit","transformée","truffes","turbulences","économiquement","écrivains","Bordeaux","CNP","Caroline","Clerck","Fina","Gilles","Gol","Prenez","Puilaetco","Puissance","Yvan","allez","bilans","blues","bordure","brin","chef-d'oeuvre","choisissent","contestation","d'Electrabel","d'Italie","d'exemplaires","d'habitude","d'équilibre","d'éventuelles","découle","déflation","emplacements","entretenir","entretiens","freinage","géographiques","jolies","l'Otan","l'affiche","linéaire","lune","manquait","nations","payante","précise-t-il","quelques-unes","radicale","rattrapage","renouveau","restaurer","semi-conducteurs","statue","sélectionnés","touchant","échanger","Atlanta","Auto","Avantage","Bel","Chemical","EMI","Indiens","L'arrivée","Packard","alcool","appétit","attirent","aurions","banquette","cessent","comte","céramique","d'analystes","d'introduction","d'événements","desquelles","directrice","décalage","déclarés","exposées","grandi","honte","incarne","interdiction","l'Indonésie","l'utilité","meurtre","monuments","must","nés","percer","performantes","préparent","prévoyait","purée","rappelé","rejeté","relié","rédiger","rétablir","révélateur","révélation","s'appliquer","s'ouvrir","temple","textiles","tirage","traversée","visuelle","échecs","étonné","Alan","Boris","Brigitte","Guillaume","Homme","Innogenetics","Jardin","Jeanne","N'est-ce","Picqué","RTL-TVi","Restaurant","Sioen","Suisses","accordés","attentivement","autorisée","banale","basilic","cacao","cauchemar","coupes","d'agneau","d'efforts","d'emballage","d'exploiter","d'installation","d'équipe","décideurs","défendu","déficits","délégués","déroulent","exercent","exécuté","finira","garantis","gardant","humanitaire","impératif","introduites","l'acteur","l'indexation","mineur","modems","passages","perquisitions","plaignent","princesse","présidente","raffinée","recueillir","refuge","soft","sterling","suffi","tragique","verront","viendront","élaborer","AGF","Charges","Grands","Gustave","Janssen","Kennedy","Kohl","L'assureur","Pascale","apprentissage","apprécient","assise","by","catholiques","chocs","chrétien","cl","collectionneur","compagnon","d'UCB","dangereuses","demandés","diffuse","diffusé","explicitement","flou","féminins","gage","honneur","imprimé","intelligents","l'immense","l'oreille","l'émergence","lacunes","libellés","majeures","mexicain","nu","olympiques","pairs","penche","pourparlers","poursuivra","prenez","privilégiées","prépension","remarques","récupéré","régulation","s'inscrire","s'inspire","s'étendre","saura","schémas","sensualité","sociaux-chrétiens","titulaires","ultérieure","urbaine","vaccin","valables","voor","voyons","écriture","épingle","épisode","Anderlecht","Bird","Citons","Droit","Goossens","J'y","Marianne","Président","Roche","agréablement","aéronautique","courtage","d'aider","debout","devoirs","distinctes","dévoile","interactive","l'assassinat","litige","microprocesseur","nomme","osé","parvenus","patienter","payement","permanents","protagonistes","ralenti","réplique","subsistent","symboles","toxiques","Astra","Benetton","Ci-dessous","Dorénavant","Eurotunnel","Floride","Hilton","Imperial","Japan","L'ancien","Navigator","Révolution","Simplement","amende","animés","caution","coquilles","corrections","d'emprunts","d'influence","déménagement","déplace","déroulement","ensembles","fermée","folles","ignorer","individuellement","installées","interdite","jambon","l'empire","libéré","magnétique","pionnier","regret","réflexions","réunissant","sud-africain","séparé","voté","éloigné","émettant","Deprez","Equipment","Fondée","L'exposition","NZD","PRATIQUE","agenda","bourses","briser","cerner","chaudes","combiner","compatibilité","concurrentielle","confrontation","consécutive","d'efficacité","déclarent","dénicher","employées","exotique","fourniture","guerres","hutu","idéalement","interdire","l'Art","l'asbl","l'impulsion","litiges","législature","marie","masques","mousseline","nipponne","octroyé","précédé","prénom","préparés","redonner","respectant","risqué","réjouir","réservation","saturation","senior","trouverez","Ahold","Ancia","Bancaire","Central","D'aucuns","LCD","Online","Pouvez-vous","Shakespeare","Sont","Vision","bougies","calmer","confondre","couloirs","cuit","d'effets","d'excellentes","duré","démontrent","euros","fautes","financée","fortunes","fournie","grains","inspiration","intéressée","intérieurs","l'automatisation","l'imprimerie","modo","motivations","moulin","mélodies","ose","persil","reproduire","régression","sain","scrutin","succédé","sursis","survient","séquences","tracé","éclat","élan","éprouve","Audi","Development","Globalement","Limbourg","Loi","Noir","Philippines","SPRL","aimerait","augmentant","bricolage","bruts","cesser","charmes","comprends","comptoir","curieusement","d'assainissement","d'avions","demandait","décoratifs","décrire","empire","emporté","emprunter","envergure","flot","fraction","féminité","islamique","jouera","l'animal","l'euphorie","luxueux","manipuler","muscles","nocturne","persiste","poches","prématuré","présentera","présidentielle","prévient","rassurant","routiers","salut","sceptiques","serons","somptueux","sursaut","séparée","télévisée","Avez-vous","Biest","Créée","Feiner","Forum","Munich","Option","Vlaanderen","abondante","adressé","arabe","bulles","cartons","casque","civiles","commercialisés","contacter","courent","d'environnement","d'intervenir","d'étonnant","diffusée","décroche","dénonce","dénoncer","désastre","entamée","espérons","except","fortiori","fournies","homologues","incertain","instituts","l'embauche","l'héritage","l'étudiant","libérale","lie","mafia","nie","nuages","oscille","pianiste","polyester","rachetant","ronde","satisfaisante","solaires","spatiale","syndicaux","séduisant","Associates","Beke","Conclusion","Eh","Expo","Fidelity","L'Allemagne","Nothomb","PET","accomplir","bulletin","cascade","chutes","cirque","colloque","compétitifs","coordonnées","coquille","couverte","créancier","d'experts","douteux","détruit","fondu","grec","grecque","houlette","inattendu","intellectuels","l'encadré","luxembourgeoises","majoritairement","marier","mariée","monsieur","nominale","notera","obligée","ouvertement","paysans","poétique","prolongée","prometteurs","provincial","reconnus","regretter","rentre","ris","rosé","réparer","réservoir","résistant","s'attendent","segmentation","sentiers","substitution","suffirait","tournoi","treize","triplé","téléspectateurs","vaudra","écologistes","Américain","Bourgogne","Burundi","Celles","Conservatoire","Cuba","Entreprendre","IV","Nations","Normalement","Saint-Jacques","Schaerbeek","Signalons","accélérée","adaptations","africaine","agressif","aidé","autonomes","chauds","civils","combustible","comités","concurrentes","consentir","d'expliquer","d'exportation","dense","découvrent","défensive","dévolu","engendré","envol","forgé","fréquentation","graisse","impliqués","imposent","inédit","l'Australie","l'anonymat","l'imaginaire","l'éclairage","laps","loisir","molécules","négoce","pape","pensait","pensez","phrases","promenades","ranger","relevant","s'inscrivent","serre","sollicité","tentés","thermique","transparente","Acheter","CBF","Canon","Donaldson","El","Espace","JP","Libre","alliés","brosse","coefficient","comparaisons","continueront","coupables","cultive","cycliste","d'Allemagne","datent","débouche","déco","entreprendre","gourmandise","gouvernementales","griffe","guichet","indéterminée","instructions","interrogé","l'adhésion","l'assemblage","l'audience","l'océan","l'énorme","mobiliser","pavé","plantation","ports","puisqu'ils","quatorze","rationalisation","renouer","ridicule","réglage","réparti","surmonter","tarification","tuyaux","uniques","Cisco","EDI","Fiscalité","Gallimard","Gandois","Jos","Mosane","Nokia","Usinor","Vande","Varsovie","abandonne","achevé","ajoutent","allusion","annexes","autoroute","brusquement","console","cran","d'Ostende","d'adopter","d'assemblage","d'exception","d'innovation","déceler","dérégulation","favorites","gourou","l'Argentine","l'allure","l'enjeu","made","moelleux","obtiennent","panoplie","photographique","qu'ailleurs","qu'entre","redevable","remonté","remplie","rentabiliser","scandinaves","serions","sud-africaines","suicide","trim","télétexte","vierge","vigoureuse","visages","voyant","électoral","Antonio","BATC","Gare","Institut","KNP","Menus","Renard","Road","Rops","Sauvage","abordable","accueillent","adresser","approprié","brillantes","cachet","casino","catastrophes","chargées","ciment","content","courantes","d'acquisition","d'avant","d'heures","delta","dysfonctionnements","filtres","gravures","innombrables","juriste","jusqu'alors","lampes","mannequin","parcouru","polonais","préférons","prévenus","pérennité","quai","quittent","radiation","refait","regrouper","revenant","réduisent","résisté","réticences","s'inquiéter","solo","temporaires","Ah","CAC","Carl","GIA","Gassée","House","Idées","Jersey","Lufthansa","Pacifique","Partenaires","Second","Travel","accompagnés","alimenter","augmentée","aval","battus","bordeaux","cohésion","compensée","complices","compliquée","conclus","construits","cuvée","défaite","flacons","fuir","habitué","historiquement","hit-parade","improbable","incluant","informés","j'y","l'argument","l'indemnité","l'originalité","l'ouest","lires","lumineuse","managing","mariages","morosité","mythes","n'offre","neuves","offrait","opérationnels","peuples","portés","prolongé","précaution","puisqu'on","pédophilie","quasi-totalité","retire","rigoureux","ristourne","réalisable","réalistes","s'entendre","tardive","vies","vital","von","Argentine","DVD","Distrigaz","Forem","Formule","In","Leys","Maria","Master","Monaco","PHOTOS","Reserve","Verbaet","Witkowska","Z","accéléré","acquises","antérieurement","avancent","avoué","chapitres","circuler","consacrent","corrélation","d'affirmer","d'élargir","d'équipements","diminuent","espérant","fonder","fusionner","für","general","heurte","hésiter","jouissent","l'aile","l'institut","laissez","parution","pleins","profils","prototypes","quiconque","rail","rareté","respire","rythmes","réservées","sauvé","scandales","superbement","trimestriel","AAA","Beethoven","Ci-contre","Commerzbank","DGX","Hormis","Inusop","PowerPC","Prins","Public","Salomon","accents","accordées","actrice","annuelles","antérieures","brille","cachés","combinaisons","contraints","crucial","d'enseignement","demandée","entoure","fascinant","fragiles","globe","haussier","hérité","inquiétudes","juif","manquera","massifs","mépris","nippons","odeurs","ombre","paille","paraissait","perceptible","pertinence","proportionnelle","proportionnellement","prévoyant","psychologue","recevront","revendu","reviendra","risquer","saga","serez","sous-jacent","techniquement","températures","tiens","échantillon","CE","CFE","Coca","Directeur","Heysel","Ier","Jacob","Paix","Samedi","Simple","Tobback","Trends","UEM","Vinck","adoptent","centimes","centraux","collants","communisme","consistait","d'ail","d'ensemble","d'exposition","devenait","défenseurs","dépendant","dévoiler","enregistrent","fameuses","gagnants","indicatif","indifférent","intégrante","intéressent","inversement","kW","l'allemand","l'apéritif","l'important","n'échappe","objective","occupants","pains","perle","phares","planétaire","pointes","poulet","prioritaire","précoce","quo","recommencer","représentées","roule","réputée","rétablissement","sonne","spirale","tendent","tiroirs","traitent","trouvaient","validité","variantes","verticale","vole","église","équilibré","Armand","Carlier","Distribution","Fred","L'ouvrage","Monceau-Zolder","Officiellement","Passer","Patriotique","Réflexions","SGB","Shop","Toshiba","absorbe","acceptée","agissent","aimer","apparences","appliqués","ascension","attaques","attitudes","autorisés","blessures","cathédrale","contrer","d'essence","d'examiner","divergences","dix-huit","détachement","formuler","implications","l'effondrement","l'habitacle","libérales","loup","magnifiques","mainframes","miroirs","offensive","ondes","organe","porte-monnaie","professionnalisme","préside","recueilli","rentes","report","retards","revendeurs","rez-de-chaussée","roulement","ruines","rénové","sienne","stimule","toilette","utilitaires","voudraient","Audiofina","BRUXELLES","Bilan","Colin","Continental","Dommage","Ferrari","Laurence","MBA","UPS","appelées","artisanale","automobilistes","berceau","bloquer","brancher","classés","communique","continentale","convivial","d'analyser","d'intérieur","détaillé","détaillée","estimait","examiné","facto","glissement","grise","incident","kit","know-how","l'Occident","légitimité","nommer","organisent","photographes","pionniers","profitera","prometteuse","prudemment","radical","rater","siens","simplification","small","spécificités","suspects","tarder","urbains","veine","vinaigrette","vingt-cinq","écoulé","Chiffre","Dimanche","Havas","Metropolitan","Presque","Viennent","Village","agréables","agréé","animations","caviar","chauffeurs","concepteur","conteneurs","coureurs","d'Union","d'envisager","d'opinion","deuil","débarque","déclenche","examens","executive","formalités","fragilité","gourmande","gratuites","inflationnistes","introductions","isolés","justifient","l'abbaye","l'épargnant","lacune","limitant","marin","new-yorkais","occupant","parvenue","relie","remplacés","reportage","représentatif","réparations","s'attaquer","sonores","soutiennent","tirés","trouvons","turbodiesel","violent","Adresse","Casino","Delaunois","Fafer","Industrie","Inversement","Longtemps","Manche","Med","Montréal","One","Partout","Pérou","Robe","Rose","Sachez","Spadel","Stefaan","Véritable","auto","bars","bizarre","blonde","cercles","chapelle","compté","conformité","créatifs","créatrice","culpabilité","d'Apple","d'exercer","d'échapper","d'étendre","désordre","exécuter","gammes","gastronomique","guider","générique","homologue","imagination","immanquablement","impressions","impératifs","initiés","irlandais","jouet","kilo","l'abonnement","l'habitation","l'inventeur","l'élection","loger","marasme","mutuelle","médiocres","n'est-elle","nobles","odeur","oignons","opinions","plates-formes","pratiquant","provocation","prédit","retomber","routine","répandu","répandue","sous-sol","soviétique","subtil","sélectionné","tabou","tache","taxer","touchés","usées","élaborée","émane","ABN-Amro","BEI","Beecham","COIB","Cirque","Claire","Direct","Interim","Louvre","Mazda","Qu'on","Recherche","Sablon","SmithKline","a-t-on","actualité","ambiant","antenne","associe","blessés","boules","bourgeois","brun","caves","chirurgie","classer","colis","consommé","d'accompagnement","d'affichage","d'aménagement","d'enregistrer","d'école","demandeur","déploie","désespoir","emplacement","ennemi","envisageable","essayons","foyers","frites","gouttes","illusion","j","jugées","l'amateur","l'axe","l'opposé","larmes","lingerie","longévité","lorsqu'elles","manches","manne","mette","muni","navires","pharmacien","potentielle","procurent","protégée","préoccupe","rigoureuse","soyez","strangle","tentant","transmet","trouvée","variant","viser","volés","Ariane","BREF","Bart","Bruxelles-Capitale","Cola"]
    en_common_words = ["the","of","and","to","a","in","for","is","on","that","by","this","with","i","you","This","it","not","or","be","are","from","at","as","your","all","have","new","more","an","was","we","will","home","can","us","about","if","page","my","has","search","free","but","our","one","other","do","no","information","time","they","site","he","up","may","what","which","their","news","out","use","any","there","see","only","so","his","when","contact","here","business","who","web","also","now","help","get","pm","view","online","c","e","first","am","been","would","how","were","me","s","services","some","these","click","its","like","service","x","than","find","price","date","back","top","people","had","list","name","just","over","state","year","day","into","email","two","health","n","world","re","next","used","go","b","work","last","most","products","music","buy","data","make","them","should","product","system","post","her","city","t","add","policy","number","such","please","available","copyright","support","message","after","best","software","then","jan","good","video","well","d","where","info","rights","public","books","high","school","through","m","each","links","she","review","years","order","very","privacy","book","items","company","r","read","group","sex","need","many","user","said","de","does","set","under","general","research","university","january","mail","full","map","reviews","program","life","know","games","way","days","management","p","part","could","great","united","hotel","real","f","item","international","center","ebay","must","store","travel","comments","made","development","report","off","member","details","line","terms","before","hotels","did","send","right","type","because","local","those","using","results","office","education","national","car","design","take","posted","internet","address","community","within","states","area","want","phone","dvd","shipping","reserved","subject","between","forum","family","l","long","based","w","code","show","o","even","black","check","special","prices","website","index","being","women","much","sign","file","link","open","today","technology","south","case","project","same","pages","uk","version","section","own","found","sports","house","related","security","both","g","county","american","photo","game","members","power","while","care","network","down","computer","systems","three","total","place","end","following","download","h","him","without","per","access","think","north","resources","current","posts","big","media","law","control","water","history","pictures","size","art","personal","since","including","guide","shop","directory","board","location","change","white","text","small","rating","rate","government","children","during","usa","return","students","v","shopping","account","times","sites","level","digital","profile","previous","form","events","love","old","john","main","call","hours","image","department","title","description","non","k","y","insurance","another","why","shall","property","class","cd","still","money","quality","every","listing","content","country","private","little","visit","save","tools","low","reply","customer","december","compare","movies","include","college","value","article","york","man","card","jobs","provide","j","food","source","author","different","press","u","learn","sale","around","print","course","job","canada","process","teen","room","stock","training","too","credit","point","join","science","men","categories","advanced","west","sales","look","english","left","team","estate","box","conditions","select","windows","photos","gay","thread","week","category","note","live","large","gallery","table","register","however","june","october","november","market","library","really","action","start","series","model","features","air","industry","plan","human","provided","tv","yes","required","second","hot","accessories","cost","movie","forums","march","la","september","better","say","questions","july","yahoo","going","medical","test","friend","come","dec","server","pc","study","application","cart","staff","articles","san","feedback","again","play","looking","issues","april","never","users","complete","street","topic","comment","financial","things","working","against","standard","tax","person","below","mobile","less","got","blog","party","payment","equipment","login","student","let","programs","offers","legal","above","recent","park","stores","side","act","problem","red","give","memory","performance","social","q","august","quote","language","story","sell","options","experience","rates","create","key","body","young","america","important","field","few","east","paper","single","ii","age","activities","club","example","girls","additional","password","z","latest","something","road","gift","question","changes","night","ca","hard","texas","oct","pay","four","poker","status","browse","issue","range","building","seller","court","february","always","result","audio","light","write","war","nov","offer","blue","groups","al","easy","given","files","event","release","analysis","request","fax","china","making","picture","needs","possible","might","professional","yet","month","major","star","areas","future","space","committee","hand","sun","cards","problems","london","washington","meeting","rss","become","interest","id","child","keep","enter","california","porn","share","similar","garden","schools","million","added","reference","companies","listed","baby","learning","energy","run","delivery","net","popular","term","film","stories","put","computers","journal","reports","co","try","welcome","central","images","president","notice","god","original","head","radio","until","cell","color","self","council","away","includes","track","australia","discussion","archive","once","others","entertainment","agreement","format","least","society","months","log","safety","friends","sure","faq","trade","edition","cars","messages","marketing","tell","further","updated","association","able","having","provides","david","fun","already","green","studies","close","common","drive","specific","several","gold","feb","living","sep","collection","called","short","arts","lot","ask","display","limited","powered","solutions","means","director","daily","beach","past","natural","whether","due","et","electronics","five","upon","period","planning","database","says","official","weather","mar","land","average","done","technical","window","france","pro","region","island","record","direct","microsoft","conference","environment","records","st","district","calendar","costs","style","url","front","statement","update","parts","aug","ever","downloads","early","miles","sound","resource","present","applications","either","ago","document","word","works","material","bill","apr","written","talk","federal","hosting","rules","final","adult","tickets","thing","centre","requirements","via","cheap","nude","kids","finance","true","minutes","else","mark","third","rock","gifts","europe","reading","topics","bad","individual","tips","plus","auto","cover","usually","edit","together","videos","percent","fast","function","fact","unit","getting","global","tech","meet","far","economic","en","player","projects","lyrics","often","subscribe","submit","germany","amount","watch","included","feel","though","bank","risk","thanks","everything","deals","various","words","linux","jul","production","commercial","james","weight","town","heart","advertising","received","choose","treatment","newsletter","archives","points","knowledge","magazine","error","camera","jun","girl","currently","construction","toys","registered","clear","golf","receive","domain","methods","chapter","makes","protection","policies","loan","wide","beauty","manager","india","position","taken","sort","listings","models","michael","known","half","cases","step","engineering","florida","simple","quick","none","wireless","license","paul","friday","lake","whole","annual","published","later","basic","sony","shows","corporate","google","church","method","purchase","customers","active","response","practice","hardware","figure","materials","fire","holiday","chat","enough","designed","along","among","death","writing","speed","html","countries","loss","face","brand","discount","higher","effects","created","remember","standards","oil","bit","yellow","political","increase","advertise","kingdom","base","near","environmental","thought","stuff","french","storage","oh","japan","doing","loans","shoes","entry","stay","nature","orders","availability","africa","summary","turn","mean","growth","notes","agency","king","monday","european","activity","copy","although","drug","pics","western","income","force","cash","employment","overall","bay","river","commission","ad","package","contents","seen","players","engine","port","album","regional","stop","supplies","started","administration","bar","institute","views","plans","double","dog","build","screen","exchange","types","soon","sponsored","lines","electronic","continue","across","benefits","needed","season","apply","someone","held","ny","anything","printer","condition","effective","believe","organization","effect","asked","eur","mind","sunday","selection","casino","pdf","lost","tour","menu","volume","cross","anyone","mortgage","hope","silver","corporation","wish","inside","solution","mature","role","rather","weeks","addition","came","supply","nothing","certain","usr","executive","running","lower","necessary","union","jewelry","according","dc","clothing","mon","com","particular","fine","names","robert","homepage","hour","gas","skills","six","bush","islands","advice","career","military","rental","decision","leave","british","teens","pre","huge","sat","woman","facilities","zip","bid","kind","sellers","middle","move","cable","opportunities","taking","values","division","coming","tuesday","object","lesbian","appropriate","machine","logo","length","actually","nice","score","statistics","client","ok","returns","capital","follow","sample","investment","sent","shown","saturday","christmas","england","culture","band","flash","ms","lead","george","choice","went","starting","registration","fri","thursday","courses","consumer","hi","airport","foreign","artist","outside","furniture","levels","channel","letter","mode","phones","ideas","wednesday","structure","fund","summer","allow","degree","contract","button","releases","wed","homes","super","male","matter","custom","virginia","almost","took","located","multiple","asian","distribution","editor","inn","industrial","cause","potential","song","cnet","ltd","los","hp","focus","late","fall","featured","idea","rooms","female","responsible","inc","communications","win","associated","thomas","primary","cancer","numbers","reason","tool","browser","spring","foundation","answer","voice","eg","friendly","schedule","documents","communication","purpose","feature","bed","comes","police","everyone","independent","ip","approach","cameras","brown","physical","operating","hill","maps","medicine","deal","hold","ratings","chicago","forms","glass","happy","tue","smith","wanted","developed","thank","safe","unique","survey","prior","telephone","sport","ready","feed","animal","sources","mexico","population","pa","regular","secure","navigation","operations","therefore","ass","simply","evidence","station","christian","round","paypal","favorite","understand","option","master","valley","recently","probably","thu","rentals","sea","built","publications","blood","cut","worldwide","improve","connection","publisher","hall","larger","anti","networks","earth","parents","nokia","impact","transfer","introduction","kitchen","strong","tel","carolina","wedding","properties","hospital","ground","overview","ship","accommodation","owners","disease","tx","excellent","paid","italy","perfect","hair","opportunity","kit","classic","basis","command","cities","william","express","anal","award","distance","tree","peter","assessment","ensure","thus","wall","ie","involved","el","extra","especially","interface","pussy","partners","budget","rated","guides","success","maximum","ma","operation","existing","quite","selected","boy","amazon","patients","restaurants","beautiful","warning","wine","locations","horse","vote","forward","flowers","stars","significant","lists","technologies","owner","retail","animals","useful","directly","manufacturer","ways","est","son","providing","rule","mac","housing","takes","iii","gmt","bring","catalog","searches","max","trying","mother","authority","considered","told","xml","traffic","programme","joined","input","strategy","feet","agent","valid","bin","modern","senior","ireland","sexy","teaching","door","grand","testing","trial","charge","units","instead","canadian","cool","normal","wrote","enterprise","ships","entire","educational","md","leading","metal","positive","fl","fitness","chinese","opinion","mb","asia","football","abstract","uses","output","funds","mr","greater","likely","develop","employees","artists","alternative","processing","responsibility","resolution","java","guest","seems","publication","pass","relations","trust","van","contains","session","multi","photography","republic","fees","components","vacation","century","academic","assistance","completed","skin","graphics","indian","prev","ads","mary","il","expected","ring","grade","dating","pacific","mountain","organizations","pop","filter","mailing","vehicle","longer","consider","int","northern","behind","panel","floor","german","buying","match","proposed","default","require","iraq","boys","outdoor","deep","morning","otherwise","allows","rest","protein","plant","reported","hit","transportation","mm","pool","mini","politics","partner","disclaimer","authors","boards","faculty","parties","fish","membership","mission","eye","string","sense","modified","pack","released","stage","internal","goods","recommended","born","unless","richard","detailed","japanese","race","approved","background","target","except","character","usb","maintenance","ability","maybe","functions","ed","moving","brands","places","php","pretty","trademarks","phentermine","spain","southern","yourself","etc","winter","rape","battery","youth","pressure","submitted","boston","incest","debt","keywords","medium","television","interested","core","break","purposes","throughout","sets","dance","wood","msn","itself","defined","papers","playing","awards","fee","studio","reader","virtual","device","established","answers","rent","las","remote","dark","programming","external","apple","le","regarding","instructions","min","offered","theory","enjoy","remove","aid","surface","minimum","visual","host","variety","teachers","isbn","martin","manual","block","subjects","agents","increased","repair","fair","civil","steel","understanding","songs","fixed","wrong","beginning","hands","associates","finally","az","updates","desktop","classes","paris","ohio","gets","sector","capacity","requires","jersey","un","fat","fully","father","electric","saw","instruments","quotes","officer","driver","businesses","dead","respect","unknown","specified","restaurant","mike","trip","pst","worth","mi","procedures","poor","teacher","xxx","eyes","relationship","workers","farm","fucking","georgia","peace","traditional","campus","tom","showing","creative","coast","benefit","progress","funding","devices","lord","grant","sub","agree","fiction","hear","sometimes","watches","careers","beyond","goes","families","led","museum","themselves","fan","transport","interesting","blogs","wife","evaluation","accepted","former","implementation","ten","hits","zone","complex","th","cat","galleries","references","die","presented","jack","flat","flow","agencies","literature","respective","parent","spanish","michigan","columbia","setting","dr","scale","stand","economy","highest","helpful","monthly","critical","frame","musical","definition","secretary","angeles","networking","path","australian","employee","chief","gives","kb","bottom","magazines","packages","detail","francisco","laws","changed","pet","heard","begin","individuals","colorado","royal","clean","switch","russian","largest","african","guy","titles","relevant","guidelines","justice","connect","bible","dev","cup","basket","applied","weekly","vol","installation","described","demand","pp","suite","vegas","na","square","chris","attention","advance","skip","diet","army","auction","gear","lee","os","difference","allowed","correct","charles","nation","selling","lots","piece","sheet","firm","seven","older","illinois","regulations","elements","species","jump","cells","module","resort","facility","random","pricing","dvds","certificate","minister","motion","looks","fashion","directions","visitors","documentation","monitor","trading","forest","calls","whose","coverage","couple","giving","chance","vision","ball","ending","clients","actions","listen","discuss","accept","automotive","naked","goal","successful","sold","wind","communities","clinical","situation","sciences","markets","lowest","highly","publishing","appear","emergency","developing","lives","currency","leather","determine","milf","temperature","palm","announcements","patient","actual","historical","stone","bob","commerce","ringtones","perhaps","persons","difficult","scientific","satellite","fit","tests","village","accounts","amateur","ex","met","pain","xbox","particularly","factors","coffee","www","settings","cum","buyer","cultural","steve","easily","oral","ford","poster","edge","functional","root","au","fi","closed","holidays","ice","pink","zealand","balance","monitoring","graduate","replies","shot","nc","architecture","initial","label","thinking","scott","llc","sec","recommend","canon","hardcore","league","waste","minute","bus","provider","optional","dictionary","cold","accounting","manufacturing","sections","chair","fishing","effort","phase","fields","bag","fantasy","po","letters","motor","va","professor","context","install","shirt","apparel","generally","continued","foot","mass","crime","count","breast","techniques","ibm","rd","johnson","sc","quickly","dollars","websites","religion","claim","driving","permission","surgery","patch","heat","wild","measures","generation","kansas","miss","chemical","doctor","task","reduce","brought","himself","nor","component","enable","exercise","bug","santa","mid","guarantee","leader","diamond","israel","se","processes","soft","servers","alone","meetings","seconds","jones","arizona","keyword","interests","flight","congress","fuel","username","walk","fuck","produced","italian","paperback","classifieds","wait","supported","pocket","saint","rose","freedom","argument","competition","creating","jim","drugs","joint","premium","providers","fresh","characters","attorney","upgrade","di","factor","growing","thousands","km","stream","apartments","pick","hearing","eastern","auctions","therapy","entries","dates","generated","signed","upper","administrative","serious","prime","samsung","limit","began","louis","steps","errors","shops","bondage","del","efforts","informed","ga","ac","thoughts","creek","ft","worked","quantity","urban","practices","sorted","reporting","essential","myself","tours","platform","load","affiliate","labor","immediately","admin","nursing","defense","machines","designated","tags","heavy","covered","recovery","joe","guys","integrated","configuration","cock","merchant","comprehensive","expert","universal","protect","drop","solid","cds","presentation","languages","became","orange","compliance","vehicles","prevent","theme","rich","im","campaign","marine","improvement","vs","guitar","finding","pennsylvania","examples","ipod","saying","spirit","ar","claims","porno","challenge","motorola","acceptance","strategies","mo","seem","affairs","touch","intended","towards","sa","goals","hire","election","suggest","branch","charges","serve","affiliates","reasons","magic","mount","smart","talking","gave","ones","latin","multimedia","xp","tits","avoid","certified","manage","corner","rank","computing","oregon","element","birth","virus","abuse","interactive","requests","separate","quarter","procedure","leadership","tables","define","racing","religious","facts","breakfast","kong","column","plants","faith","chain","developer","identify","avenue","missing","died","approximately","domestic","sitemap","recommendations","moved","houston","reach","comparison","mental","viewed","moment","extended","sequence","inch","attack","sorry","centers","opening","damage","lab","reserve","recipes","cvs","gamma","plastic","produce","snow","placed","truth","counter","failure","follows","eu","weekend","dollar","camp","ontario","automatically","des","minnesota","films","bridge","native","fill","williams","movement","printing","baseball","owned","approval","draft","chart","played","contacts","cc","jesus","readers","clubs","lcd","wa","jackson","equal","adventure","matching","offering","shirts","profit","leaders","posters","institutions","assistant","variable","ave","dj","advertisement","expect","parking","headlines","yesterday","compared","determined","wholesale","workshop","russia","gone","codes","kinds","extension","seattle","statements","golden","completely","teams","fort","cm","wi","lighting","senate","forces","funny","brother","gene","turned","portable","tried","electrical","applicable","disc","returned","pattern","ct","hentai","boat","named","theatre","laser","earlier","manufacturers","sponsor","classical","icon","warranty","dedicated","indiana","direction","harry","basketball","objects","ends","delete","evening","assembly","nuclear","taxes","mouse","signal","criminal","issued","brain","sexual","wisconsin","powerful","dream","obtained","false","da","cast","flower","felt","personnel","passed","supplied","identified","falls","pic","soul","aids","opinions","promote","stated","stats","hawaii","professionals","appears","carry","flag","decided","nj","covers","hr","em","advantage","hello","designs","maintain","tourism","priority","newsletters","adults","clips","savings","iv","graphic","atom","payments","rw","estimated","binding","brief","ended","winning","eight","anonymous","iron","straight","script","served","wants","miscellaneous","prepared","void","dining","alert","integration","atlanta","dakota","tag","interview","mix","framework","disk","installed","queen","vhs","credits","clearly","fix","handle","sweet","desk","criteria","pubmed","dave","massachusetts","diego","hong","vice","associate","ne","truck","behavior","enlarge","ray","frequently","revenue","measure","changing","votes","du","duty","looked","discussions","bear","gain","festival","laboratory","ocean","flights","experts","signs","lack","depth","iowa","whatever","logged","laptop","vintage","train","exactly","dry","explore","maryland","spa","concept","nearly","eligible","checkout","reality","forgot","handling","origin","knew","gaming","feeds","billion","destination","scotland","faster","intelligence","dallas","bought","con","ups","nations","route","followed","specifications","broken","tripadvisor","frank","alaska","zoom","blow","battle","residential","anime","speak","decisions","industries","protocol","query","clip","partnership","editorial","nt","expression","es","equity","provisions","speech","wire","principles","suggestions","rural","shared","sounds","replacement","tape","strategic","judge","spam","economics","acid","bytes","cent","forced","compatible","fight","apartment","height","null","zero","speaker","filed","gb","netherlands","obtain","bc","consulting","recreation","offices","designer","remain","managed","pr","failed","marriage","roll","korea","banks","fr","participants","secret","bath","aa","kelly","leads","negative","austin","favorites","toronto","theater","springs","missouri","andrew","var","perform","healthy","translation","estimates","font","assets","injury","mt","joseph","ministry","drivers","lawyer","figures","married","protected","proposal","sharing","philadelphia","portal","waiting","birthday","beta","fail","gratis","banking","officials","brian","toward","won","slightly","assist","conduct","contained","lingerie","shemale","legislation","calling","parameters","jazz","serving","bags","profiles","miami","comics","matters","houses","doc","postal","relationships","tennessee","wear","controls","breaking","combined","ultimate","wales","representative","frequency","introduced","minor","finish","departments","residents","noted","displayed","mom","reduced","physics","rare","spent","performed","extreme","samples","davis","daniel","bars","reviewed","row","oz","forecast","removed","helps","singles","administrator","cycle","amounts","contain","accuracy","dual","rise","usd","sleep","mg","bird","pharmacy","brazil","creation","static","scene","hunter","addresses","lady","crystal","famous","writer","chairman","violence","fans","oklahoma","speakers","drink","academy","dynamic","gender","eat","permanent","agriculture","dell","cleaning","constitutes","portfolio","practical","delivered","collectibles","infrastructure","exclusive","seat","concerns","colour","vendor","originally","intel","utilities","philosophy","regulation","officers","reduction","aim","bids","referred","supports","nutrition","recording","regions","junior","toll","les","cape","ann","rings","meaning","tip","secondary","wonderful","mine","ladies","henry","ticket","announced","guess","agreed","prevention","whom","ski","soccer","math","import","posting","presence","instant","mentioned","automatic","healthcare","viewing","maintained","ch","increasing","majority","connected","christ","dan","dogs","sd","directors","aspects","austria","ahead","moon","participation","scheme","utility","preview","fly","manner","matrix","containing","combination","devel","amendment","despite","strength","guaranteed","turkey","libraries","proper","distributed","degrees","singapore","enterprises","delta","fear","seeking","inches","phoenix","rs","convention","shares","principal","daughter","standing","voyeur","comfort","colors","wars","cisco","ordering","kept","alpha","appeal","cruise","bonus","certification","previously","hey","bookmark","buildings","specials","beat","disney","household","batteries","adobe","smoking","bbc","becomes","drives","arms","alabama","tea","improved","trees","avg","achieve","positions","dress","subscription","dealer","contemporary","sky","utah","nearby","rom","carried","happen","exposure","panasonic","hide","permalink","signature","gambling","refer","miller","provision","outdoors","clothes","caused","luxury","babes","frames","viagra","certainly","indeed","newspaper","toy","circuit","layer","printed","slow","removal","easier","src","liability","trademark","hip","printers","faqs","nine","adding","kentucky","mostly","eric","spot","taylor","trackback","prints","spend","factory","interior","revised","grow","americans","optical","promotion","relative","amazing","clock","dot","hiv","identity","suites","conversion","feeling","hidden","reasonable","victoria","serial","relief","revision","broadband","influence","ratio","pda","importance","rain","onto","dsl","planet","webmaster","copies","recipe","zum","permit","seeing","proof","dna","diff","tennis","bass","prescription","bedroom","empty","instance","hole","pets","ride","licensed","orlando","specifically","tim","bureau","maine","sql","represent","conservation","pair","ideal","specs","recorded","don","pieces","finished","parks","dinner","lawyers","sydney","stress","cream","ss","runs","trends","yeah","discover","sexo","ap","patterns","boxes","louisiana","hills","javascript","fourth","nm","advisor","mn","marketplace","nd","evil","aware","wilson","shape","evolution","irish","certificates","objectives","stations","suggested","gps","op","remains","acc","greatest","firms","concerned","euro","operator","structures","generic","encyclopedia","usage","cap","ink","charts","continuing","mixed","census","interracial","peak","tn","competitive","exist","wheel","transit","dick","suppliers","salt","compact","poetry","lights","tracking","angel","bell","keeping","preparation","attempt","receiving","matches","accordance","width","noise","engines","forget","array","discussed","accurate","stephen","elizabeth","climate","reservations","pin","playstation","alcohol","greek","instruction","managing","annotation","sister","raw","differences","walking","explain","smaller","newest","establish","gnu","happened","expressed","jeff","extent","sharp","lesbians","ben","lane","paragraph","kill","mathematics","aol","compensation","ce","export","managers","aircraft","modules","sweden","conflict","conducted","versions","employer","occur","percentage","knows","mississippi","describe","concern","backup","requested","citizens","connecticut","heritage","personals","immediate","holding","trouble","spread","coach","kevin","agricultural","expand","supporting","audience","assigned","jordan","collections","ages","participate","plug","specialist","cook","affect","virgin","experienced","investigation","raised","hat","institution","directed","dealers","searching","sporting","helping","perl","affected","lib","bike","totally","plate","expenses","indicate","blonde","ab","proceedings","favourite","transmission","anderson","utc","characteristics","der","lose","organic","seek","experiences","albums","cheats","extremely","verzeichnis","contracts","guests","hosted","diseases","concerning","developers","equivalent","chemistry","tony","neighborhood","nevada","kits","thailand","variables","agenda","anyway","continues","tracks","advisory","cam","curriculum","logic","template","prince","circle","soil","grants","anywhere","psychology","responses","atlantic","wet","circumstances","edward","investor","identification","ram","leaving","wildlife","appliances","matt","elementary","cooking","speaking","sponsors","fox","unlimited","respond","sizes","plain","exit","entered","iran","arm","keys","launch","wave","checking","costa","belgium","printable","holy","acts","guidance","mesh","trail","enforcement","symbol","crafts","highway","buddy","hardcover","observed","dean","setup","poll","booking","glossary","fiscal","celebrity","styles","denver","unix","filled","bond","channels","ericsson","appendix","notify","blues","chocolate","pub","portion","scope","hampshire","supplier","cables","cotton","bluetooth","controlled","requirement","authorities","biology","dental","killed","border","ancient","debate","representatives","starts","pregnancy","causes","arkansas","biography","leisure","attractions","learned","transactions","notebook","explorer","historic","attached","opened","tm","husband","disabled","authorized","crazy","upcoming","britain","concert","retirement","scores","financing","efficiency","sp","comedy","adopted","efficient","weblog","linear","commitment","specialty","bears","jean","hop","carrier","edited","constant","visa","mouth","jewish","meter","linked","portland","interviews","concepts","nh","gun","reflect","pure","deliver","wonder","hell","lessons","fruit","begins","qualified","reform","lens","alerts","treated","discovery","draw","mysql","classified","relating","assume","confidence","alliance","fm","confirm","warm","neither","lewis","howard","offline","leaves","engineer","lifestyle","consistent","replace","clearance","connections","inventory","converter","suck","organisation","babe","checks","reached","becoming","blowjob","safari","objective","indicated","sugar","crew","legs","sam","stick","securities","allen","pdt","relation","enabled","genre","slide","montana","volunteer","tested","rear","democratic","enhance","switzerland","exact","bound","parameter","adapter","processor","node","formal","dimensions","contribute","lock","hockey","storm","micro","colleges","laptops","mile","showed","challenges","editors","mens","threads","bowl","supreme","brothers","recognition","presents","ref","tank","submission","dolls","estimate","encourage","navy","kid","regulatory","inspection","consumers","cancel","limits","territory","transaction","manchester","weapons","paint","delay","pilot","outlet","contributions","continuous","db","czech","resulting","cambridge","initiative","novel","pan","execution","disability","increases","ultra","winner","idaho","contractor","ph","episode","examination","potter","dish","plays","bulletin","ia","pt","indicates","modify","oxford","adam","truly","epinions","painting","committed","extensive","affordable","universe","candidate","databases","patent","slot","psp","outstanding","ha","eating","perspective","planned","watching","lodge","messenger","mirror","tournament","consideration","ds","discounts","sterling","sessions","kernel","boobs","stocks","buyers","journals","gray","catalogue","ea","jennifer","antonio","charged","broad","taiwan","und","chosen","demo","greece","lg","swiss","sarah","clark","labour","hate","terminal","publishers","nights","behalf","caribbean","liquid","rice","nebraska","loop","salary","reservation","foods","gourmet","guard","properly","orleans","saving","nfl","remaining","empire","resume","twenty","newly","raise","prepare","avatar","gary","depending","illegal","expansion","vary","hundreds","rome","arab","lincoln","helped","premier","tomorrow","purchased","milk","decide","consent","drama","visiting","performing","downtown","keyboard","contest","collected","nw","bands","boot","suitable","ff","absolutely","millions","lunch","dildo","audit","push","chamber","guinea","findings","muscle","featuring","iso","implement","clicking","scheduled","polls","typical","tower","yours","sum","misc","calculator","significantly","chicken","temporary","attend","shower","alan","sending","jason","tonight","dear","sufficient","holdem","shell","province","catholic","oak","vat","awareness","vancouver","governor","beer","seemed","contribution","measurement","swimming","spyware","formula","constitution","packaging","solar","jose","catch","jane","pakistan","ps","reliable","consultation","northwest","sir","doubt","earn","finder","unable","periods","classroom","tasks","democracy","attacks","kim","wallpaper","merchandise","const","resistance","doors","symptoms","resorts","biggest","memorial","visitor","twin","forth","insert","baltimore","gateway","ky","dont","alumni","drawing","candidates","charlotte","ordered","biological","fighting","transition","happens","preferences","spy","romance","instrument","bruce","split","themes","powers","heaven","br","bits","pregnant","twice","classification","focused","egypt","physician","hollywood","bargain","wikipedia","cellular","norway","vermont","asking","blocks","normally","lo","spiritual","hunting","diabetes","suit","ml","shift","chip","res","sit","bodies","photographs","cutting","wow","simon","writers","marks","flexible","loved","favourites","mapping","numerous","relatively","birds","satisfaction","represents","char","indexed","pittsburgh","superior","preferred","saved","paying","cartoon","shots","intellectual","moore","granted","choices","carbon","spending","comfortable","magnetic","interaction","listening","effectively","registry","crisis","outlook","massive","denmark","employed","bright","treat","header","cs","poverty","formed","piano","echo","que","grid","sheets","patrick","experimental","puerto","revolution","consolidation","displays","plasma","allowing","earnings","voip","mystery","landscape","dependent","mechanical","journey","delaware","bidding","consultants","risks","banner","applicant","charter","fig","barbara","cooperation","counties","acquisition","ports","implemented","sf","directories","recognized","dreams","blogger","notification","kg","licensing","stands","teach","occurred","textbooks","rapid","pull","hairy","diversity","cleveland","ut","reverse","deposit","seminar","investments","latina","nasa","wheels","sexcam","specify","accessibility","dutch","sensitive","templates","formats","tab","depends","boots","holds","router","concrete","si","editing","poland","folder","womens","css","completion","upload","pulse","universities","technique","contractors","milfhunter","voting","courts","notices","subscriptions","calculate","mc","detroit","alexander","broadcast","converted","metro","toshiba","anniversary","improvements","strip","specification","pearl","accident","nick","accessible","accessory","resident","plot","qty","possibly","airline","typically","representation","regard","pump","exists","arrangements","smooth","conferences","uniprotkb","beastiality","strike","consumption","birmingham","flashing","lp","narrow","afternoon","threat","surveys","sitting","putting","consultant","controller","ownership","committees","penis","legislative","researchers","vietnam","trailer","anne","castle","gardens","missed","malaysia","unsubscribe","antique","labels","willing","bio","molecular","upskirt","acting","heads","stored","exam","logos","residence","attorneys","milfs","antiques","density","hundred","ryan","operators","strange","sustainable","philippines","statistical","beds","breasts","mention","innovation","pcs","employers","grey","parallel","honda","amended","operate","bills","bold","bathroom","stable","opera","definitions","von","doctors","lesson","cinema","asset","ag","scan","elections","drinking","blowjobs","reaction","blank","enhanced","entitled","severe","generate","stainless","newspapers","hospitals","vi","deluxe","humor","aged","monitors","exception","lived","duration","bulk","successfully","indonesia","pursuant","sci","fabric","edt","visits","primarily","tight","domains","capabilities","pmid","contrast","recommendation","flying","recruitment","sin","berlin","cute","organized","ba","para","siemens","adoption","improving","cr","expensive","meant","capture","pounds","buffalo","organisations","plane","pg","explained","seed","programmes","desire","expertise","mechanism","camping","ee","jewellery","meets","welfare","peer","caught","eventually","marked","driven","measured","medline","bottle","agreements","considering","innovative","marshall","massage","rubber","conclusion","closing","tampa","thousand","meat","legend","grace","susan","ing","ks","adams","python","monster","alex","bang","villa","bone","columns","disorders","bugs","collaboration","hamilton","detection","ftp","cookies","inner","formation","tutorial","med","engineers","entity","cruises","gate","holder","proposals","moderator","sw","tutorials","settlement","portugal","lawrence","roman","duties","valuable","erotic","tone","collectables","ethics","forever","dragon","busy","captain","fantastic","imagine","brings","heating","leg","neck","hd","wing","governments","purchasing","scripts","abc","stereo","appointed","taste","dealing","commit","tiny","operational","rail","airlines","liberal","livecam","jay","trips","gap","sides","tube","turns","corresponding","descriptions","cache","belt","jacket","determination","animation","oracle","er","matthew","lease","productions","aviation","hobbies","proud","excess","disaster","console","commands","jr","telecommunications","instructor","giant","achieved","injuries","shipped","bestiality","seats","approaches","biz","alarm","voltage","anthony","nintendo","usual","loading","stamps","appeared","franklin","angle","rob","vinyl","highlights","mining","designers","melbourne","ongoing","worst","imaging","betting","scientists","liberty","wyoming","blackjack","argentina","era","convert","possibility","analyst","commissioner","dangerous","garage","exciting","reliability","thongs","gcc","unfortunately","respectively","volunteers","attachment","ringtone","finland","morgan","derived","pleasure","honor","asp","oriented","eagle","desktops","pants","columbus","nurse","prayer","appointment","workshops","hurricane","quiet","luck","postage","producer","represented","mortgages","dial","responsibilities","cheese","comic","carefully","jet","productivity","investors","crown","par","underground","diagnosis","maker","crack","principle","picks","vacations","gang","semester","calculated","cumshot","fetish","applies","casinos","appearance","smoke","apache","filters","incorporated","nv","craft","cake","notebooks","apart","fellow","blind","lounge","mad","algorithm","semi","coins","andy","gross","strongly","cafe","valentine","hilton","ken","proteins","horror","su","exp","familiar","capable","douglas","debian","till","involving","pen","investing","christopher","admission","epson","shoe","elected","carrying","victory","sand","madison","terrorism","joy","editions","cpu","mainly","ethnic","ran","parliament","actor","finds","seal","situations","fifth","allocated","citizen","vertical","corrections","structural","municipal","describes","prize","sr","occurs","jon","absolute","disabilities","consists","anytime","substance","prohibited","addressed","lies","pipe","soldiers","nr","guardian","lecture","simulation","layout","initiatives","ill","concentration","classics","lbs","lay","interpretation","horses","lol","dirty","deck","wayne","donate","taught","bankruptcy","mp","worker","optimization","alive","temple","substances","prove","discovered","wings","breaks","genetic","restrictions","participating","waters","promise","thin","exhibition","prefer","ridge","cabinet","modem","harris","mph","bringing","sick","dose","evaluate","tiffany","tropical","collect","bet","composition","toyota","streets","nationwide","vector","definitely","shaved","turning","buffer","purple","existence","commentary","larry","limousines","developments","def","immigration","destinations","lets","mutual","pipeline","necessarily","syntax","li","attribute","prison","skill","chairs","nl","everyday","apparently","surrounding","mountains","moves","popularity","inquiry","ethernet","checked","exhibit","throw","trend","sierra","visible","cats","desert","postposted","ya","oldest","rhode","nba","busty","coordinator","obviously","mercury","steven","handbook","greg","navigate","worse","summit","victims","epa","spaces","fundamental","burning","escape","coupons","somewhat","receiver","substantial","tr","progressive","cialis","bb","boats","glance","scottish","championship","arcade","richmond","sacramento","impossible","ron","russell","tells","obvious","fiber","depression","graph","covering","platinum","judgment","bedrooms","talks","filing","foster","modeling","passing","awarded","testimonials","trials","tissue","nz","memorabilia","clinton","masters","bonds","cartridge","alberta","explanation","folk","org","commons","cincinnati","subsection","fraud","electricity","permitted","spectrum","arrival","okay","pottery","emphasis","roger","aspect","workplace","awesome","mexican","confirmed","counts","priced","wallpapers","hist","crash","lift","desired","inter","closer","assumes","heights","shadow","riding","infection","firefox","lisa","expense","grove","eligibility","venture","clinic","korean","healing","princess","mall","entering","packet","spray","studios","involvement","dad","buttons","placement","observations","vbulletin","funded","thompson","winners","extend","roads","subsequent","pat","dublin","rolling","fell","motorcycle","yard","disclosure","establishment","memories","nelson","te","arrived","creates","faces","tourist","cocks","av","mayor","murder","sean","adequate","senator","yield","presentations","grades","cartoons","pour","digest","reg","lodging","tion","dust","hence","wiki","entirely","replaced","radar","rescue","undergraduate","losses","combat","reducing","stopped","occupation","lakes","butt","donations","associations","citysearch","closely","radiation","diary","seriously","kings","shooting","kent","adds","nsw","ear","flags","pci","baker","launched","elsewhere","pollution","conservative","guestbook","shock","effectiveness","walls","abroad","ebony","tie","ward","drawn","arthur","ian","visited","roof","walker","demonstrate","atmosphere","suggests","kiss","beast","ra","operated","experiment","targets","overseas","purchases","dodge","counsel","federation","pizza","invited","yards","assignment","chemicals","gordon","mod","farmers","rc","queries","bmw","rush","ukraine","absence","nearest","cluster","vendors","mpeg","whereas","yoga","serves","woods","surprise","lamp","rico","partial","shoppers","phil","everybody","couples","nashville","ranking","jokes","cst","http","ceo","simpson","twiki","sublime","counseling","palace","acceptable","satisfied","glad","wins","measurements","verify","globe","trusted","copper","milwaukee","rack","medication","warehouse","shareware","ec","rep","dicke","kerry","receipt","supposed","ordinary","nobody","ghost","violation","configure","stability","mit","applying","southwest","boss","pride","institutional","expectations","independence","knowing","reporter","metabolism","keith","champion","cloudy","linda","ross","personally","chile","anna","plenty","solo","sentence","throat","ignore","maria","uniform","excellence","wealth","tall","rm","somewhere","vacuum","dancing","attributes","recognize","brass","writes","plaza","pdas","outcomes","survival","quest","publish","sri","screening","toe","thumbnail","trans","jonathan","whenever","nova","lifetime","api","pioneer","booty","forgotten","acrobat","plates","acres","venue","athletic","thermal","essays","behaviour","vital","telling","fairly","coastal","config","cf","charity","intelligent","edinburgh","vt","excel","modes","obligation","campbell","wake","stupid","harbor","hungary","traveler","urw","segment","realize","regardless","lan","enemy","puzzle","rising","aluminum","wells","wishlist","opens","insight","sms","shit","restricted","republican","secrets","lucky","latter","merchants","thick","trailers","repeat","syndrome","philips","attendance","penalty","drum","glasses","enables","nec","iraqi","builder","vista","jessica","chips","terry","flood","foto","ease","arguments","amsterdam","orgy","arena","adventures","pupils","stewart","announcement","tabs","outcome","xx","appreciate","expanded","casual","grown","polish","lovely","extras","gm","centres","jerry","clause","smile","lands","ri","troops","indoor","bulgaria","armed","broker","charger","regularly","believed","pine","cooling","tend","gulf","rt","rick","trucks","cp","mechanisms","divorce","laura","shopper","tokyo","partly","nikon","customize","tradition","candy","pills","tiger","donald","folks","sensor","exposed","telecom","hunt","angels","deputy","indicators","sealed","thai","emissions","physicians","loaded","fred","complaint","scenes","experiments","balls","afghanistan","dd","boost","spanking","scholarship","governance","mill","founded","supplements","chronic","icons","tranny","moral","den","catering","aud","finger","keeps","pound","locate","camcorder","pl","trained","burn","implementing","roses","labs","ourselves","bread","tobacco","wooden","motors","tough","roberts","incident","gonna","dynamics","lie","crm","rf","conversation","decrease","cumshots","chest","pension","billy","revenues","emerging","worship","bukkake","capability","ak","fe","craig","herself","producing","churches","precision","damages","reserves","contributed","solve","shorts","reproduction","minority","td","diverse","amp","ingredients","sb","ah","johnny","sole","franchise","recorder","complaints","facing","sm","nancy","promotions","tones","passion","rehabilitation","maintaining","sight","laid","clay","defence","patches","weak","refund","usc","towns","environments","trembl","divided","blvd","reception","amd","wise","emails","cyprus","wv","odds","correctly","insider","seminars","consequences","makers","hearts","geography","appearing","integrity","worry","ns","discrimination","eve","carter","legacy","marc","pleased","danger","vitamin","widely","processed","phrase","genuine","raising","implications","functionality","paradise","hybrid","reads","roles","intermediate","emotional","sons","leaf","pad","glory","platforms","ja","bigger","billing","diesel","versus","combine","overnight","geographic","exceed","bs","rod","saudi","fault","cuba","hrs","preliminary","districts","introduce","silk","promotional","kate","chevrolet","babies","bi","karen","compiled","romantic","revealed","specialists","generator","albert","examine","jimmy","graham","suspension","bristol","margaret","compaq","sad","correction","wolf","slowly","authentication","communicate","rugby","supplement","showtimes","cal","portions","infant","promoting","sectors","samuel","fluid","grounds","fits","kick","regards","meal","ta","hurt","machinery","bandwidth","unlike","equation","baskets","probability","pot","dimension","wright","img","barry","proven","schedules","admissions","cached","warren","slip","studied","reviewer","involves","quarterly","rpm","profits","devil","grass","comply","marie","florist","illustrated","cherry","continental","alternate","deutsch","achievement","limitations","kenya","webcam","cuts","funeral","nutten","earrings","enjoyed","automated","chapters","pee","charlie","quebec","nipples","passenger","convenient","dennis","mars","francis","tvs","sized","manga","noticed","socket","silent","literary","egg","mhz","signals","caps","orientation","pill","theft","childhood","swing","symbols","lat","meta","humans","analog","facial","choosing","talent","dated","flexibility","seeker","wisdom","shoot","boundary","mint","packard","offset","payday","philip","elite","gi","spin","holders","believes","swedish","poems","deadline","jurisdiction","robot","displaying","witness","collins","equipped","stages","encouraged","sur","winds","powder","broadway","acquired","assess","wash","cartridges","stones","entrance","gnome","roots","declaration","losing","attempts","gadgets","noble","glasgow","automation","impacts","rev","gospel","advantages","shore","loves","induced","ll","knight","preparing","loose","aims","recipient","linking","extensions","appeals","cl","earned","illness","islamic","athletics","southeast","ieee","ho","alternatives","pending","parker","determining","lebanon","corp","personalized","kennedy","gt","sh","conditioning","teenage","soap","ae","triple","cooper","nyc","vincent","jam","secured","unusual","answered","partnerships","destruction","slots","increasingly","migration","disorder","routine","toolbar","basically","rocks","conventional","titans","applicants","wearing","axis","sought","genes","mounted","habitat","firewall","median","guns","scanner","herein","occupational","animated","horny","judicial","rio","hs","adjustment","hero","integer","treatments","bachelor","attitude","camcorders","engaged","falling","basics","montreal","carpet","rv","struct","lenses","binary","genetics","attended","difficulty","punk","collective","coalition","pi","dropped","enrollment","duke","walter","ai","pace","besides","wage","producers","ot","collector","arc","hosts","interfaces","advertisers","moments","atlas","strings","dawn","representing","observation","feels","torture","carl","deleted","coat","mitchell","mrs","rica","restoration","convenience","returning","ralph","opposition","container","yr","defendant","warner","confirmation","app","embedded","inkjet","supervisor","wizard","corps","actors","liver","peripherals","liable","brochure","morris","bestsellers","petition","eminem","recall","antenna","picked","assumed","departure","minneapolis","belief","killing","bikini","memphis","shoulder","decor","lookup","texts","harvard","brokers","roy","ion","diameter","ottawa","doll","ic","podcast","tit","seasons","peru","interactions","refine","bidder","singer","evans","herald","literacy","fails","aging","nike","intervention","pissing","fed","plugin","attraction","diving","invite","modification","alice","latinas","suppose","customized","reed","involve","moderate","terror","younger","thirty","mice","opposite","understood","rapidly","dealtime","ban","temp","intro","mercedes","zus","assurance","fisting","clerk","happening","vast","mills","outline","amendments","tramadol","holland","receives","jeans","metropolitan","compilation","verification","fonts","ent","odd","wrap","refers","mood","favor","veterans","quiz","mx","sigma","gr","attractive","xhtml","occasion","recordings","jefferson","victim","demands","sleeping","careful","ext","beam","gardening","obligations","arrive","orchestra","sunset","tracked","moreover","minimal","polyphonic","lottery","tops","framed","aside","outsourcing","licence","adjustable","allocation","michelle","essay","discipline","amy","ts","demonstrated","dialogue","identifying","alphabetical","camps","declared","dispatched","aaron","handheld","trace","disposal","shut","florists","packs","ge","installing","switches","romania","voluntary","ncaa","thou","consult","phd","greatly","blogging","mask","cycling","midnight","ng","commonly","pe","photographer","inform","turkish","coal","cry","messaging","pentium","quantum","murray","intent","tt","zoo","largely","pleasant","announce","constructed","additions","requiring","spoke","aka","arrow","engagement","sampling","rough","weird","tee","refinance","lion","inspired","holes","weddings","blade","suddenly","oxygen","cookie","meals","canyon","goto","meters","merely","calendars","arrangement","conclusions","passes","bibliography","pointer","compatibility","stretch","durham","furthermore","permits","cooperative","muslim","xl","neil","sleeve","netscape","cleaner","cricket","beef","feeding","stroke","township","rankings","measuring","cad","hats","robin","robinson","jacksonville","strap","headquarters","sharon","crowd","tcp","transfers","surf","olympic","transformation","remained","attachments","dv","dir","entities","customs","administrators","personality","rainbow","hook","roulette","decline","gloves","israeli","medicare","cord","skiing","cloud","facilitate","subscriber","valve","val","hewlett","explains","proceed","flickr","feelings","knife","jamaica","priorities","shelf","bookstore","timing","liked","parenting","adopt","denied","fotos","incredible","britney","freeware","fucked","donation","outer","crop","deaths","rivers","commonwealth","pharmaceutical","manhattan","tales","katrina","workforce","islam","nodes","tu","fy","thumbs","seeds","cited","lite","ghz","hub","targeted","organizational","skype","realized","twelve","founder","decade","gamecube","rr","dispute","portuguese","tired","titten","adverse","everywhere","excerpt","eng","steam","discharge","ef","drinks","ace","voices","acute","halloween","climbing","stood","sing","tons","perfume","carol","honest","albany","hazardous","restore","stack","methodology","somebody","sue","ep","housewares","reputation","resistant","democrats","recycling","hang","gbp","curve","creator","amber","qualifications","museums","coding","slideshow","tracker","variation","passage","transferred","trunk","hiking","lb","damn","pierre","jelsoft","headset","photograph","oakland","colombia","waves","camel","distributor","lamps","underlying","hood","wrestling","suicide","archived","photoshop","jp","chi","bt","arabia","gathering","projection","juice","chase","mathematical","logical","sauce","fame","extract","specialized","diagnostic","panama","indianapolis","af","payable","corporations","courtesy","criticism","automobile","confidential","rfc","statutory","accommodations","athens","northeast","downloaded","judges","sl","seo","retired","isp","remarks","detected","decades","paintings","walked","arising","nissan","bracelet","ins","eggs","juvenile","injection","yorkshire","populations","protective","afraid","acoustic","railway","cassette","initially","indicator","pointed","hb","jpg","causing","mistake","norton","locked","eliminate","tc","fusion","mineral","sunglasses","ruby","steering","beads","fortune","preference","canvas","threshold","parish","claimed","screens","cemetery","planner","croatia","flows","stadium","venezuela","exploration","mins","fewer","sequences","coupon","nurses","ssl","stem","proxy","gangbang","astronomy","lanka","opt","edwards","drew","contests","flu","translate","announces","mlb","costume","tagged","berkeley","voted","killer","bikes","gates","adjusted","rap","tune","bishop","pulled","corn","gp","shaped","compression","seasonal","establishing","farmer","counters","puts","constitutional","grew","perfectly","tin","slave","instantly","cultures","norfolk","coaching","examined","trek","encoding","litigation","submissions","oem","heroes","painted","lycos","ir","zdnet","broadcasting","horizontal","artwork","cosmetic","resulted","portrait","terrorist","informational","ethical","carriers","ecommerce","mobility","floral","builders","ties","struggle","schemes","suffering","neutral","fisher","rat","spears","prospective","dildos","bedding","ultimately","joining","heading","equally","artificial","bearing","spectacular","coordination","connector","brad","combo","seniors","worlds","guilty","affiliated","activation","naturally","haven","tablet","jury","dos","tail","subscribers","charm","lawn","violent","mitsubishi","underwear","basin","soup","potentially","ranch","constraints","crossing","inclusive","dimensional","cottage","drunk","considerable","crimes","resolved","mozilla","byte","toner","nose","latex","branches","anymore","oclc","delhi","holdings","alien","locator","selecting","processors","pantyhose","plc","broke","nepal","zimbabwe","difficulties","juan","complexity","msg","constantly","browsing","resolve","barcelona","presidential","documentary","cod","territories","melissa","moscow","thesis","thru","jews","nylon","palestinian","discs","rocky","bargains","frequent","trim","nigeria","ceiling","pixels","ensuring","hispanic","cv","cb","legislature","hospitality","gen","anybody","procurement","diamonds","espn","fleet","untitled","bunch","totals","marriott","singing","theoretical","afford","exercises","starring","referral","nhl","surveillance","optimal","quit","distinct","protocols","lung","highlight","substitute","inclusion","hopefully","brilliant","turner","sucking","cents","reuters","ti","fc","gel","todd","spoken","omega","evaluated","stayed","civic","assignments","fw","manuals","doug","sees","termination","watched","saver","thereof","grill","households","gs","redeem","rogers","grain","aaa","authentic","regime","wanna","wishes","bull","montgomery","architectural","louisville","depend","differ","macintosh","movements","ranging","monica","repairs","breath","amenities","virtually","cole","mart","candle","hanging","colored","authorization","tale","verified","lynn","formerly","projector","bp","situated","comparative","std","seeks","herbal","loving","strictly","routing","docs","stanley","psychological","surprised","retailer","vitamins","elegant","gains","renewal","vid","genealogy","opposed","deemed","scoring","expenditure","panties","brooklyn","liverpool","sisters","critics","connectivity","spots","oo","algorithms","hacker","madrid","similarly","margin","coin","bbw","solely","fake","salon","collaborative","norman","fda","excluding","turbo","headed","voters","cure","madonna","commander","arch","ni","murphy","thinks","thats","suggestion","hdtv","soldier","phillips","asin","aimed","justin","bomb","harm","interval","mirrors","spotlight","tricks","reset","brush","investigate","thy","expansys","panels","repeated","assault","connecting","spare","logistics","deer","kodak","tongue","bowling","tri","danish","pal","monkey","proportion","filename","skirt","florence","invest","honey","um","analyses","drawings","significance","scenario","ye","fs","lovers","atomic","approx","symposium","arabic","gauge","essentials","junction","protecting","nn","faced","mat","rachel","solving","transmitted","weekends","screenshots","produces","oven","ted","intensive","chains","kingston","sixth","engage","deviant","noon","switching","quoted","adapters","correspondence","farms","imports","supervision","cheat","bronze","expenditures","sandy","separation","testimony","suspect","celebrities","macro","sender","mandatory","boundaries","crucial","syndication","gym","celebration","kde","adjacent","filtering","tuition","spouse","exotic","viewer","signup","threats","luxembourg","puzzles","reaching","vb","damaged","cams","receptor","piss","laugh","joel","surgical","destroy","citation","pitch","autos","yo","premises","perry","proved","offensive","imperial","dozen","benjamin","deployment","teeth","cloth","studying","colleagues","stamp","lotus","salmon","olympus","separated","proc","cargo","tan","directive","fx","salem","mate","dl","starter","upgrades","likes","butter","pepper","weapon","luggage","burden","chef","tapes","zones","races","isle","stylish","slim","maple","luke","grocery","offshore","governing","retailers","depot","kenneth","comp","alt","pie","blend","harrison","ls","julie","occasionally","cbs","attending","emission","pete","spec","finest","realty","janet","bow","penn","recruiting","apparent","instructional","phpbb","autumn","traveling","probe","midi","permissions","biotechnology","toilet","ranked","jackets","routes","packed","excited","outreach","helen","mounting","recover","tied","lopez","balanced","prescribed","catherine","timely","talked","upskirts","debug","delayed","chuck","reproduced","hon","dale","explicit","calculation","villas","ebook","consolidated","boob","exclude","peeing","occasions","brooks","equations","newton","oils","sept","exceptional","anxiety","bingo","whilst","spatial","respondents","unto","lt","ceramic","prompt","precious","minds","annually","considerations","scanners","atm","xanax","eq","pays","cox","fingers","sunny","ebooks","delivers","je","queensland","necklace","musicians","leeds","composite","unavailable","cedar","arranged","lang","theaters","advocacy","raleigh","stud","fold","essentially","designing","threaded","uv","qualify","fingering","blair","hopes","assessments","cms","mason","diagram","burns","pumps","slut","ejaculation","footwear","sg","vic","beijing","peoples","victor","mario","pos","attach","licenses","utils","removing","advised","brunswick","spider","phys","ranges","pairs","sensitivity","trails","preservation","hudson","isolated","calgary","interim","assisted","divine","streaming","approve","chose","compound","intensity","technological","syndicate","abortion","dialog","venues","blast","wellness","calcium","newport","antivirus","addressing","pole","discounted","indians","shield","harvest","membrane","prague","previews","bangladesh","constitute","locally","concluded","pickup","desperate","mothers","nascar","iceland","demonstration","governmental","manufactured","candles","graduation","mega","bend","sailing","variations","moms","sacred","addiction","morocco","chrome","tommy","springfield","refused","brake","exterior","greeting","ecology","oliver","congo","glen","botswana","nav","delays","synthesis","olive","undefined","unemployment","cyber","verizon","scored","enhancement","newcastle","clone","dicks","velocity","lambda","relay","composed","tears","performances","oasis","baseline","cab","angry","fa","societies","silicon","brazilian","identical","petroleum","compete","ist","norwegian","lover","belong","honolulu","beatles","lips","escort","retention","exchanges","pond","rolls","thomson","barnes","soundtrack","wondering","malta","daddy","lc","ferry","rabbit","profession","seating","dam","cnn","separately","physiology","lil","collecting","das","exports","omaha","tire","participant","scholarships","recreational","dominican","chad","electron","loads","friendship","heather","passport","motel","unions","treasury","warrant","sys","solaris","frozen","occupied","josh","royalty","scales","rally","observer","sunshine","strain","drag","ceremony","somehow","arrested","expanding","provincial","investigations","icq","ripe","yamaha","rely","medications","hebrew","gained","rochester","dying","laundry","stuck","solomon","placing","stops","homework","adjust","assessed","advertiser","enabling","encryption","filling","downloadable","sophisticated","imposed","silence","scsi","focuses","soviet","possession","cu","laboratories","treaty","vocal","trainer","organ","stronger","volumes","advances","vegetables","lemon","toxic","dns","thumbnails","darkness","pty","ws","nuts","nail","bizrate","vienna","implied","span","stanford","sox","stockings","joke","respondent","packing","statute","rejected","satisfy","destroyed","shelter","chapel","gamespot","manufacture","layers","wordpress","guided","vulnerability","accountability","celebrate","accredited","appliance","compressed","bahamas","powell","mixture","zoophilia","bench","univ","tub","rider","scheduling","radius","perspectives","mortality","logging","hampton","christians","borders","therapeutic","pads","butts","inns","bobby","impressive","sheep","accordingly","architect","railroad","lectures","challenging","wines","nursery","harder","cups","ash","microwave","cheapest","accidents","travesti","relocation","stuart","contributors","salvador","ali","salad","np","monroe","tender","violations","foam","temperatures","paste","clouds","competitions","discretion","tft","tanzania","preserve","jvc","poem","vibrator","unsigned","staying","cosmetics","easter","theories","repository","praise","jeremy","venice","jo","concentrations","vibrators","estonia","christianity","veteran","streams","landing","signing","executed","katie","negotiations","realistic","dt","cgi","showcase","integral","asks","relax","namibia","generating","christina","congressional","synopsis","hardly","prairie","reunion","composer","bean","sword","absent","photographic","sells","ecuador","hoping","accessed","spirits","modifications","coral","pixel","float","colin","bias","imported","paths","bubble","por","acquire","contrary","millennium","tribune","vessel","acids","focusing","viruses","cheaper","admitted","dairy","admit","mem","fancy","equality","samoa","gc","achieving","tap","stickers","fisheries","exceptions","reactions","leasing","lauren","beliefs","ci","macromedia","companion","squad","analyze","ashley","scroll","relate","divisions","swim","wages","additionally","suffer","forests","fellowship","nano","invalid","concerts","martial","males","victorian","retain","colours","execute","tunnel","genres","cambodia","patents","copyrights","yn","chaos","lithuania","mastercard","wheat","chronicles","obtaining","beaver","updating","distribute","readings","decorative","kijiji","confused","compiler","enlargement","eagles","bases","vii","accused","bee","campaigns","unity","loud","conjunction","bride","rats","defines","airports","instances","indigenous","begun","cfr","brunette","packets","anchor","socks","validation","parade","corruption","stat","trigger","incentives","cholesterol","gathered","essex","slovenia","notified","differential","beaches","folders","dramatic","surfaces","terrible","routers","cruz","pendant","dresses","baptist","scientist","starsmerchant","hiring","clocks","arthritis","bios","females","wallace","nevertheless","reflects","taxation","fever","pmc","cuisine","surely","practitioners","transcript","myspace","theorem","inflation","thee","nb","ruth","pray","stylus","compounds","pope","drums","contracting","topless","arnold","structured","reasonably","jeep","chicks","bare","hung","cattle","mba","radical","graduates","rover","recommends","controlling","treasure","reload","distributors","flame","levitra","tanks","assuming","monetary","elderly","pit","arlington","mono","particles","floating","extraordinary","tile","indicating","bolivia","spell","hottest","stevens","coordinate","kuwait","exclusively","emily","alleged","limitation","widescreen","compile","squirting","webster","struck","rx","illustration","plymouth","warnings","construct","apps","inquiries","bridal","annex","mag","gsm","inspiration","tribal","curious","affecting","freight","rebate","meetup","eclipse","sudan","ddr","downloading","rec","shuttle","aggregate","stunning","cycles","affects","forecasts","detect","sluts","actively","ciao","ampland","knee","prep","pb","complicated","chem","fastest","butler","shopzilla","injured","decorating","payroll","cookbook","expressions","ton","courier","uploaded","shakespeare","hints","collapse","americas","connectors","twinks","unlikely","oe","gif","pros","conflicts","techno","beverage","tribute","wired","elvis","immune","latvia","travelers","forestry","barriers","cant","jd","rarely","gpl","infected","offerings","martha","genesis","barrier","argue","incorrect","trains","metals","bicycle","furnishings","letting","arise","guatemala","celtic","thereby","irc","jamie","particle","perception","minerals","advise","humidity","bottles","boxing","wy","dm","bangkok","renaissance","pathology","sara","bra","ordinance","hughes","photographers","bitch","infections","jeffrey","chess","operates","brisbane","configured","survive","oscar","festivals","menus","joan","possibilities","duck","reveal","canal","amino","phi","contributing","herbs","clinics","mls","cow","manitoba","analytical","missions","watson","lying","costumes","strict","dive","saddam","circulation","drill","offense","threesome","bryan","cet","protest","handjob","assumption","jerusalem","hobby","tries","transexuales","invention","nickname","fiji","technician","inline","executives","enquiries","washing","audi","staffing","cognitive","exploring","trick","enquiry","closure","raid","ppc","timber","volt","intense","div","playlist","registrar","showers","supporters","ruling","steady","dirt","statutes","withdrawal","myers","drops","predicted","wider","saskatchewan","jc","cancellation","plugins","enrolled","sensors","screw","ministers","publicly","hourly","blame","geneva","freebsd","veterinary","acer","prostores","reseller","dist","handed","suffered","intake","informal","relevance","incentive","butterfly","tucson","mechanics","heavily","swingers","fifty","headers","mistakes","numerical","ons","geek","uncle","defining","xnxx","counting","reflection","sink","accompanied","assure","invitation","devoted","princeton","jacob","sodium","randy","spirituality","hormone","meanwhile","proprietary","timothy","childrens","brick","grip","naval","thumbzilla","medieval","porcelain","avi","bridges","pichunter","captured","watt","thehun","decent","casting","dayton","translated","shortly","cameron","columnists","pins","carlos","reno","donna","andreas","warrior","diploma","cabin","innocent","bdsm","scanning","ide","consensus","polo","valium","copying","rpg","delivering","cordless","patricia","horn","eddie","uganda","fired","journalism","pd","prot","trivia","adidas","perth","frog","grammar","intention","syria","disagree","klein","harvey","tires","logs","undertaken","tgp","hazard","retro","leo","livesex","statewide","semiconductor","gregory","episodes","boolean","circular","anger","diy","mainland","illustrations","suits","chances","interact","snap","happiness","arg","substantially","bizarre","glenn","ur","auckland","olympics","fruits","identifier","geo","worldsex","ribbon","calculations","doe","jpeg","conducting","startup","suzuki","trinidad","ati","kissing","wal","handy","swap","exempt","crops","reduces","accomplished","calculators","geometry","impression","abs","slovakia","flip","guild","correlation","gorgeous","capitol","sim","dishes","rna","barbados","chrysler","nervous","refuse","extends","fragrance","mcdonald","replica","plumbing","brussels","tribe","neighbors","trades","superb","buzz","transparent","nuke","rid","trinity","charleston","handled","legends","boom","calm","champions","floors","selections","projectors","inappropriate","exhaust","comparing","shanghai","speaks","burton","vocational","davidson","copied","scotia","farming","gibson","pharmacies","fork","troy","ln","roller","introducing","batch","organize","appreciated","alter","nicole","latino","ghana","edges","uc","mixing","handles","skilled","fitted","albuquerque","harmony","distinguished","asthma","projected","assumptions","shareholders","twins","developmental","rip","zope","regulated","triangle","amend","anticipated","oriental","reward","windsor","zambia","completing","gmbh","buf","ld","hydrogen","webshots","sprint","comparable","chick","advocate","sims","confusion","copyrighted","tray","inputs","warranties","genome","escorts","documented","thong","medal","paperbacks","coaches","vessels","harbour","walks","sucks","sol","keyboards","sage","knives","eco","vulnerable","arrange","artistic","bat","honors","booth","indie","reflected","unified","bones","breed","detector","ignored","polar","fallen","precise","sussex","respiratory","notifications","msgid","transexual","mainstream","invoice","evaluating","lip","subcommittee","sap","gather","suse","maternity","backed","alfred","colonial","mf","carey","motels","forming","embassy","cave","journalists","danny","rebecca","slight","proceeds","indirect","amongst","wool","foundations","msgstr","arrest","volleyball","mw","adipex","horizon","nu","deeply","toolbox","ict","marina","liabilities","prizes","bosnia","browsers","decreased","patio","dp","tolerance","surfing","creativity","lloyd","describing","optics","pursue","lightning","overcome","eyed","ou","quotations","grab","inspector","attract","brighton","beans","bookmarks","ellis","disable","snake","succeed","leonard","lending","oops","reminder","nipple","xi","searched","behavioral","riverside","bathrooms","plains","sku","ht","raymond","insights","abilities","initiated","sullivan","za","midwest","karaoke","trap","lonely","fool","ve","nonprofit","lancaster","suspended","hereby","observe","julia","containers","attitudes","karl","berry","collar","simultaneously","racial","integrate","bermuda","amanda","sociology","mobiles","screenshot","exhibitions","kelkoo","confident","retrieved","exhibits","officially","consortium","dies","terrace","bacteria","pts","replied","seafood","novels","rh","rrp","recipients","playboy","ought","delicious","traditions","fg","jail","safely","finite","kidney","periodically","fixes","sends","durable","mazda","allied","throws","moisture","hungarian","roster","referring","symantec","spencer","wichita","nasdaq","uruguay","ooo","hz","transform","timer","tablets","tuning","gotten","educators","tyler","futures","vegetable","verse","highs","humanities","independently","wanting","custody","scratch","launches","ipaq","alignment","masturbating","henderson","bk","britannica","comm","ellen","competitors","nhs","rocket","aye","bullet","towers","racks","lace","nasty","visibility","latitude","consciousness","ste","tumor","ugly","deposits","beverly","mistress","encounter","trustees","watts","duncan","reprints","hart","bernard","resolutions","ment","accessing","forty","tubes","attempted","col","midlands","priest","floyd","ronald","analysts","queue","dx","sk","trance","locale","nicholas","biol","yu","bundle","hammer","invasion","witnesses","runner","rows","administered","notion","sq","skins","mailed","oc","fujitsu","spelling","arctic","exams","rewards","beneath","strengthen","defend","aj","frederick","medicaid","treo","infrared","seventh","gods","une","welsh","belly","aggressive","tex","advertisements","quarters","stolen","cia","sublimedirectory","soonest","haiti","disturbed","determines","sculpture","poly","ears","dod","wp","fist","naturals","neo","motivation","lenders","pharmacology","fitting","fixtures","bloggers","mere","agrees","passengers","quantities","petersburg","consistently","powerpoint","cons","surplus","elder","sonic","obituaries","cheers","dig","taxi","punishment","appreciation","subsequently","om","belarus","nat","zoning","gravity","providence","thumb","restriction","incorporate","backgrounds","treasurer","guitars","essence","flooring","lightweight","ethiopia","tp","mighty","athletes","humanity","transcription","jm","holmes","complications","scholars","dpi","scripting","gis","remembered","galaxy","chester","snapshot","caring","loc","worn","synthetic","shaw","vp","segments","testament","expo","dominant","twist","specifics","itunes","stomach","partially","buried","cn","newbie","minimize","darwin","ranks","wilderness","debut","generations","tournaments","bradley","deny","anatomy","bali","judy","sponsorship","headphones","fraction","trio","proceeding","cube","defects","volkswagen","uncertainty","breakdown","milton","marker","reconstruction","subsidiary","strengths","clarity","rugs","sandra","adelaide","encouraging","furnished","monaco","settled","folding","emirates","terrorists","airfare","comparisons","beneficial","distributions","vaccine","belize","crap","fate","viewpicture","promised","volvo","penny","robust","bookings","threatened","minolta","republicans","discusses","gui","porter","gras","jungle","ver","rn","responded","rim","abstracts","zen","ivory","alpine","dis","prediction","pharmaceuticals","andale","fabulous","remix","alias","thesaurus","individually","battlefield","literally","newer","kay","ecological","spice","oval","implies","cg","soma","ser","cooler","appraisal","consisting","maritime","periodic","submitting","overhead","ascii","prospect","shipment","breeding","citations","geographical","donor","mozambique","tension","href","benz","trash","shapes","wifi","tier","fwd","earl","manor","envelope","diane","homeland","disclaimers","championships","excluded","andrea","breeds","rapids","disco","sheffield","bailey","aus","endif","finishing","emotions","wellington","incoming","prospects","lexmark","cleaners","bulgarian","hwy","eternal","cashiers","guam","cite","aboriginal","remarkable","rotation","nam","preventing","productive","boulevard","eugene","ix","gdp","pig","metric","compliant","minus","penalties","bennett","imagination","hotmail","refurbished","joshua","armenia","varied","grande","closest","activated","actress","mess","conferencing","assign","armstrong","politicians","trackbacks","lit","accommodate","tigers","aurora","una","slides","milan","premiere","lender","villages","shade","chorus","christine","rhythm","digit","argued","dietary","symphony","clarke","sudden","accepting","precipitation","marilyn","lions","findlaw","ada","pools","tb","lyric","claire","isolation","speeds","sustained","matched","approximate","rope","carroll","rational","programmer","fighters","chambers","dump","greetings","inherited","warming","incomplete","vocals","chronicle","fountain","chubby","grave","legitimate","biographies","burner","yrs","foo","investigator","gba","plaintiff","finnish","gentle","bm","prisoners","deeper","muslims","hose","mediterranean","nightlife","footage","howto","worthy","reveals","architects","saints","entrepreneur","carries","sig","freelance","duo","excessive","devon","screensaver","helena","saves","regarded","valuation","unexpected","cigarette","fog","characteristic","marion","lobby","egyptian","tunisia","metallica","outlined","consequently","headline","treating","punch","appointments","str","gotta","cowboy","narrative","bahrain","enormous","karma","consist","betty","queens","academics","pubs","quantitative","shemales","lucas","screensavers","subdivision","tribes","vip","defeat","clicks","distinction","honduras","naughty","hazards","insured","harper","livestock","mardi","exemption","tenant","sustainability","cabinets","tattoo","shake","algebra","shadows","holly","formatting","silly","nutritional","yea","mercy","hartford","freely","marcus","sunrise","wrapping","mild","fur","nicaragua","weblogs","timeline","tar","belongs","rj","readily","affiliation","soc","fence","nudist","infinite","diana","ensures","relatives","lindsay","clan","legally","shame","satisfactory","revolutionary","bracelets","sync","civilian","telephony","mesa","fatal","remedy","realtors","breathing","briefly","thickness","adjustments","graphical","genius","discussing","aerospace","fighter","meaningful","flesh","retreat","adapted","barely","wherever","estates","rug","democrat","borough","maintains","failing","shortcuts","ka","retained","voyeurweb","pamela","andrews","marble","extending","jesse","specifies","hull","logitech","surrey","briefing","belkin","dem","accreditation","wav","blackberry","highland","meditation","modular","microphone","macedonia","combining","brandon","instrumental","giants","organizing","shed","balloon","moderators","winston","memo","ham","solved","tide","kazakhstan","hawaiian","standings","partition","invisible","gratuit","consoles","funk","fbi","qatar","magnet","translations","porsche","cayman","jaguar","reel","sheer","commodity","posing","wang","kilometers","rp","bind","thanksgiving","rand","hopkins","urgent","guarantees","infants","gothic","cylinder","witch","buck","indication","eh","congratulations","tba","cohen","sie","usgs","puppy","kathy","acre","graphs","surround","cigarettes","revenge","expires","enemies","lows","controllers","aqua","chen","emma","consultancy","finances","accepts","enjoying","conventions","eva","patrol","smell","pest","hc","italiano","coordinates","rca","fp","carnival","roughly","sticker","promises","responding","reef","physically","divide","stakeholders","hydrocodone","gst","consecutive","cornell","satin","bon","deserve","attempting","mailto","promo","jj","representations","chan","worried","tunes","garbage","competing","combines","mas","beth","bradford","len","phrases","kai","peninsula","chelsea","boring","reynolds","dom","jill","accurately","speeches","reaches","schema","considers","sofa","catalogs","ministries","vacancies","quizzes","parliamentary","obj","prefix","lucia","savannah","barrel","typing","nerve","dans","planets","deficit","boulder","pointing","renew","coupled","viii","myanmar","metadata","harold","circuits","floppy","texture","handbags","jar","ev","somerset","incurred","acknowledge","thoroughly","antigua","nottingham","thunder","tent","caution","identifies","questionnaire","qualification","locks","modelling","namely","miniature","dept","hack","dare","euros","interstate","pirates","aerial","hawk","consequence","rebel","systematic","perceived","origins","hired","makeup","textile","lamb","madagascar","nathan","tobago","presenting","cos","troubleshooting","uzbekistan","indexes","pac","rl","erp","centuries","gl","magnitude","ui","richardson","hindu","dh","fragrances","vocabulary","licking","earthquake","vpn","fundraising","fcc","markers","weights","albania","geological","assessing","lasting","wicked","eds","introduces","kills","roommate","webcams","pushed","webmasters","ro","df","computational","acdbentity","participated","junk","handhelds","wax","lucy","answering","hans","impressed","slope","reggae","failures","poet","conspiracy","surname","theology","nails","evident","whats","rides","rehab","epic","saturn","organizer","nut","allergy","sake","twisted","combinations","preceding","merit","enzyme","cumulative","zshops","planes","edmonton","tackle","disks","condo","pokemon","amplifier","ambien","arbitrary","prominent","retrieve","lexington","vernon","sans","worldcat","titanium","irs","fairy","builds","contacted","shaft","lean","bye","cdt","recorders","occasional","leslie","casio","deutsche","ana","postings","innovations","kitty","postcards","dude","drain","monte","fires","algeria","blessed","luis","reviewing","cardiff","cornwall","favors","potato","panic","explicitly","sticks","leone","transsexual","ez","citizenship","excuse","reforms","basement","onion","strand","pf","sandwich","uw","lawsuit","alto","informative","girlfriend","bloomberg","cheque","hierarchy","influenced","banners","reject","eau","abandoned","bd","circles","italic","beats","merry","mil","scuba","gore","complement","cult","dash","passive","mauritius","valued","cage","checklist","bangbus","requesting","courage","verde","lauderdale","scenarios","gazette","hitachi","divx","extraction","batman","elevation","hearings","coleman","hugh","lap","utilization","beverages","calibration","jake","eval","efficiently","anaheim","ping","textbook","dried","entertaining","prerequisite","luther","frontier","settle","stopping","refugees","knights","hypothesis","palmer","medicines","flux","derby","sao","peaceful","altered","pontiac","regression","doctrine","scenic","trainers","muze","enhancements","renewable","intersection","passwords","sewing","consistency","collectors","conclude","recognised","munich","oman","celebs","gmc","propose","hh","azerbaijan","lighter","rage","adsl","uh","prix","astrology","advisors","pavilion","tactics","trusts","occurring","supplemental","travelling","talented","annie","pillow","induction","derek","precisely","shorter","harley","spreading","provinces","relying","finals","paraguay","steal","parcel","refined","fd","bo","fifteen","widespread","incidence","fears","predict","boutique","acrylic","rolled","tuner","avon","incidents","peterson","rays","asn","shannon","toddler","enhancing","flavor","alike","walt","homeless","horrible","hungry","metallic","acne","blocked","interference","warriors","palestine","listprice","libs","undo","cadillac","atmospheric","malawi","wm","pk","sagem","knowledgestorm","dana","halo","ppm","curtis","parental","referenced","strikes","lesser","publicity","marathon","ant","proposition","gays","pressing","gasoline","apt","dressed","scout","belfast","exec","dealt","niagara","inf","eos","warcraft","charms","catalyst","trader","bucks","allowance","vcr","denial","uri","designation","thrown","prepaid","raises","gem","duplicate","electro","criterion","badge","wrist","civilization","analyzed","vietnamese","heath","tremendous","ballot","lexus","varying","remedies","validity","trustee","maui","handjobs","weighted","angola","squirt","performs","plastics","realm","corrected","jenny","helmet","salaries","postcard","elephant","yemen","encountered","tsunami","scholar","nickel","internationally","surrounded","psi","buses","expedia","geology","pct","wb","creatures","coating","commented","wallet","cleared","smilies","vids","accomplish","boating","drainage","shakira","corners","broader","vegetarian","rouge","yeast","yale","newfoundland","sn","qld","pas","clearing","investigated","dk","ambassador","coated","intend","stephanie","contacting","vegetation","doom","findarticles","louise","kenny","specially","owen","routines","hitting","yukon","beings","bite","issn","aquatic","reliance","habits","striking","myth","infectious","podcasts","singh","gig","gilbert","sas","ferrari","continuity","brook","fu","outputs","phenomenon","ensemble","insulin","assured","biblical","weed","conscious","accent","mysimon","eleven","wives","ambient","utilize","mileage","oecd","prostate","adaptor","auburn","unlock","hyundai","pledge","vampire","angela","relates","nitrogen","xerox","dice","merger","softball","referrals","quad","dock","differently","firewire","mods","nextel","framing","organised","musician","blocking","rwanda","sorts","integrating","vsnet","limiting","dispatch","revisions","papua","restored","hint","armor","riders","chargers","remark","dozens","varies","msie","reasoning","wn","liz","rendered","picking","charitable","guards","annotated","ccd","sv","convinced","openings","buys","burlington","replacing","researcher","watershed","councils","occupations","acknowledged","nudity","kruger","pockets","granny","pork","zu","equilibrium","viral","inquire","pipes","characterized","laden","aruba","cottages","realtor","merge","privilege","edgar","develops","qualifying","chassis","dubai","estimation","barn","pushing","llp","fleece","pediatric","boc","fare","dg","asus","pierce","allan","dressing","techrepublic","sperm","vg","bald","filme","craps","fuji","frost","leon","institutes","mold","dame","fo","sally","yacht","tracy","prefers","drilling","brochures","herb","tmp","alot","ate","breach","whale","traveller","appropriations","suspected","tomatoes","benchmark","beginners","instructors","highlighted","bedford","stationery","idle","mustang","unauthorized","clusters","antibody","competent","momentum","fin","wiring","io","pastor","mud","calvin","uni","shark","contributor","demonstrates","phases","grateful","emerald","gradually","laughing","grows","cliff","desirable","tract","ul","ballet","ol","journalist","abraham","js","bumper","afterwards","webpage","religions","garlic","hostels","shine","senegal","explosion","pn","banned","wendy","briefs","signatures","diffs","cove","mumbai","ozone","disciplines","casa","mu","daughters","conversations","radios","tariff","nvidia","opponent","pasta","simplified","muscles","serum","wrapped","swift","motherboard","runtime","inbox","focal","bibliographic","vagina","eden","distant","incl","champagne","ala","decimal","hq","deviation","superintendent","propecia","dip","nbc","samba","hostel","housewives","employ","mongolia","penguin","magical","influences","inspections","irrigation","miracle","manually","reprint","reid","wt","hydraulic","centered","robertson","flex","yearly","penetration","wound","belle","rosa","conviction","hash","omissions","writings","hamburg","lazy","mv","mpg","retrieval","qualities","cindy","lolita","fathers","carb","charging","cas","marvel","lined","cio","dow","prototype","importantly","rb","petite","apparatus","upc","terrain","dui","pens","explaining","yen","strips","gossip","rangers","nomination","empirical","mh","rotary","worm","dependence","discrete","beginner","boxed","lid","sexuality","polyester","cubic","deaf","commitments","suggesting","sapphire","kinase","skirts","mats","remainder","crawford","labeled","privileges","televisions","specializing","marking","commodities","pvc","serbia","sheriff","griffin","declined","guyana","spies","blah","mime","neighbor","motorcycles","elect","highways","thinkpad","concentrate","intimate","reproductive","preston","deadly","cunt","feof","bunny","chevy","molecules","rounds","longest","refrigerator","tions","intervals","sentences","dentists","usda","exclusion","workstation","holocaust","keen","flyer","peas","dosage","receivers","urls","customise","disposition","variance","navigator","investigators","cameroon","baking","marijuana","adaptive","computed","needle","baths","enb","gg","cathedral","brakes","og","nirvana","ko","fairfield","owns","til","invision","sticky","destiny","generous","madness","emacs","climb","blowing","fascinating","landscapes","heated","lafayette","jackie","wto","computation","hay","cardiovascular","ww","sparc","cardiac","salvation","dover","adrian","predictions","accompanying","vatican","brutal","learners","gd","selective","arbitration","configuring","token","editorials","zinc","sacrifice","seekers","guru","isa","removable","convergence","yields","gibraltar","levy","suited","numeric","anthropology","skating","kinda","aberdeen","emperor","grad","malpractice","dylan","bras","belts","blacks","educated","rebates","reporters","burke","proudly","pix","necessity","rendering","mic","inserted","pulling","basename","kyle","obesity","curves","suburban","touring","clara","vertex","bw","hepatitis","nationally","tomato","andorra","waterproof","expired","mj","travels","flush","waiver","pale","specialties","hayes","humanitarian","invitations","functioning","delight","survivor","garcia","cingular","economies","alexandria","bacterial","moses","counted","undertake","declare","continuously","johns","valves","gaps","impaired","achievements","donors","tear","jewel","teddy","lf","convertible","ata","teaches","ventures","nil","bufing","stranger","tragedy","julian","nest","pam","dryer","painful","velvet","tribunal","ruled","nato","pensions","prayers","funky","secretariat","nowhere","cop","paragraphs","gale","joins","adolescent","nominations","wesley","dim","lately","cancelled","scary","mattress","mpegs","brunei","likewise","banana","introductory","slovak","cakes","stan","reservoir","occurrence","idol","bloody","mixer","remind","wc","worcester","sbjct","demographic","charming","mai","tooth","disciplinary","annoying","respected","stays","disclose","affair","drove","washer","upset","restrict","springer","beside","mines","portraits","rebound","logan","mentor","interpreted","evaluations","fought","baghdad","elimination","metres","hypothetical","immigrants","complimentary","helicopter","pencil","freeze","hk","performer","abu","titled","commissions","sphere","powerseller","moss","ratios","concord","graduated","endorsed","ty","surprising","walnut","lance","ladder","italia","unnecessary","dramatically","liberia","sherman","cork","maximize","cj","hansen","senators","workout","mali","yugoslavia","bleeding","characterization","colon","likelihood","lanes","purse","fundamentals","contamination","mtv","endangered","compromise","masturbation","optimize","stating","dome","caroline","leu","expiration","namespace","align","peripheral","bless","engaging","negotiation","crest","opponents","triumph","nominated","confidentiality","electoral","changelog","welding","orgasm","deferred","alternatively","heel","alloy","condos","plots","polished","yang","gently","greensboro","tulsa","locking","casey","controversial","draws","fridge","blanket","bloom","qc","simpsons","lou","elliott","recovered","fraser","justify","upgrading","blades","pgp","loops","surge","frontpage","trauma","aw","tahoe","advert","possess","demanding","defensive","sip","flashers","subaru","forbidden","tf","vanilla","programmers","pj","monitored","installations","deutschland","picnic","souls","arrivals","spank","cw","practitioner","motivated","wr","dumb","smithsonian","hollow","vault","securely","examining","fioricet","groove","revelation","rg","pursuit","delegation","wires","bl","dictionaries","mails","backing","greenhouse","sleeps","vc","blake","transparency","dee","travis","wx","endless","figured","orbit","currencies","niger","bacon","survivors","positioning","heater","colony","cannon","circus","promoted","forbes","mae","moldova","mel","descending","paxil","spine","trout","enclosed","feat","temporarily","ntsc","cooked","thriller","transmit","apnic","fatty","gerald","pressed","frequencies","scanned","reflections","hunger","mariah","sic","municipality","usps","joyce","detective","surgeon","cement","experiencing","fireplace","endorsement","bg","planners","disputes","textiles","missile","intranet","closes","seq","psychiatry","persistent","deborah","conf","marco","assists","summaries","glow","gabriel","auditor","wma","aquarium","violin","prophet","cir","bracket","looksmart","isaac","oxide","oaks","magnificent","erik","colleague","naples","promptly","modems","adaptation","hu","harmful","paintball","prozac","sexually","enclosure","acm","dividend","newark","kw","paso","glucose","phantom","norm","playback","supervisors","westminster","turtle","ips","distances","absorption","treasures","dsc","warned","neural","ware","fossil","mia","hometown","badly","transcripts","apollo","wan","disappointed","persian","continually","communist","collectible","handmade","greene","entrepreneurs","robots","grenada","creations","jade","scoop","acquisitions","foul","keno","gtk","earning","mailman","sanyo","nested","biodiversity","excitement","somalia","movers","verbal","blink","presently","seas","carlo","workflow","mysterious","novelty","bryant","tiles","voyuer","librarian","subsidiaries","switched","stockholm","tamil","garmin","ru","pose","fuzzy","indonesian","grams","therapist","richards","mrna","budgets","toolkit","promising","relaxation","goat","render","carmen","ira","sen","thereafter","hardwood","erotica","temporal","sail","forge","commissioners","dense","dts","brave","forwarding","qt","awful","nightmare","airplane","reductions","southampton","istanbul","impose","organisms","sega","telescope","viewers","asbestos","portsmouth","cdna","meyer","enters","pod","savage","advancement","wu","harassment","willow","resumes","bolt","gage","throwing","existed","whore","generators","lu","wagon","barbie","dat","favour","soa","knock","urge","smtp","generates","potatoes","thorough","replication","inexpensive","kurt","receptors","peers","roland","optimum","neon","interventions","quilt","huntington","creature","ours","mounts","syracuse","internship","lone","refresh","aluminium","snowboard","beastality","webcast","michel","evanescence","subtle","coordinated","notre","shipments","maldives","stripes","firmware","antarctica","cope","shepherd","lm","canberra","cradle","chancellor","mambo","lime","kirk","flour","controversy","legendary","bool","sympathy","choir","avoiding","beautifully","blond","expects","cho","jumping","fabrics","antibodies","polymer","hygiene","wit","poultry","virtue","burst","examinations","surgeons","bouquet","immunology","promotes","mandate","wiley","departmental","bbs","spas","ind","corpus","johnston","terminology","gentleman","fibre","reproduce","convicted","shades","jets","indices","roommates","adware","qui","intl","threatening","spokesman","zoloft","activists","frankfurt","prisoner","daisy","halifax","encourages","ultram","cursor","assembled","earliest","donated","stuffed","restructuring","insects","terminals","crude","morrison","maiden","simulations","cz","sufficiently","examines","viking","myrtle","bored","cleanup","yarn","knit","conditional","mug","crossword","bother","budapest","conceptual","knitting","attacked","hl","bhutan","liechtenstein","mating","compute","redhead","arrives","translator","automobiles","tractor","allah","continent","ob","unwrap","fares","longitude","resist","challenged","telecharger","hoped","pike","safer","insertion","instrumentation","ids","hugo","wagner","constraint","groundwater","touched","strengthening","cologne","gzip","wishing","ranger","smallest","insulation","newman","marsh","ricky","ctrl","scared","theta","infringement","bent","laos","subjective","monsters","asylum","lightbox","robbie","stake","cocktail","outlets","swaziland","varieties","arbor","mediawiki","configurations","poison"]
    secmon_common_words = ['HHS','HSE','IC','Insight Partners','Interpol',"L'acteur",'Ledger','Microsoft Windows','NASA','NET','NordVPN','OTAN','Patch','Password',"Play Store","Pakistan","Outlook","PDC","NoxPlayer","Python","Rapid","Red Hat","RiskIQ","Samsung","SEO","San Francisco","SolarWinds","SSL-VPN","SSN","Suivez","Taîwan","Trustwave","TTP","Ubuntu","UC","VirusTotal","VSA","Washington Post","Windows Server","Xbox","ZIP","Zyxel",'Apprenez','Agent Tesla','AlmaLinux','Avast','Big Data','Cofense','Command','Comprendre','Congrès','CSP','Cyberpunk','CyberNews','DVR','ESXi','FIN','Google Play','Google Play Store','GreatHorn','Anne Neuberger','Bitcoin Fog','Excel','Word','Powerpoint','Nbsp','LDAP','FortiWLC','SSL Forti','CISO','CIO',"L'expert",'SIM','Taiwan','Maryland','SSL Forti','New York Times','CWE-','NFT','National Cyber','Xcode','Acronis','Ten Eleven Ventures','Advent International','Conférence','Rust','Minnesota','Le National Cyber','SITA','UEFI','X-Force','Ministère','LastPass','Lituanie','NIST','Devermont','Supermicro','HackerOne','TeamViewer','Ukraine','UDP','Internet Explorer','SEC','Chainalysis','Netflix','NCC','Surnommée','Commentaire','FAI','Le Do','Phones','Pwn','SASE','TAG','Patch Tuesday','SAML','Géorgie','Google Cloud','Joker','Informa','DoH','Stanford','FortiGuard Labs','Illustrator','Photoshop','Colorado','Instagram','Naked Security Live',"L'acquisition",'LLC','Virginie','Veritas','Pysa','Citrix','AHK',"Google's",'Golang','Cybersecurity','PIN','According','KrebsOnSecurity','Things','MIME','Amazon Web Services','China','India','How','Cyber Security Hub','FormBook','APKPure','Armorblox','SMB',"L'organisation",'Kansas','OF','Verkada','Security',"Jusqu'à",'Mercredi','UE','Suivi','Talos','Français','PayPal','WiFi',"L'outil",'Windows Defender','Jeudi',"L'exploitation",'IBM Security','La Cybersecurity','CI','ZDNet','AD','DC','Malwarebytes','CMS',"L'enquête",'IcedID','Adobe','Microsoft Defender','Omdia','PII','Cloud','Bluetooth','Recorded Future','Message','Dragos','Orion',"Joker's",'Investigation','Project Zero','Naked Security','FA','AI','ForgePoint Capital','Elasticsearch','But','Iran','Edge','December','Magento','DBIR','TLD','FiveHands','Have','Been Pwned','Bash Uploader','Jen Easterly','Pierre Fabre','RV','Group-IB','Zero Trust',"Accellion","L'Université",'Clubhouse','Unité','Mimecast',"L'Agence",'Azure','Mozilla','Git','MSP','Stash','FortiClient','Fortune',"Here's",'We','SSH','Dark Web','Cellebrite','Chromium','PowerShell','Microsoft Corp','Positive Technologies','YouTube','ICS','SMA','SentinelOne','Proofpoint','Infrastructure Security Agency','DevOps','DoJ','Akamai','TikTok','Claroty','Nouvelle-Zélande','Tesla','Mandiant','Zscaler','Emsisoft','VoIP','GPS','Soumettez','TA','CLI','ID','IRS','To','Russia','If','Cisco Talos',"L'incident","DLL","Verizon","The Edge","Journée","Day","FTC","Département","CVE-","COVID-","DDoS","APT","L'attaque",'''"L'Université"''',"Kubernetes","Docker","Gitlab","GitLab","L'application","RDP","Maison Blanche","Security Center","Près","L'équipe","New York","Bitdefender","UNC","FTA",'''"L'incident"''',"FortiGate","Check Point Research","Pulse Secure","Exchange Server","RAT","Palo Alto Networks","Joe Biden","Microsoft Office","Wi-Fi","Tor","IoT","OT","IA","GitHub","ESET","XDR","RCE","DHS","EDR","AV","Codecov","JavaScript","Tom Merritt","Après","WhatsApp","Découvrez","Bitcoin","And","SPONSORISÉ","Discord","URL","CONTENU","CVSS","Phone","HTML","CSS","Résumé","Google Chrome","Apr","Go","This","DNS","HTTP","HTTPS","SSL","TLS","TCP","FTP","Zoom","Teams","Malgr","NAS","QNAP","PoC","Amazon","Kaspersky","McAfee","CrowdStrike","Active Directory","Signal","DoS","SVR","Telegram","TheHackersNews","Fi","Europol","As","FortiOS","Orient","Russie","Russian","Israël","MITRE","ATT","IaaS","SaaS","Dark Reading","Wordpress","WordPress","Wi","Microsoft Exchange","Microsoft Exchange Server","BEC","Détails","CK","BleepingComputer","Fortinet","Trend Micro","Fortigate","NCSC","Même","SOC","AWS","RSA","FireEye","Cor","LinkedIn","Linkedin","VMware","Agence","PHP","SQL","Lisez","Github","SonicWall","NSA","SMS","Firefox","Javascript","An","VPN","XSS","API","Jack Wallen","Cyber","Here","Est","Check Point","Checkpoint","Sophos","FortiProxy","FortiWeb","Unis","Université","Chrome","Troie","Facebook","Twitter","Uni","Linux","Google","Android","CVE","IT","COVID","Biden","ZATAZ","PDF","RSSI","CISA","IoT","FBI"]

    conn = get_db_connection()
    dbresult = conn.execute("SELECT title,summary,rss_f FROM RSS_DATA").fetchall()
    global_news = []
    for dbresult1 in dbresult:
        if not "us-cert" in dbresult1[2] and not "cert.ssi" in dbresult1[2]:
            global_news.append(dbresult1[1])

    proper_names = []
    for news in global_news:
        multiple_capitalized = re.findall('([A-Z][a-zA-ZÀ-ÿ]+(?=\s[A-Z])(?:\s[A-Z][a-z]+)+)', news)
        for e in multiple_capitalized:
            proper_names.append(e)
        single_capitalized = re.findall("""([A-Z][a-zA-ZÀ-ÿ-']+)""",news)
        for e in single_capitalized:
            if e not in ' '.join(multiple_capitalized):
                proper_names.append(e)

    unique_word_list = list(dict.fromkeys(proper_names))
    word_stats = []
    for word in unique_word_list:
        if word not in fr_common_words and word not in en_common_words and word not in secmon_common_words:
            occurrences = proper_names.count(word)
            word_stats.append([word,occurrences])
    top_word = sortWordsByOccurrences(word_stats)
    for word in top_word:
        if word[1] >= 4:
            top_words.append(word[0])
    return top_words
def handleException(error):
    now = datetime.now()
    now = now.strftime("%d/%m/%Y, %H:%M:%S")
    print(f"################## NEW SCRIPT ERROR AT {now} ##################")
    print(traceback.print_exc())
    print("################## PLEASE REPORT THIS ON GITHUB ##################")
def generateCveReport(start_date,end_date,isFull):
    bold_font = Font(bold=True,color='00FAAA3C', size=14)
    grey = Color(rgb='00202020')
    center_aligned_text = Alignment(horizontal="center")
    wrapped_text = Alignment(horizontal="center",wrapText=True,shrinkToFit=True,vertical="center")
    thin_border = Side(border_style="thin")
    square_border = Border(top=thin_border,right=thin_border,bottom=thin_border,left=thin_border)
    title_bg = PatternFill(fill_type = "solid",fgColor=grey)
    fr = [
    'ID CVE',
    'Produit matché',
    'Description',
    'Date de publication',
    'Score CVSS',
    'Sources',
    'Statut',
    'Produits affectés']
    en = [
    'CVE ID',
    'Matching product',
    'Summary',
    'Published date',
    'CVSS Score',
    'Sources',
    'Status',
    'Affected products']
    cols = ["A","B","C","D","E","F","G","H"]
    script_path = os.path.abspath(__file__)
    dir_path = script_path.replace("secmon_lib.py","")  
    xlsx_report = Workbook()
    sheet = xlsx_report.active
    sheet.title = "Global"
    lang = getUserLanguage()
    sheet.column_dimensions['A'].width = 21
    sheet.column_dimensions['B'].width = 58
    sheet.column_dimensions['C'].width = 73
    sheet.column_dimensions['D'].width = 15
    sheet.column_dimensions['E'].width = 10
    sheet.column_dimensions['F'].width = 110
    sheet.column_dimensions['G'].width = 20
    sheet.column_dimensions['H'].width = 68
    sheet.auto_filter.ref = sheet.dimensions
    i = 0
    while i != 8:
        if lang == "fr":
            letter = cols[i]
            sheet["{}1".format(letter)].font = bold_font
            sheet["{}1".format(letter)].alignment = center_aligned_text
            sheet["{}1".format(letter)].border = square_border
            sheet["{}1".format(letter)].fill = title_bg
            sheet["{}1".format(letter)].value = fr[i]
            i+=1
        else:
            letter = cols[i]
            sheet["{}1".format(letter)].font = bold_font
            sheet["{}1".format(letter)].alignment = center_aligned_text
            sheet["{}1".format(letter)].border = square_border
            sheet["{}1".format(letter)].fill = title_bg
            sheet["{}1".format(letter)].value = en[i]
            i+=1            
    cves = []
    script_path = os.path.abspath(__file__)
    dir_path = script_path.replace("secmon_lib.py","")
    con = sqlite3.connect(dir_path+'secmon.db')
    cur = con.cursor()
    cur.execute("SELECT CVE_ID,KEYWORD,CVE_DESCRIPTION,CVE_DATE,CVE_SCORE,CVE_SOURCES,STATUS,CVE_CPE,INDEXING_DATE FROM CVE_DATA")
    db_result_list = cur.fetchall()
    cves = []
    for db_result_tuple in db_result_list:
        current_cve = []
        for result in db_result_tuple:
            current_cve.append(result)
        cves.append(current_cve)
    for cve in cves:
        cve_date = datetime.strptime(cve[8], "%d/%m/%Y")
        if start_date <= cve_date <= end_date or isFull:
            match = cve[1]
            if "cpe" in match:
                match = getParsedCpe(match)
            cve[1] = match
            cpes = cve[7]
            new_product_list = ""
            if "cpe" in cpes:
                cpe_list = cpes.split("\n")
                for cpe in cpe_list:
                    if "cpe" in cpe:
                        hname = getParsedCpe(cpe)
                        new_product_list+=(hname+"\n")
            if new_product_list == "" or new_product_list == " ":
                new_product_list = "N/A"
            cve[7] = new_product_list
            sheet.append(cve[:-1])
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = wrapped_text

    xlsx_report.save(filename=dir_path+"SECMON-CVE-Report.xlsx")
def writeCveTypeLog(source_script,cve_id,temporal_type,matched_product,cvss_score,cve_date,cpes,report,alert,*args):
    server = socket.gethostbyname(socket.gethostname())
    if cvss_score != "N/A" and cvss_score != "":
        if 0 < float(cvss_score) <= 3.9:
            severity = "Low"
        elif 4.0 < float(cvss_score) <= 6.9:
            severity = "Medium"
        elif 7.0 < float(cvss_score) <= 8.9:
            severity = "High"
        elif 9.0 < float(cvss_score) <= 10:
            severity = "Critical"
        elif cvss_score == "N/A":
            severity = "N/A"
    else:
        severity = "N/A"
    timestamp = datetime.now().isoformat()
    if temporal_type == "new":
        message = f"The {cve_id} has been added."
    else:
        message = f"The {cve_id} has been updated."
    for arg in args:
        message+=arg

    cpes = cpes.replace("\n","+")

    log = f"""{timestamp} source_script="{source_script}" server="{server}" cve_id="{cve_id}" type="{temporal_type}" matched_product="{matched_product}" cvss_score="{cvss_score}" severity="{severity}" cve_date="{cve_date}" cpes="{cpes}" report="{report}" alert="{alert}" message="{message}" \n"""
    script_path = os.path.abspath(__file__)
    dir_path = script_path.replace(os.path.basename(__file__),"")
    log_file = dir_path+"logs.txt"
    with open(log_file, "a") as f:
        f.write(log.replace('+" report=','" report='))
        
def writeNewExploitFoundLog(source_script,exploit_source,cve,message):
    timestamp = datetime.now().isoformat()
    server = socket.gethostbyname(socket.gethostname())
    log = f"""{timestamp} source_script="{source_script}" server="{server}" cve="{cve}" exploit_source="{exploit_source}" message="{message}" \n"""
    script_path = os.path.abspath(__file__)
    dir_path = script_path.replace(os.path.basename(__file__),"")
    log_file = dir_path+"logs.txt"
    with open(log_file, "a") as f:
        f.write(log)

def writeNewHighRiskProductLog(source_script,cpe,message):
    timestamp = datetime.now().isoformat()
    server = socket.gethostbyname(socket.gethostname())
    if "cpe" in cpe:
        hname = getParsedCpe(cpe)
    else:
        hname = cpe
    log = f"""{timestamp} source_script="{source_script}" server="{server}" cpe="{cpe}" hname="{hname}" message="{message}" \n"""
    script_path = os.path.abspath(__file__)
    dir_path = script_path.replace(os.path.basename(__file__),"")
    log_file = dir_path+"logs.txt"
    with open(log_file, "a") as f:
        f.write(log)
def writeAuthLog(source_script,username,auth_status,msg,src_ip):
    timestamp = datetime.now().isoformat()
    server = socket.gethostbyname(socket.gethostname())
    log = f"""{timestamp} source_script="{source_script}" server="{server}" src_ip="{src_ip}" username="{username}" auth_status="{auth_status}" message="{msg}" \n"""
    script_path = os.path.abspath(__file__)
    dir_path = script_path.replace(os.path.basename(__file__),"")
    log_file = dir_path+"logs.txt"
    with open(log_file, "a") as f:
        f.write(log)
def getTasks():
    script_path = os.path.abspath(__file__)
    dir_path = script_path.replace("secmon_lib.py","")
    logfile = dir_path+"logs.txt"
    count = 0
    tid = []
    if os.path.isfile(logfile) == True:
        for line in reversed(list(open(logfile))):
            if "task_id" in line:
                timestamp = line.split(" ")[0]
                formatted_date = datetime.fromisoformat(timestamp)
                date_2_days_ago = datetime.now() - timedelta(hours=2)
                if (date_2_days_ago<=formatted_date) and (formatted_date<=datetime.now()):
                    task_id = line.split('"')[5]
                    if task_id not in tid:
                        tid.append(task_id)
    tasks = []
    for task in tid:
        status = ""
        operation = ""
        if os.path.isfile(logfile) == True:
            for line in reversed(list(open(logfile))):
                if f'task_id="{task}"' in line and "loaded" in line:
                    operation += line.split('"')[9].replace("....",".")
                elif f'task_id="{task}"' in line and "failed" in line:
                    status += "Failed"
                elif f'task_id="{task}"' in line and "success" in line:
                    status += "Success"
            if status =="":
                status="In progress"
        tasks.append([task,operation,status])
    return tasks
def writeTaskLog(source_script,task_id,task_status,message):
    timestamp = datetime.now().isoformat()
    server = socket.gethostbyname(socket.gethostname())
    log = f"""{timestamp} source_script="{source_script}" server="{server}" task_id="{task_id}" task_status="{task_status}" message="{message}" \n"""
    script_path = os.path.abspath(__file__)
    dir_path = script_path.replace(os.path.basename(__file__),"")
    log_file = dir_path+"logs.txt"
    with open(log_file, "a") as f:
        f.write(log)

def writeMgmtTasksLog(source_script,message):
    timestamp = datetime.now().isoformat()
    server = socket.gethostbyname(socket.gethostname())
    log = f"""{timestamp} source_script="{source_script}" server="{server}" message="{message}" \n"""
    script_path = os.path.abspath(__file__)
    dir_path = script_path.replace(os.path.basename(__file__),"")
    log_file = dir_path+"logs.txt"
    with open(log_file, "a") as f:
        f.write(log)


def writeNewRssNewLog(source_script,feed,feed_url,news_title,news_url,mail):
    timestamp = datetime.now().isoformat()
    server = socket.gethostbyname(socket.gethostname()) 
    log = f"""{timestamp} source_script="{source_script}" server="{server}" feed="{feed}" feed_url="{feed_url}" news_title="{news_title}" news_url="{news_url}" mail="{mail}" \n"""   
    script_path = os.path.abspath(__file__)
    dir_path = script_path.replace(os.path.basename(__file__),"")
    log_file = dir_path+"logs.txt"
    with open(log_file, "a") as f:
        f.write(log)
def translateText(lang,text):
    try:
        translation = GoogleTranslator(source='auto', target=lang).translate(text=text)
        return translation
    except Exception as e:
        handleException(e)
        try:
            translation = MyMemoryTranslator(source='auto', target=lang).translate(text)
            return translation
        except Exception as e:
            handleException(e)
            return text +" (TRANSLATION FAILED)"

def getGithubAPISettings():
    key,username = "",""
    script_path = os.path.abspath(__file__)
    dir_path = script_path.replace("secmon_lib.py","")
    con = sqlite3.connect(dir_path+'secmon.db')
    cur = con.cursor()
    cur.execute("SELECT github_api_key FROM config")
    db_result_tuple = cur.fetchone()
    for value in db_result_tuple:
        key = value
    cur.execute("SELECT github_username FROM config")
    db_result_tuple = cur.fetchone()
    for value in db_result_tuple:
        username = value
    return key,username

def getUserLanguage():
    script_path = os.path.abspath(__file__)
    dir_path = script_path.replace("secmon_lib.py","")
    con = sqlite3.connect(dir_path+'secmon.db')
    cur = con.cursor()
    cur.execute("SELECT language FROM config")
    db_result_tuple = cur.fetchone()
    for value in db_result_tuple:
        language = value
    return language
def searchExploit(cve, sleeping):
    exploits = CS.edbid_from_cve(cve)
    result = []
    for exploit in exploits: 
        current_exploit = []
        exploitdb_url = "https://www.exploit-db.com/exploits/"+str(exploit)
        r = requests.get(exploitdb_url,headers={"User-Agent":"Mozilla/5.0"})
        soup = BeautifulSoup(r.text,"html.parser")
        exploit_infos = soup.find_all(class_="stats-title")
        eID = exploit_infos[0].get_text().replace(" ","").replace("\n","")
        eAuthor = exploit_infos[2].get_text().replace(" ","").replace("\n","")
        eType = exploit_infos[3].get_text().replace(" ","").replace("\n","")
        ePlatform = exploit_infos[4].get_text().replace(" ","").replace("\n","")
        eDate = exploit_infos[5].get_text().replace(" ","").replace("\n","")

        current_exploit.append(eID)
        current_exploit.append(exploitdb_url)
        current_exploit.append(eAuthor)
        current_exploit.append(eType)
        current_exploit.append(ePlatform)
        created_date = eDate.split("-")
        current_exploit.append(created_date[2]+"/"+created_date[1]+"/"+created_date[0])
        current_exploit.append("Exploit-DB")
        result.append(current_exploit)
    github_query = "https://api.github.com/search/repositories?q=exploit+"+cve
    github_response = requests.get(url=github_query)
    github_data = github_response.json()
    if sleep == True:
        if "message" in github_data.keys():
            sleep(60)
    if "items" in github_data.keys():
        for exploit in github_data['items']:
            current_exploit = []
            current_exploit.append(exploit["id"])
            current_exploit.append(exploit["html_url"])
            current_exploit.append(exploit["owner"]["login"])
            current_exploit.append("N/A")
            current_exploit.append("N/A")
            created_date = exploit["created_at"].split("T")[0].split("-")
            current_exploit.append(created_date[2]+"/"+created_date[1]+"/"+created_date[0])
            current_exploit.append("Github")
            result.append(current_exploit)
    return result
def pollCveIdFromCpe(cpe):
    cve_ids = []
    if "cpe" in cpe:
        try:
            nvd_query = "https://services.nvd.nist.gov/rest/json/cves/1.0?cpeMatchString="+cpe+"&resultsPerPage=1000"
            nvd_response = requests.get(url=nvd_query)
            nvd_data = nvd_response.json()
            for cve in nvd_data["result"]["CVE_Items"]:
                cve_id = cve["cve"]["CVE_data_meta"]["ID"]
                if cve_id not in cve_ids:
                    cve_ids.append(cve_id)
            result_nb = int(nvd_response.json()['totalResults'])
            current_result = 1000
            page = 1
            while current_result < result_nb:
                page+=1
                nvd_query = "https://services.nvd.nist.gov/rest/json/cves/1.0?cpeMatchString="+cpe+"&resultsPerPage=1000"+f"&startIndex={current_result}"
                nvd_response = requests.get(url=nvd_query)
                nvd_data = nvd_response.json()
                for cve in nvd_data["result"]["CVE_Items"]:
                    cve_id = cve["cve"]["CVE_data_meta"]["ID"]
                    if cve_id not in cve_ids:
                        cve_ids.append(cve_id)
                current_result+=1000
        except Exception as e:
            print("Unable to poll CVE related to :",cpe)
            handleException(e)
            pass
    else:
        cve_rss = []
        current_feed = feedparser.parse("https://nvd.nist.gov/feeds/xml/cve/misc/nvd-rss.xml")
        for entries in current_feed.entries:
            title = entries.title
            summary = entries.summary
            cve_id = re.findall('CVE-\d{4}-\d{4,7}',title)
            cve_id = cve_id[0]
            cve_rss.append(cve_id+"(=)"+summary)
        for cve in cve_rss:
            cve_summary = cve.split("(=)")[1]
            cve_id = cve.split("(=)")[0]
            if cpe in cve_summary:
               cve_ids.append(cve_id)
    return cve_ids

def getSecretKey():
    script_path = os.path.abspath(__file__)
    dir_path = script_path.replace("secmon_lib.py","")
    con = sqlite3.connect(dir_path+'secmon.db')
    cur = con.cursor()
    db_result = cur.execute("SELECT app_secret_key FROM secret_key")
    for data in db_result:
        secret_key = data[0]
    return secret_key

def getHighRiskProducts(dentype):
    if dentype == "hname": 
        critical_products = []
        conn = get_db_connection()
        db_result = conn.execute('SELECT hname FROM high_risk_products',).fetchall()
        for result in db_result:
            current_products = []
            for tuples in result:
                critical_products.append(tuples)
    elif dentype == "cpe": 
        critical_products = []
        conn = get_db_connection()
        db_result = conn.execute('SELECT cpe FROM high_risk_products',).fetchall()
        for result in db_result:
            current_products = []
            for tuples in result:
                critical_products.append(tuples)
    elif dentype == "all": 
        critical_products = []
        conn = get_db_connection()
        db_result = conn.execute('SELECT cpe,hname FROM high_risk_products',).fetchall()
        for result in db_result:
            current_product = []
            for tuples in result:
                current_product.append(tuples)
                if len(current_product) == 2:
                    critical_products.append(current_product)
                    continue
    return critical_products

def returnUsername():
    script_path = os.path.abspath(__file__)
    dir_path = script_path.replace("secmon_lib.py","")
    con = sqlite3.connect(dir_path+'secmon.db')
    cur = con.cursor()
    db_result = cur.execute("SELECT username FROM users")
    for data in db_result:
        for users in data:
            if is_logged_in(users):
                return users
            else:
                continue
    return "Not logged !!!"
def deleteProduct(ptype, key_or_cpe):    
    if ptype == "CPE":
        high_risk_products = getHighRiskProducts("cpe")
        if key_or_cpe in high_risk_products:
            con = get_db_connection()
            cur = con.cursor()
            cur.execute("DELETE FROM high_risk_products WHERE cpe = (?);", (key_or_cpe,))
            con.commit()  
        con = get_db_connection()
        cur = con.cursor()
        try:
            cpelist = []
            cur.execute("SELECT cpe FROM cpe_list")
            db_result_list = cur.fetchall()
            for db_result_tuple in db_result_list:
                for result in db_result_tuple:  
                    cpelist.append(result)
            if key_or_cpe not in cpelist:
                raise
            if not "cpe" in key_or_cpe:
                raise
            cur.execute("DELETE FROM cpe_list WHERE cpe = (?);", (key_or_cpe,))
            cur.execute("DELETE FROM CVE_DATA WHERE KEYWORD = (?);", (key_or_cpe,))
            con.commit()
            return "OK"
        except:
            return "NOK"
    else:
        con = get_db_connection()
        cur = con.cursor()
        try:
            keylist = []
            cur.execute("SELECT keyword FROM keyword_list")
            db_result_list = cur.fetchall()
            for db_result_tuple in db_result_list:
                for result in db_result_tuple:  
                    keylist.append(result)
            if key_or_cpe not in keylist:
                raise
            cur.execute("DELETE FROM keyword_list WHERE keyword = (?);", (key_or_cpe,))
            cur.execute("DELETE FROM CVE_DATA WHERE KEYWORD = (?);", (key_or_cpe,))
            con.commit()
            return "OK"
        except:
            return "NOK"

def registerNewCve(cve_id,reason,product):
    try:
        now_date = str(datetime.now()).split(" ")[0].split("-")
        idx_date = now_date[2]+"/"+now_date[1]+"/"+now_date[0]
        con = get_db_connection()
        cur = con.cursor()
        cur.execute("SELECT language FROM config")
        db_result_tuple = cur.fetchone()
        for value in db_result_tuple:
            language = value
        cur.execute("SELECT CVE_ID FROM CVE_DATA WHERE CVE_ID = (?);", (cve_id,))
        db_result_list = cur.fetchall()
        db_result_str = ""
        for db_result_tuple in db_result_list:
            for result in db_result_tuple:
                db_result_str+=result
        if cve_id not in db_result_str:
            nvd_base_url = "https://services.nvd.nist.gov/rest/json/cve/1.0/"
            nvd_query = nvd_base_url+cve_id
            nvd_response = requests.get(url=nvd_query)
            try:
                nvd_data = nvd_response.json()
            except:
                pass
            if 'result' in nvd_data.keys():
                if 'reference_data' in nvd_data['result']['CVE_Items'][0]['cve']['references']:
                    nvd_links = nvd_data['result']['CVE_Items'][0]['cve']['references']['reference_data']
                    cve_sources = ""
                    for link in nvd_links:
                        cve_sources += (link['url']+"\n")
                else:
                        cve_sources = "N/A"
                cve_score = ""
                if not "impact" in nvd_data['result']['CVE_Items'][0].keys():
                    cve_score = "N/A"
                    try:    
                        webpage = requests.get("https://nvd.nist.gov/vuln/detail/{}".format(cve_id))
                        soup = BeautifulSoup(webpage.content, 'html.parser')
                        cve_cna_score = soup.find(id='Cvss3CnaCalculatorAnchor')
                        cve_cna_score = cve_cna_score.get_text()
                        if cve_cna_score != "":
                            cve_score = cve_cna_score
                        else:
                            cve_score = "N/A"
                    except:
                            cve_score = "N/A"
                else:
                    try:
                        cve_score = nvd_data['result']['CVE_Items'][0]['impact']['baseMetricV3']['cvssV3']['baseScore']
                    except:
                        cve_score = "N/A"
                published_date = nvd_data['result']['CVE_Items'][0]['publishedDate']
                formatted_date = published_date.split("T")[0]
                formatted_date = datetime.strptime(formatted_date,"%Y-%m-%d")
                date_2_days_ago = datetime.now() - timedelta(days=7)
                if (date_2_days_ago<=formatted_date) and (formatted_date<=datetime.now()):
                    cve_date_status = "Potentially new CVE"
                else:
                    cve_date_status = "Potentially an update"
                cve_date = published_date.split("T")[0]
                if language == "fr" or language == "FR":
                    fr_date = cve_date.split("-")
                    cve_date = fr_date[2]+"/"+fr_date[1]+"/"+fr_date[0]
                if (cve_score=="N/A"):
                    cve_status = cve_date_status+" - "+"Not yet rated (No score, no CPE)"
                else:
                    cve_status = cve_date_status+" - "+"Valid evaluation"
                cve_description = nvd_data['result']['CVE_Items'][0]['cve']['description']['description_data'][0]['value']
                cve_cpe = ""
                if "configurations" in nvd_data['result']['CVE_Items'][0].keys():
                    if nvd_data['result']['CVE_Items'][0]['configurations']['nodes']:
                        for node in nvd_data['result']['CVE_Items'][0]['configurations']['nodes']:
                            if 'cpe_match' in node.keys():
                                for cpe_node in node['cpe_match']:
                                    cve_cpe += (cpe_node['cpe23Uri']+'\n')
            else:
                cve_score = "N/A"
                cve_date = "N/A"
                cve_cpe = "N/A"
                cve_description = "N/A"
                cve_status = "N/A"

            if cve_cpe == "":
                cve_cpe = "N/A"
            key_match = product
            if reason == "NewProduct":  
                status = "Found when the product was added."
            elif reason == "NewPolling":
                status = "Unread"
            elif reason == "Setup":
                status = "Native"
            else:
                status = "N/A"        
            cur.execute("INSERT INTO CVE_DATA (CVE_ID,KEYWORD,STATUS,CVE_SCORE,CVE_DATE,CVE_DESCRIPTION,CVE_EVAL,CVE_CPE,CVE_SOURCES,EXPLOIT_FIND,INDEXING_DATE) VALUES (?,?,?,?,?,?,?,?,?,?,?);", (cve_id,key_match,status,str(cve_score),cve_date,cve_description,cve_status,cve_cpe,cve_sources,"False",idx_date))
            con.commit()
            return "ok"
        else:
            return "already"
    except Exception as e:
        print("ERROR when SECMON would to register this CVE : ",cve_id)
        handleException(e)
def addProduct(ptype, key_or_cpe):
    task_id = random.randint(10000,99999)
    writeTaskLog("secmon_web",task_id,"loaded",f"Adding the following product : {key_or_cpe}....")
    if ptype == "CPE":
        con = get_db_connection()
        cur = con.cursor()
        try:
            cpe = key_or_cpe
            cve_ids = pollCveIdFromCpe(cpe)
            for cve_id in cve_ids:
                try:
                    registerNewCve(cve_id,"NewProduct",cpe)
                except Exception as e:
                    handleException(e)
                    pass
            cpelist = []
            cur.execute("SELECT cpe FROM cpe_list")
            db_result_list = cur.fetchall()
            for db_result_tuple in db_result_list:
                for result in db_result_tuple:  
                    cpelist.append(result)
            if key_or_cpe in cpelist:
                raise Exception('Product already in the product list.')
            if not "cpe" in key_or_cpe:
                raise Exception('The CPE is in a bad format.')         
            cur.execute("INSERT INTO cpe_list (cpe) VALUES (?);", (key_or_cpe,))
            con.commit()
            writeTaskLog("secmon_web",task_id,"success",f"Following product successfully added : {key_or_cpe} !")        

        except Exception as e:
            handleException(e)
            writeTaskLog("secmon_web",task_id,"failed",f"Unable to add this product : {key_or_cpe}. {str(e)}.") 
    else:
        con = get_db_connection()
        cur = con.cursor()
        try:
            key = key_or_cpe
            cve_ids = pollCveIdFromCpe(key)
            for cve_id in cve_ids:
                try:
                    registerNewCve(cve_id,"NewProduct",key)
                except Exception as e:
                    handleException(e)
                    pass
            keylist = []
            cur.execute("SELECT keyword FROM keyword_list")
            db_result_list = cur.fetchall()
            for db_result_tuple in db_result_list:
                for result in db_result_tuple:  
                    keylist.append(result)
            if key_or_cpe in keylist:
                 raise Exception('Product already exist ! ')
            cur.execute("INSERT INTO keyword_list (keyword) VALUES (?);", (key_or_cpe,))
            con.commit()
            writeTaskLog("secmon_web",task_id,"success",f"Following product successfully added : {key_or_cpe} !") 
            return "OK"
        except Exception as e:
            handleException(e)
            writeTaskLog("secmon_web",task_id,"failed",f"Unable to add this product : {key_or_cpe}. {str(e)}.") 
            return "NOK"
def getProductsStats():
    product_number = 0
    con = get_db_connection()
    cur = con.cursor()
    cur.execute("SELECT keyword FROM keyword_list")
    db_result_list = cur.fetchall()
    for db_result_tuple in db_result_list:
        for result in db_result_tuple:  
            product_number+=1
    cur.execute("SELECT cpe FROM cpe_list")
    db_result_list = cur.fetchall()
    for db_result_tuple in db_result_list:
        for result in db_result_tuple:  
            product_number+=1
    return product_number
def getCveByProduct(product,sleeping):
    con = get_db_connection()
    cur = con.cursor()
    if "cpe" in product:
        pversion = product.split(":")[5]
        all_version_cpe = product.replace(pversion,"*")
        req = cur.execute("SELECT CVE_ID FROM CVE_DATA WHERE KEYWORD = (?) UNION SELECT CVE_ID FROM CVE_DATA WHERE CVE_CPE LIKE (?) UNION SELECT CVE_ID FROM CVE_DATA WHERE KEYWORD = (?) UNION SELECT CVE_ID FROM CVE_DATA WHERE CVE_CPE LIKE (?);", (product,"*"+product+"*",all_version_cpe,"*"+all_version_cpe+"*",))
    else:
        req = cur.execute("SELECT CVE_ID FROM CVE_DATA WHERE KEYWORD = (?);", (product,)) 

    db_result_list = req.fetchall()
    cve_list = []
    for db_result_tuple in db_result_list:
        for result in db_result_tuple:
            if result not in cve_list:
                cve_list.append(result)

    critical_cve, high_cve, medium_cve, low_cve, na_cve, exploitable_cve, expl_cve= [], [], [], [], [], [], []
    if "cpe" in product:
        req2 = cur.execute('SELECT CVE_ID FROM CVE_DATA WHERE EXPLOIT_FIND = "True" AND KEYWORD = (?) UNION SELECT CVE_ID FROM CVE_DATA WHERE EXPLOIT_FIND = "True" AND CVE_CPE LIKE (?) UNION SELECT CVE_ID FROM CVE_DATA WHERE EXPLOIT_FIND = "True" AND KEYWORD = (?) UNION SELECT CVE_ID FROM CVE_DATA WHERE EXPLOIT_FIND = "True" AND CVE_CPE LIKE (?);',(product,"*"+product+"*",all_version_cpe,"*"+all_version_cpe+"*",))
        db_result_list2 = req2.fetchall()
        for db_result_tuple2 in db_result_list2:
            for result2 in db_result_tuple2:
                if result2 not in expl_cve:
                    expl_cve.append(result2)
    else:
        req2 = cur.execute('SELECT CVE_ID FROM CVE_DATA WHERE EXPLOIT_FIND = "True" AND KEYWORD = (?);',(product,))
        db_result_list2 = req2.fetchall()
        for db_result_tuple2 in db_result_list2:
            for result2 in db_result_tuple2:
                if result2 not in expl_cve:
                    expl_cve.append(result2)

    for cve in cve_list:
        db_result_list = cur.execute("SELECT CVE_ID,CVE_SCORE,STATUS FROM CVE_DATA WHERE CVE_ID = (?);", (cve,)).fetchall()
        for db_result_tuple in db_result_list:
            current_cve = []
            for result in db_result_tuple:
                current_cve.append(result)
            if current_cve[1] != None and current_cve[1] != "N/A" :
                current_cve[1] = current_cve[1].split(" ")[0]
                if 0 < float(current_cve[1]) <= 3.9:
                    low_cve.append(current_cve)
                elif 4.0 < float(current_cve[1]) <= 6.9:
                    medium_cve.append(current_cve)
                elif 7.0 < float(current_cve[1]) <= 8.9:
                    high_cve.append(current_cve)
                elif 9.0 < float(current_cve[1]) <= 10:
                    critical_cve.append(current_cve)
            elif current_cve[1] == "N/A":
                na_cve.append(current_cve)
            if cve in expl_cve:
                exploitable_cve.append(current_cve)

    return critical_cve, high_cve, medium_cve, low_cve, na_cve, exploitable_cve

def getParsedCpe(cpe):
    try:
        cpe = cpe.replace("*","- All versions")
        current_product = []
        disassmbled_cpe = cpe.split(":")
        pvendor = disassmbled_cpe[3].replace("_"," ").title()
        pproduct = disassmbled_cpe[4].replace("_"," ").title()
        subversion_infos = "("+disassmbled_cpe[6] +" "+ disassmbled_cpe[7] +" "+ disassmbled_cpe[8]+")"
        subversion_infos = subversion_infos.replace("- All versions","").replace(" - All versions","").replace("   )",")").replace("  )",")").replace(" )",")").replace("(   ","(").replace("(  ","(").replace("( ","(")
        if subversion_infos == "()" or subversion_infos == " " or subversion_infos == "( )":
            subversion_infos = ""
        try:    
            pversion = disassmbled_cpe[5].title()
        except Exception as e:
            handleException(e)
            pversion = ""
        if pversion == "-":
            product = pvendor+" "+pproduct
        else:
            product = pvendor+" "+pproduct+" "+pversion + subversion_infos
        return product
    except:
        print("Unable to parse this CPE :",cpe)
        return "Unparsed product : "+cpe
def getProductInfos(product):
	current_product = []
	if "cpe" in product:
		secmon_ptype = "CPE"
        fcpe = product
		cpe = product.replace("*","All")
		disassmbled_cpe = product.split(":")
		cpeptype = disassmbled_cpe[2]
		if cpeptype == "a":
			ptype = "Application"
		elif cpeptype == "o":
			ptype = "Operating System"
		elif cpeptype == "h":
			ptype = "Hardware"
		else: 
			ptype = "N/A" 
		pvendor = disassmbled_cpe[3].replace("_"," ").title()
		pproduct = disassmbled_cpe[4].replace("_"," ").title()
		try:
			pversion = disassmbled_cpe[5]
		except:
			pversion = ""
		if pversion == "-" or pversion == "":
			pversion = "All versions"
	else:
		secmon_ptype = "Keyword"
		fcpe = "N/A"
		pvendor = "N/A"
		pproduct = product
		ptype = "N/A"
		pversion = "All versions"
	current_product.append(secmon_ptype)
	current_product.append(fcpe)
	current_product.append(ptype)
	current_product.append(pvendor)
	current_product.append(pproduct)
	current_product.append(pversion)  
	return current_product
def getFormatedProductList():
    con = get_db_connection()
    cur = con.cursor()
    cur.execute("SELECT keyword FROM keyword_list")
    db_result_list = cur.fetchall()
    keywords = []
    for db_result_tuple in db_result_list:
        for result in db_result_tuple:  
            keywords.append(result)
    cur.execute("SELECT cpe FROM cpe_list")
    db_result_list = cur.fetchall()
    cpes = []
    for db_result_tuple in db_result_list:
        for result in db_result_tuple:  
            cpes.append(result)

    cpe_result = []
    for cpe in cpes:
        cpe = cpe.replace("*","All")
        current_product = []
        disassmbled_cpe = cpe.split(":")
        pvendor = disassmbled_cpe[3].replace("_"," ").title()
        pproduct = disassmbled_cpe[4].replace("_"," ").title()
        try:    
            pversion = disassmbled_cpe[5].title()
            if pversion == "All":
            	pversion = "- All versions"
            elif pversion == "-":
            	pversion = "- All versions"
        except Exception as e:
            handleException(e)
            pversion = ""
        product = pvendor+" "+pproduct+" "+pversion
        current_product.append(product)
        current_product.append(cpe.replace("All","*"))
        cpe_result.append(current_product)

    keyword_result = []
    for key in keywords:
        current_product = []
        current_product.append(key)
        current_product.append("N/A")
        keyword_result.append(current_product)
    return cpe_result+keyword_result
def showProducts(ptype_search):
    try:
        con = get_db_connection()
        cur = con.cursor()
        cur.execute("SELECT keyword FROM keyword_list")
        db_result_list = cur.fetchall()
        keywords = []
        for db_result_tuple in db_result_list:
            for result in db_result_tuple:  
                keywords.append(result)
        cur.execute("SELECT cpe FROM cpe_list")
        db_result_list = cur.fetchall()
        cpes = []
        for db_result_tuple in db_result_list:
            for result in db_result_tuple:  
                cpes.append(result)

        cpe_result = []
        for cpe in cpes:
            cpe = cpe.replace("*","All")
            current_product = []
            disassmbled_cpe = cpe.split(":")
            ptype = ""
            if disassmbled_cpe[2] == "o":
                ptype = "Operating System"
            elif disassmbled_cpe[2] == "a":
                ptype = "Application"
            else:
                ptype = "Hardware"

            try:    
                pvendor = disassmbled_cpe[3].replace("_"," ")
                pproduct = disassmbled_cpe[4].replace("_"," ")
                pversion = disassmbled_cpe[5]
            except Exception as e:
                handleException(e)
                pversion = "All"
                pvendor = "N/A"
                pproduct = "N/A"
            current_product.append(ptype)
            current_product.append(pvendor)
            current_product.append(pproduct)
            current_product.append(pversion)
            current_product.append(cpe.replace("All","*"))
            cpe_result.append(current_product)

        keyword_result = []
        for key in keywords:
            current_product = []
            current_product.append("N/A")
            current_product.append("N/A")
            current_product.append(key)
            current_product.append("N/A")
            current_product.append("N/A")
            keyword_result.append(current_product)

        if ptype_search == "All":
            result = ["OK",keyword_result+cpe_result]
            return result
        elif "CPE" in ptype_search:
            result = ["OK",cpe_result]
            return result
        elif "keyword" in ptype_search:
            result = ["OK",keyword_result]
            return result
        else:
            result = ["NOK",""]
            return result
    except Exception as e:
        handleException(e)
        result = ["NOK",""]
        return result   
def mailTester(sender, passwd, smtpsrv, port, tls, receivers):
    try:
        if tls == "True":
            tls = "yes"
        else:
            tls = "no"
        script_path = os.path.abspath(__file__)
        dir_path = script_path.replace("secmon_lib.py","")
        language = ""
        con = sqlite3.connect(dir_path+'secmon.db')
        cur = con.cursor()
        cur.execute("SELECT language FROM config")
        db_result_tuple = cur.fetchone()
        for value in db_result_tuple:
            language = value
            body = ""
        with open(dir_path+'rss_template.html', 'r') as template:
            html_code = template.read()
            html_code = html_code.replace("<strong>Source : </strong> $SOURCE <br><br>","")
            html_code = html_code.replace("Notification","SECMON")
            html_code = html_code.replace("<strong>Title : </strong> $TITLE <br><br>","")
        if language =="en" or language == "EN":
            html_code = html_code.replace("RSS module","Test email")
            html_code = html_code.replace("<strong>Summary :</strong> $SUMMARY <br>","SECMON is an active security watch program that alerts you to the latest vulnerabilities and threats.")
            html_code = html_code.replace("Responsive HTML email templates","SECMON - Test email.")
            html_code = html_code.replace("See more details","Follow me on Github")
            body = html_code.replace("URL OF THE NEWS","https://github.com/Guezone")

        elif language =="fr" or language == "FR":
            html_code = html_code.replace("RSS module","Mail de test")
            html_code = html_code.replace("<strong>Summary :</strong> $SUMMARY <br>","SECMON est un programme de veille sécurité active\net vous averti des dernières vulnérabilités et menaces.")
            html_code = html_code.replace("Responsive HTML email templates","SECMON - Mail de test.")
            html_code = html_code.replace("See more details","Suivez-moi sur Github")
            html_code = html_code.replace("This email was sent by", "Cet email a été envoyé par")
            body = html_code.replace("URL OF THE NEWS","https://github.com/Guezone")   
        for receiver in receivers:
            smtpserver = smtplib.SMTP(smtpsrv,port)
            msg = MIMEMultipart()
            if language =="en" or language == "EN":
                msg['Subject'] = 'SECMON - Test email.'
            elif language =="fr" or language == "FR":
                msg['Subject'] = 'SECMON - Mail de test.'
                msg['From'] = sender
                msg['To'] = receiver
                msg.attach(MIMEText(body, 'html'))
            if tls == "yes":
                smtpserver.ehlo()
                smtpserver.starttls()
                smtpserver.login(sender, passwd)
                smtpserver.sendmail(sender, receiver, msg.as_string())
            elif tls == "no":
                smtpserver.login(sender, passwd)
                smtpserver.sendmail(sender, receiver, msg.as_string())
        receiver_conf = ';'.join(receivers)
        con = get_db_connection()
        cur = con.cursor()
        cur.execute("UPDATE config SET sender = (?)", (sender,))
        cur.execute("UPDATE config SET password = (?)", ((str(base64.b64encode(passwd.encode("UTF-8"))).replace("b'","")).replace("'",""),))
        cur.execute("UPDATE config SET smtpsrv = (?)", (smtpsrv,))
        cur.execute("UPDATE config SET port = (?)", (port,))
        cur.execute("UPDATE config SET receiver = (?)", (receiver_conf,))
        cur.execute("UPDATE config SET tls = (?)", (tls,))
        con.commit()
        return True
    except Exception as e:
        handleException(e)
        return False

def get_db_connection():
    script_path = os.path.abspath(__file__)
    dir_path = script_path.replace("secmon_lib.py","")
    conn = sqlite3.connect(dir_path+'secmon.db')
    conn.row_factory = sqlite3.Row
    return conn

def removeHTMLtags(text):
    clean = re.compile('<.*?>')
    return re.sub(clean, '', text)
def changeCVEState(action, cve):
    
    db_change = ""
    if action == "ack":
        db_change = "Read"
    elif action == "trt":
        db_change = "Corrected"
    elif action == "false":
        db_change = "False positive"
    elif action == "utrt":
        db_change = "Unread"
    script_path = os.path.abspath(__file__)
    dir_path = script_path.replace("secmon_lib.py","")
    conn = sqlite3.connect(dir_path+'secmon.db')
    cur = conn.cursor()
    # Check if CVE exist
    cur.execute("SELECT CVE_ID FROM CVE_DATA WHERE CVE_ID = (?);", (cve.replace(" ",""),))
    db_result_list = cur.fetchall()
    db_result_str = ""
    for db_result_tuple in db_result_list:
        for result in db_result_tuple:
            db_result_str+=result
    if cve not in db_result_str:
        raise
    else:
        cur.execute("UPDATE CVE_DATA SET STATUS = (?) WHERE CVE_ID = (?)", (db_change,cve.replace(" ","")))
        conn.commit()
        return db_change

def getRegisteredCveInfos(cve,full):
    cve_infos = {}
    script_path = os.path.abspath(__file__)
    dir_path = script_path.replace("secmon_lib.py","")
    conn = get_db_connection()
    cur = conn.cursor()
    cve_av = "N/A"
    cve_ac = "N/A"
    cve_pr = "N/A"
    cve_ui = "N/A"
    cve_scope = "N/A"
    cve_confid = "N/A"
    cve_integrity = "N/A"
    cve_avail = "N/A"
    if full == True:
        nvd_base_url = "https://services.nvd.nist.gov/rest/json/cve/1.0/"   
        nvd_query = nvd_base_url+cve
        nvd_response = requests.get(url=nvd_query)
        try:
            nvd_data = nvd_response.json()
        except Exception as e:
            handleException(e)
            pass    
        try:
            cve_av = nvd_data['result']['CVE_Items'][0]['impact']['baseMetricV3']['cvssV3']['attackVector']
            cve_ac = nvd_data['result']['CVE_Items'][0]['impact']['baseMetricV3']['cvssV3']['attackComplexity']
            cve_pr = nvd_data['result']['CVE_Items'][0]['impact']['baseMetricV3']['cvssV3']['privilegesRequired']
            cve_ui = nvd_data['result']['CVE_Items'][0]['impact']['baseMetricV3']['cvssV3']['userInteraction']
            cve_scope = nvd_data['result']['CVE_Items'][0]['impact']['baseMetricV3']['cvssV3']['scope']
            cve_confid = nvd_data['result']['CVE_Items'][0]['impact']['baseMetricV3']['cvssV3']['confidentialityImpact']
            cve_integrity = nvd_data['result']['CVE_Items'][0]['impact']['baseMetricV3']['cvssV3']['integrityImpact']
            cve_avail = nvd_data['result']['CVE_Items'][0]['impact']['baseMetricV3']['cvssV3']['availabilityImpact']
        except Exception as e:
            pass
    req = cur.execute("SELECT CVE_ID, KEYWORD, STATUS,CVE_SCORE,CVE_DATE, CVE_DESCRIPTION, CVE_EVAL, CVE_CPE,CVE_SOURCES FROM CVE_DATA WHERE CVE_ID = (?);", (cve.replace(" ",""),))
    db_result = []
    for tup in req:
        for item in tup:
            db_result.append(item)
    try:
        cve_infos['cve_id'] = db_result[0]
        cve_infos['cve_description'] = db_result[5]
        cve_infos['cve_date'] = db_result[4]
        cve_infos['cve_score'] = db_result[3]
        cve_infos['cve_status'] = db_result[6]
        cve_infos['cve_cpe'] = db_result[7].split("\n")
        cve_infos['cve_sources'] = db_result[8].split("\n")
        cve_infos['cve_mgmt_status'] = db_result[2]
        cve_infos['cve_keymatch'] = db_result[1]
        cve_infos['cve_av'] = cve_av
        cve_infos['cve_ac'] = cve_ac
        cve_infos['cve_pr'] = cve_pr
        cve_infos['cve_ui'] = cve_ui
        cve_infos['cve_scope'] = cve_scope
        cve_infos['cve_confid'] = cve_confid
        cve_infos['cve_integrity'] = cve_integrity
        cve_infos['cve_avail'] = cve_avail
    except Exception as e:
        handleException(e)
        return None

    return cve_infos
def getUnregisteredCveInfos(cve):
    cve_infos = {}
    script_path = os.path.abspath(__file__)
    dir_path = script_path.replace("secmon_lib.py","")
    conn = sqlite3.connect(dir_path+'secmon.db')
    cur = conn.cursor()
    cur.execute("SELECT CVE_ID FROM CVE_DATA WHERE CVE_ID = (?);", (cve.replace(" ",""),))
    cve_infos['cve_id'] = cve
    cve_id = cve
    db_result_list = cur.fetchall()
    db_result_str = ""
    for db_result_tuple in db_result_list:
        for result in db_result_tuple:
            db_result_str+=result
    if cve not in db_result_str:
        cve_infos["cve_dbs"]="False"
    else:
        cve_infos["cve_dbs"]="True"

    nvd_base_url = "https://services.nvd.nist.gov/rest/json/cve/1.0/"
    nvd_query = nvd_base_url+cve_id
    nvd_response = requests.get(url=nvd_query)
    nvd_data = nvd_response.json()
    cve_av = "N/A"
    cve_ac = "N/A"
    cve_pr = "N/A"
    cve_ui = "N/A"
    cve_scope = "N/A"
    cve_confid = "N/A"
    cve_integrity = "N/A"
    cve_avail = "N/A"
    if 'result' in nvd_data.keys():
        cve_score = ""
        if not "impact" in nvd_data['result']['CVE_Items'][0].keys():
            cve_score = "N/A"
            try:    
                webpage = requests.get("https://nvd.nist.gov/vuln/detail/{}".format(cve_id))
                soup = BeautifulSoup(webpage.content, 'html.parser')
                cve_cna_score = soup.find(id='Cvss3CnaCalculatorAnchor')
                cve_cna_score = cve_cna_score.get_text()
                if cve_cna_score != "":
                    cve_score = cve_cna_score
                else:
                    cve_score = "N/A"
            except:
                    cve_score = "N/A"
        else:
            try:
                cve_score = nvd_data['result']['CVE_Items'][0]['impact']['baseMetricV3']['cvssV3']['baseScore']
                cve_av = nvd_data['result']['CVE_Items'][0]['impact']['baseMetricV3']['cvssV3']['attackVector']
                cve_ac = nvd_data['result']['CVE_Items'][0]['impact']['baseMetricV3']['cvssV3']['attackComplexity']
                cve_pr = nvd_data['result']['CVE_Items'][0]['impact']['baseMetricV3']['cvssV3']['privilegesRequired']
                cve_ui = nvd_data['result']['CVE_Items'][0]['impact']['baseMetricV3']['cvssV3']['userInteraction']
                cve_scope = nvd_data['result']['CVE_Items'][0]['impact']['baseMetricV3']['cvssV3']['scope']
                cve_confid = nvd_data['result']['CVE_Items'][0]['impact']['baseMetricV3']['cvssV3']['confidentialityImpact']
                cve_integrity = nvd_data['result']['CVE_Items'][0]['impact']['baseMetricV3']['cvssV3']['integrityImpact']
                cve_avail = nvd_data['result']['CVE_Items'][0]['impact']['baseMetricV3']['cvssV3']['availabilityImpact']
            except:
                cve_score = "N/A"
        weakness = ""
        try:
            weakness = nvd_data['result']['CVE_Items'][0]['cve']["problemtype"]["problemtype_data"][0]["description"][0]["value"]
        except:
            weakness = "N/A"
        published_date = nvd_data['result']['CVE_Items'][0]['publishedDate']
        cve_date = published_date.split("T")[0]
        fr_date = cve_date.split("-")
        cve_infos['cve_date'] = fr_date[2]+"/"+fr_date[1]+"/"+fr_date[0]
        cve_cpe = []
        if "configurations" in nvd_data['result']['CVE_Items'][0].keys():
            if nvd_data['result']['CVE_Items'][0]['configurations']['nodes']:
                for node in nvd_data['result']['CVE_Items'][0]['configurations']['nodes']:
                    if 'cpe_match' in node.keys():
                        for cpe_node in node['cpe_match']:
                            cve_cpe.append(cpe_node['cpe23Uri'])
        products = []
        if cve_cpe == []:
            products = ["N/A"]
        else:
            con = get_db_connection()
            cur = con.cursor()
            cur.execute("SELECT cpe FROM cpe_list")
            db_result_list = cur.fetchall()
            cpes = []
            for db_result_tuple in db_result_list:
                for result in db_result_tuple:  
                    cpes.append(result)
            for cpe in cve_cpe:
                cpe = cpe.replace("*","All")
                current_product = []
                disassmbled_cpe = cpe.split(":")
                ptype = ""
                if disassmbled_cpe[2] == "o":
                    ptype = "Operating System"
                elif disassmbled_cpe[2] == "a":
                    ptype = "Application"
                else:
                    ptype = "Hardware"
                pvendor = disassmbled_cpe[3].replace("_"," ")
                pproduct = disassmbled_cpe[4].replace("_"," ")
                try:    
                    pversion = disassmbled_cpe[5]
                except:
                    pversion = "All"
                current_product.append(ptype)
                current_product.append(pvendor)
                current_product.append(pproduct)
                current_product.append(pversion)
                current_product.append(cpe.replace("All","*"))
                match = ""
                if cpe.replace("All","*") in cpes:
                    match = "True - exact match"
                else:
                    for p in cpes:
                        partial_cpe = disassmbled_cpe[2]+":"+disassmbled_cpe[3]+":"+disassmbled_cpe[4]
                        if partial_cpe in p:
                            match = "True - partial match (require a manual check)"
                            break
                        else:
                            match = "False"
                current_product.append(match)
                products.append(current_product)




        cve_infos["cve_cpe"] = products
        published_date = nvd_data['result']['CVE_Items'][0]['publishedDate']
        formatted_date = published_date.split("T")[0]
        formatted_date = datetime.strptime(formatted_date,"%Y-%m-%d")
        date_2_days_ago = datetime.now() - timedelta(days=30)
        if (date_2_days_ago<=formatted_date) and (formatted_date<=datetime.now()):
            cve_date_status = "Recent"
        else:
            cve_date_status = "Old"
        if (cve_score=="N/A"):
            cve_status = cve_date_status+" - "+"Not yet rated (No score, no CPE)"
        else:
            cve_status = cve_date_status+" - "+"Valid evaluation"  
        cve_infos['cve_score'] = cve_score
        cve_infos['cve_av'] = cve_av
        cve_infos['cve_ac'] = cve_ac
        cve_infos['cve_pr'] = cve_pr
        cve_infos['cve_ui'] = cve_ui
        cve_infos['cve_scope'] = cve_scope
        cve_infos['cve_confid'] = cve_confid
        cve_infos['cve_integrity'] = cve_integrity
        cve_infos['cve_avail'] = cve_avail
        cve_infos['cve_status'] = cve_status  
        cve_infos['cve_weakness'] = weakness  
        exploits = []
        try:
            for e in CS.edbid_from_cve(cve):
                exploits.append(e)
        except:
            pass
        user,key = getGithubAPISettings()
        con = get_db_connection()
        cur = con.cursor()
        if user != "None" or key != "None":
            github_query = "https://api.github.com/search/repositories?q=exploit+"+cve
            github_response = requests.get(url=github_query,auth=(user,key))
            github_data = github_response.json()
            if "items" in github_data.keys():
                for i in github_data['items']:
                    exploits.append(i)
        if exploits == []:
            cve_infos['cve_expa'] = False
        else:
            cve_infos['cve_expa'] = True
        if 'reference_data' in nvd_data['result']['CVE_Items'][0]['cve']['references']:
            nvd_links = nvd_data['result']['CVE_Items'][0]['cve']['references']['reference_data']
            cve_sources = []
            for link in nvd_links:
                cve_sources.append(link['url'])
        else:
            cve_sources = []
        cve_infos['cve_sources'] = cve_sources
        cve_infos['cve_description'] = nvd_data['result']['CVE_Items'][0]['cve']['description']['description_data'][0]['value']
        
        req = cur.execute("SELECT STATUS FROM CVE_DATA WHERE CVE_ID = (?);", (cve,)).fetchone()
        if req == None:
            cve_infos['cve_mgmt_status'] = "Unavailable"
        else:
            cve_infos['cve_mgmt_status'] = ""
            for res in req:
                cve_infos['cve_mgmt_status'] += res
    else:
        raise ValueError('Failed to find CVE details on NVD API.')
    return cve_infos
def getCpeList():
    cur = get_db_connection().cursor()
    cur.execute("SELECT cpe FROM cpe_list")
    db_result_list = cur.fetchall()
    cpes = []
    for db_result_tuple in db_result_list:
        for result in db_result_tuple:    
            cpes.append(result)
    return cpes
def getKeywordsList():    
    cur = get_db_connection().cursor()
    cur.execute("SELECT keyword FROM keyword_list")
    db_result_list = cur.fetchall()
    keywords = []
    for db_result_tuple in db_result_list:
        for result in db_result_tuple:    
            keywords.append(result)
    return keywords
def getRegisteredCveStats():
    conn = get_db_connection()
    db_result = conn.execute('SELECT * FROM CVE_DATA',).fetchall()
    low_counter = 0
    medium_counter = 0
    high_counter = 0
    critical_counter = 0
    for cve in db_result:
        current_cve = []
        
        for data in cve:
            current_cve.append(data)
        current_cve.pop(0)
        if current_cve[3] != "N/A" and current_cve[3] != None:
            current_cve[3] = current_cve[3].split(" ")[0]
            if 0 < float(current_cve[3]) <= 3.9:
                low_counter +=1
            elif 4.0 < float(current_cve[3]) <= 6.9:
                medium_counter +=1
            elif 7.0 < float(current_cve[3]) <= 8.9:
                high_counter +=1
            elif 9.0 < float(current_cve[3]) <= 10:
                critical_counter +=1
    stats = [low_counter,medium_counter,high_counter,critical_counter]
    return stats

def getRegisteredCveStatus():
    conn = get_db_connection()
    db_result = conn.execute('SELECT * FROM CVE_DATA',).fetchall()
    read_counter = 0
    unread_counter = 0
    corrected_counter = 0
    falsep_counter = 0
    total = 0
    for cve in db_result:
        current_cve = []
        total += 1
        for data in cve:
            current_cve.append(data)
        current_cve.pop(0)
        if current_cve[2] != "N/A" and current_cve[2] != None:
            if current_cve[2] == "Read":
                read_counter +=1
            elif current_cve[2] == "Unread" or current_cve[2] == "Native" or current_cve[2] == "Found when the product was added.":
                unread_counter +=1
            elif current_cve[2] == "Corrected":
                corrected_counter +=1
            elif current_cve[2] == "False positive": 
                falsep_counter +=1
    date = datetime.now()
    y = date.strftime("%Y")
    months = []
    months.append(date.strftime("%B"))
    months.append((date + relativedelta(months=-1)).strftime("%B"))
    months.append((date + relativedelta(months=-2)).strftime("%B"))
    months.append((date + relativedelta(months=-3)).strftime("%B"))
    months.append((date + relativedelta(months=-4)).strftime("%B"))
    months.append((date + relativedelta(months=-5)).strftime("%B"))
    digit_months = []
    digit_months.append(date.strftime("%m/%Y"))
    digit_months.append((date + relativedelta(months=-1)).strftime("%m/%Y"))
    digit_months.append((date + relativedelta(months=-2)).strftime("%m/%Y"))
    digit_months.append((date + relativedelta(months=-3)).strftime("%m/%Y"))
    digit_months.append((date + relativedelta(months=-4)).strftime("%m/%Y"))
    digit_months.append((date + relativedelta(months=-5)).strftime("%m/%Y"))
    months_data = [0,0,0,0,0,0]
    for cve in db_result:
        current_cve = []
        for data in cve:
            current_cve.append(data)
        current_cve.pop(0)
        if current_cve[4] != "N/A" and current_cve[4] != None:
            if digit_months[0] in current_cve[4]:
                months_data[0] +=1
            elif digit_months[1] in current_cve[4]:
                months_data[1] +=1
            elif digit_months[2] in current_cve[4]:
                months_data[2] +=1
            elif digit_months[3] in current_cve[4]:
                months_data[3] +=1
            elif digit_months[4] in current_cve[4]:
                months_data[4] +=1
            elif digit_months[5] in current_cve[4]:
                months_data[5] +=1
    stats = [read_counter,unread_counter,corrected_counter,falsep_counter,total,[ele for ele in reversed(months)],[ele for ele in reversed(months_data)]]
    return stats


def getRegisteredCve():
    conn = get_db_connection()
    db_result = conn.execute('SELECT * FROM CVE_DATA',).fetchall()
    cvelist = []
    for cve in db_result:
        current_cve = []
        for data in cve:
            current_cve.append(data)
        current_cve.pop(0)
        if "cpe" in current_cve[1]:
            current_cve[1] = getParsedCpe(current_cve[1])
        if current_cve[2] == "Unread" or "Native" in current_cve[2] or "Found" in current_cve[2]:
            current_cve.append("#e74a3b")
            current_cve.append("white")
        elif current_cve[2] == "Read":
            current_cve.append("#faaa3c")
            current_cve.append("black")
        elif current_cve[2] == "False positive":
            current_cve.append("#6c757d")
            current_cve.append("white")
        else:
            current_cve.append("#1cc88a")
            current_cve.append("white")
        try:
            float(current_cve[3])
        except:
            if current_cve[3] != "N/A":
                current_cve[3] = current_cve[3].split(" ")[0]

        if current_cve[3] != "N/A":
            current_cve[3] = current_cve[3].split(" ")[0]
            if 0 < float(current_cve[3]) <= 3.9:
                current_cve.append("#36b9cc")
                current_cve.append("white")
            elif 3.9 < float(current_cve[3]) <= 6.9:
                current_cve.append("#faaa3c")
                current_cve.append("black")
            elif 6.9 < float(current_cve[3]) <= 8.9:
                current_cve.append("#e74a3b")
                current_cve.append("white")
            elif 8.9 < float(current_cve[3]) <= 10:
                current_cve.append("#202020")
                current_cve.append("white")
        else:
            current_cve.append("#6c757d")
            current_cve.append("white")
        cvelist.append(current_cve)
    conn.close()
    return cvelist

def getUnexploitableCveIdList():
    conn = get_db_connection()
    db_result = conn.execute('SELECT CVE_ID FROM CVE_DATA WHERE EXPLOIT_FIND = "False"',).fetchall()
    cvelist = []
    for cve in db_result:
        for t in cve:
            cvelist.append(t)
    conn.close()
    return cvelist

def getExploitableCveIdList():
    conn = get_db_connection()
    db_result = conn.execute('SELECT CVE_ID FROM CVE_DATA WHERE EXPLOIT_FIND = "True"',).fetchall()
    cvelist = []
    for cve in db_result:
        for t in cve:
            cvelist.append(t)
    conn.close()
    return cvelist
